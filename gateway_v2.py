import asyncio
import json
import logging
import os
import re
import sys
import time
import traceback
import uuid
import httpx

from datetime import datetime
from collections import defaultdict
from personality import GalacticPersonality
from model_manager import (TRANSIENT_ERRORS, PERMANENT_ERRORS,
                           ERROR_RATE_LIMIT, ERROR_TIMEOUT, ERROR_AUTH)
from spinner import spinner

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger("GalacticGateway")

# Silence noisy HTTP libraries
logging.getLogger("httpx").setLevel(logging.WARNING)
logging.getLogger("httpcore").setLevel(logging.WARNING)

# NVIDIA models where the streaming endpoint is broken or unreliable.
# These will be forced to non-streaming mode even when streaming is enabled.
# (e.g. Qwen 3.5 397B returns 200 on stream request but never sends SSE data)
_NVIDIA_NO_STREAM = {
    "qwen/qwen3.5-397b-a17b",
}

# OpenRouter models/routes where SSE streaming is unreliable (finish_reason=error,
# empty chunks). Force these to non-streaming to avoid noisy fallback warnings.
_OPENROUTER_NO_STREAM = {
    "google/gemini-3.1-pro-preview:nitro",
    "openai/gpt-5.2-pro",
    "qwen/qwen3-coder-next",
}

# NVIDIA models that require extra body params for thinking/reasoning.
# These are injected into the payload when the active model is on this list.
_NVIDIA_THINKING_MODELS = {
    "z-ai/glm5":              {"chat_template_kwargs": {"enable_thinking": True, "clear_thinking": False}},
    "moonshotai/kimi-k2.5":   {"chat_template_kwargs": {"thinking": True}},
    "qwen/qwen3.5-397b-a17b": {"chat_template_kwargs": {"enable_thinking": True}},
    "deepseek-ai/deepseek-v3.2": {"chat_template_kwargs": {"thinking": True}},
    "nvidia/nemotron-3-nano-30b-a3b": {
        "reasoning_budget": 16384,
        "chat_template_kwargs": {"enable_thinking": True},
    },
    "nvidia/nvidia-nemotron-nano-9b-v2": {
                "min_thinking_tokens": 1024,
                "max_thinking_tokens": 2048,
            },
        }
        
_PLANNING_PROMPT_TEMPLATE = """
Act as an expert Project Manager and Planner. Your goal is to break down a complex user request into a sequence of small, concrete, and actionable steps. Each step should be a specific instruction or a question to resolve before moving to the next. Focus on identifying information gaps, research needs, and logical progression.

Output a numbered list of steps, clearly outlining the plan. DO NOT execute anything yet; just provide the plan.

Example Complex Request:
"Research the best ways to train a dog for agility competitions and then write an email to a local dog trainer asking for a consultation, including some of the research findings."

Example Plan:
1. Research "best dog breeds for agility competitions" and "agility training techniques for beginners."
2. Summarize key findings regarding training methods and recommended breeds.
3. Find contact information for local dog trainers specializing in agility.
4. Draft an email to a chosen trainer, introducing the user, mentioning their dog's breed (if known), incorporating summarized research findings, and asking about consultation availability and costs.
5. Review and send the email.

User Request: {user_input}

Your Plan:
"""
        
# ── Token pricing (USD per 1M tokens) ────────────────────────────────
MODEL_PRICING = {
    # OpenRouter — Frontier
    "google/gemini-3.1-pro-preview":   {"input": 1.25,  "output": 10.00},
    "anthropic/claude-opus-4.6":       {"input": 15.00, "output": 75.00},    "openai/gpt-5.2":                  {"input": 2.50,  "output": 10.00},
    "openai/gpt-5.2-codex":            {"input": 2.50,  "output": 10.00},
    "x-ai/grok-4.1-fast":              {"input": 3.00,  "output": 15.00},
    "deepseek/deepseek-v3.2":          {"input": 0.27,  "output": 1.10},
    "qwen/qwen3.5-plus-02-15":         {"input": 0.30,  "output": 1.20},
    # OpenRouter — Strong
    "google/gemini-3-pro-preview":     {"input": 1.25,  "output": 5.00},
    "google/gemini-3-flash-preview":   {"input": 0.10,  "output": 0.40},
    "anthropic/claude-sonnet-4.6":     {"input": 3.00,  "output": 15.00},
    "anthropic/claude-opus-4.5":       {"input": 15.00, "output": 75.00},
    "openai/gpt-5.2-pro":             {"input": 2.50,  "output": 10.00},
    "openai/gpt-5.1":                  {"input": 2.00,  "output": 8.00},
    "openai/gpt-5.1-codex":            {"input": 2.00,  "output": 8.00},
    "qwen/qwen3.5-397b-a17b":          {"input": 0.40,  "output": 1.60},
    "qwen/qwen3-coder-next":           {"input": 0.30,  "output": 1.20},
    "moonshotai/kimi-k2.5":            {"input": 0.60,  "output": 2.40},
    "deepseek/deepseek-v3.2-speciale": {"input": 0.27,  "output": 1.10},
    "z-ai/glm-5":                      {"input": 0.50,  "output": 2.00},
    # OpenRouter — Fast
    "mistralai/mistral-large-2512":    {"input": 2.00,  "output": 6.00},
    "mistralai/devstral-2512":         {"input": 0.10,  "output": 0.30},
    "minimax/minimax-m2.5":            {"input": 0.15,  "output": 0.60},
    "perplexity/sonar-pro-search":     {"input": 3.00,  "output": 15.00},
    "nvidia/nemotron-3-nano-30b-a3b":  {"input": 0,     "output": 0},
    "stepfun/step-3.5-flash":          {"input": 0.02,  "output": 0.16},
    "openai/gpt-5.2-chat":             {"input": 2.50,  "output": 10.00},
    # Direct providers
    "claude-sonnet-4-20250514":        {"input": 3.00,  "output": 15.00},
    "gemini-2.5-flash":                {"input": 0.15,  "output": 0.60},
    "gpt-4o":                          {"input": 2.50,  "output": 10.00},
    "grok-3":                          {"input": 3.00,  "output": 15.00},
    "mistral-large-latest":            {"input": 2.00,  "output": 6.00},
    "deepseek-chat":                   {"input": 0.27,  "output": 1.10},
}
_PRICING_FALLBACK = {"input": 1.00, "output": 3.00}
FREE_PROVIDERS = {"nvidia", "cerebras", "groq", "huggingface", "ollama"}


class CostTracker:
    """Tracks per-request token costs, persists to JSONL, computes dashboard stats."""

    def __init__(self, logs_dir='./logs'):
        self.logs_dir = logs_dir
        self.log_file = os.path.join(logs_dir, 'cost_log.jsonl')
        os.makedirs(logs_dir, exist_ok=True)
        self.session_start = datetime.now().isoformat()
        self.session_cost = 0.0
        self.last_request_cost = 0.0
        self.entries = []  # in-memory cache of recent entries
        self._load_existing()
        self._prune_old()

    def _load_existing(self):
        """Load existing JSONL entries into memory."""
        if not os.path.exists(self.log_file):
            return
        try:
            with open(self.log_file, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if line:
                        try:
                            self.entries.append(json.loads(line))
                        except json.JSONDecodeError:
                            continue
        except Exception:
            pass

    def _prune_old(self):
        """Remove entries older than 90 days and rewrite the file."""
        from datetime import timedelta
        cutoff = (datetime.now() - timedelta(days=90)).isoformat()
        before = len(self.entries)
        self.entries = [e for e in self.entries if e.get('ts', '') >= cutoff]
        if len(self.entries) < before:
            self._rewrite_file()

    def _rewrite_file(self):
        """Rewrite the JSONL file from memory (after prune)."""
        try:
            with open(self.log_file, 'w', encoding='utf-8') as f:
                for entry in self.entries:
                    f.write(json.dumps(entry) + '\n')
        except Exception:
            pass

    def log_usage(self, model, provider, tokens_in, tokens_out, actual_cost=None):
        """Calculate cost, append to JSONL, update running totals.

        If actual_cost is provided (e.g. from OpenRouter's generation API),
        it overrides the local estimate for accurate tracking.
        """
        is_free = provider in FREE_PROVIDERS

        if actual_cost is not None:
            # Use the real cost reported by the provider
            total_cost = actual_cost
            cost_in = 0.0   # breakdown unavailable for actual costs
            cost_out = 0.0
        else:
            pricing = MODEL_PRICING.get(model, _PRICING_FALLBACK)
            if is_free:
                pricing = {"input": 0, "output": 0}
            cost_in = (tokens_in / 1_000_000) * pricing["input"]
            cost_out = (tokens_out / 1_000_000) * pricing["output"]
            total_cost = cost_in + cost_out

        entry = {
            "ts": datetime.now().isoformat(),
            "model": model,
            "provider": provider,
            "tin": tokens_in,
            "tout": tokens_out,
            "cost_in": round(cost_in, 6),
            "cost_out": round(cost_out, 6),
            "cost": round(total_cost, 6),
            "free": is_free,
            "actual": actual_cost is not None,
        }

        self.entries.append(entry)
        self.session_cost += total_cost
        self.last_request_cost = total_cost

        # Append to file
        try:
            with open(self.log_file, 'a', encoding='utf-8') as f:
                f.write(json.dumps(entry) + '\n')
        except Exception:
            pass

    def get_stats(self):
        """Compute dashboard statistics from in-memory entries."""
        from datetime import timedelta
        now = datetime.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        week_start = (now - timedelta(days=7)).isoformat()
        month_start = (now - timedelta(days=30)).isoformat()
        fourteen_days_ago = (now - timedelta(days=14)).isoformat()

        today_cost = 0.0
        week_cost = 0.0
        month_cost = 0.0
        month_messages = 0
        daily_map = defaultdict(lambda: {"cost": 0.0, "models": defaultdict(float)})
        model_map = defaultdict(lambda: {"cost": 0.0, "messages": 0, "tokens_in": 0, "tokens_out": 0})
        free_models = set()

        for e in self.entries:
            ts = e.get('ts', '')
            cost = e.get('cost', 0.0)
            model = e.get('model', 'unknown')
            is_free = e.get('free', False)

            if is_free:
                free_models.add(model)

            if ts >= today_start:
                today_cost += cost
            if ts >= week_start:
                week_cost += cost
            if ts >= month_start:
                month_cost += cost
                month_messages += 1

            # Daily series (last 14 days)
            if ts >= fourteen_days_ago:
                day = ts[:10]  # YYYY-MM-DD
                daily_map[day]["cost"] += cost
                short_model = model.split('/')[-1] if '/' in model else model
                daily_map[day]["models"][short_model] += cost

            # By-model aggregation (last 30 days)
            if ts >= month_start and not is_free:
                model_map[model]["cost"] += cost
                model_map[model]["messages"] += 1
                model_map[model]["tokens_in"] += e.get('tin', 0)
                model_map[model]["tokens_out"] += e.get('tout', 0)

        # Build daily series (sorted, last 14 days)
        daily_series = []
        for day in sorted(daily_map.keys()):
            d = daily_map[day]
            daily_series.append({
                "date": day,
                "cost": round(d["cost"], 4),
                "models": {k: round(v, 4) for k, v in d["models"].items()},
            })

        # Build by-model list (sorted by cost descending, top 8)
        by_model = []
        for model, stats in sorted(model_map.items(), key=lambda x: x[1]["cost"], reverse=True)[:8]:
            by_model.append({
                "model": model,
                "cost": round(stats["cost"], 4),
                "messages": stats["messages"],
                "tokens_in": stats["tokens_in"],
                "tokens_out": stats["tokens_out"],
            })

        avg_per_message = (month_cost / month_messages) if month_messages > 0 else 0.0

        return {
            "session_cost": round(self.session_cost, 4),
            "today_cost": round(today_cost, 4),
            "week_cost": round(week_cost, 4),
            "month_cost": round(month_cost, 4),
            "last_request_cost": round(self.last_request_cost, 6),
            "avg_per_message": round(avg_per_message, 4),
            "message_count_month": month_messages,
            "daily": daily_series,
            "by_model": by_model,
            "free_models_used": sorted(list(free_models)),
        }


class GalacticGateway:
    def __init__(self, core):
        self.core = core
        self.config = core.config.get('gateway', {})
        # Prefer models.primary_provider/model (canonical source of truth written by
        # ModelManager._save_config), fall back to legacy gateway.* fields, and only
        # use hardcoded defaults when the config has never been written at all.
        models_cfg = core.config.get('models', {})
        self.provider = (
            models_cfg.get('primary_provider')
            or self.config.get('provider')
            or 'google'
        )
        self.model = (
            models_cfg.get('primary_model')
            or self.config.get('model')
            or 'gemini-2.5-flash'
        )
        self.api_key = self.config.get('api_key', 'NONE')
        
        # Load Personality (dynamic: reads .md files, config, or Byte defaults)
        workspace = core.config.get('paths', {}).get('workspace', '')
        self.personality = GalacticPersonality(config=core.config, workspace=workspace)

        # Token tracking (for /status compatibility)
        self.total_tokens_in = 0
        self.total_tokens_out = 0
        self._last_usage = None  # Populated by provider methods with real API token counts
        self._last_generation_id = None  # OpenRouter generation ID for cost lookup

        # TTS voice file tracking — set by speak() when text_to_speech tool is invoked
        self.last_voice_file = None
        
        # LLM reference (for /status compatibility and model switching)
        from types import SimpleNamespace
        self.llm = SimpleNamespace(
            provider=self.provider,
            model=self.model,
            api_key=self.api_key
        )

        # Anti-spin: flag indicating an active speak() call is in progress
        self._speaking = False
        # Set of active speak() asyncio.Tasks for reliable global cancellation
        self._active_tasks = set()
        # Queued model switch: if user switches model during active speak(), apply after
        self._queued_switch = None
        # Lock to serialize sub-agent speak_isolated() calls (prevents concurrent state corruption)
        self._speak_lock = asyncio.Lock()
        
        # Tool Registry
        self.tools = {}
        self.register_tools()
        
        # Conversation History
        self.history = []

        # Resumable Workflows State
        logs_dir = core.config.get('paths', {}).get('logs', './logs')
        self.runs_dir = os.path.join(logs_dir, 'runs')
        os.makedirs(self.runs_dir, exist_ok=True)
        self.checkpoint_uuid = None
        self._tool_count_since_cp = 0
        self._consecutive_failures = 0
        self._recent_tools = []
        self._trace_sid = None
        self.active_plan = None

        # Persistent chat log (JSONL) — survives page refreshes
        self.history_file = os.path.join(logs_dir, 'chat_history.jsonl')
        os.makedirs(os.path.dirname(self.history_file), exist_ok=True)

    def _log_chat(self, role, content, source="web"):
        """Append a chat entry to the persistent JSONL log and update the 30-min hot buffer."""
        entry = {
            "ts": datetime.now().isoformat(),
            "role": role,
            "content": content[:2000],  # Cap stored content to prevent log bloat
            "source": source,
        }
        try:
            with open(self.history_file, 'a', encoding='utf-8') as f:
                f.write(json.dumps(entry) + '\n')
            
            # Update the Hot Buffer (last 30 mins)
            try:
                from hot_memory_buffer import update_hot_buffer
                update_hot_buffer()
            except ImportError:
                pass
        except Exception:
            pass

    def register_tools(self):
        """Registers available tools for the LLM."""
        self.tools = {
            "read_file": {
                "description": "Read the contents of a file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file."}
                    },
                    "required": ["path"]
                },
                "fn": self.tool_read_file
            },
            "write_file": {
                "description": "Write content to a file.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file."},
                        "content": {"type": "string", "description": "Content to write."}
                    },
                    "required": ["path", "content"]
                },
                "fn": self.tool_write_file
            },
            "schedule_task": {
                "description": "Schedule a reminder or task execution.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Name of the task."},
                        "delay_seconds": {"type": "number", "description": "Delay in seconds before execution."},
                        "message": {"type": "string", "description": "Message to display/log when task fires."}
                    },
                    "required": ["name", "delay_seconds", "message"]
                },
                "fn": self.tool_schedule_task
            },
            "list_tasks": {
                "description": "List all scheduled tasks.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
                "fn": self.tool_list_tasks
            },
            "web_search": {
                "description": "Search the web.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Search query."}
                    },
                    "required": ["query"]
                },
                "fn": self.tool_web_search
            },
            "edit_file": {
                "description": "Edit a file by replacing exact text (safer than write_file).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file."},
                        "old_text": {"type": "string", "description": "Exact text to find and replace."},
                        "new_text": {"type": "string", "description": "New text to replace with."}
                    },
                    "required": ["path", "old_text", "new_text"]
                },
                "fn": self.tool_edit_file
            },
            # ── NVIDIA FLUX image generation ──────────────────────────────
            "generate_image": {
                "description": "Generate an image using FLUX AI via NVIDIA. Returns the path to the saved PNG file. Models: 'black-forest-labs/flux.1-schnell' (fast, 4 steps) or 'black-forest-labs/flux.1-dev' (quality, 50 steps).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt": {"type": "string", "description": "Image description / prompt"},
                        "model": {"type": "string", "description": "FLUX model ID (default: flux.1-schnell)"},
                        "width": {"type": "integer", "description": "Image width in pixels (default: 1024)"},
                        "height": {"type": "integer", "description": "Image height in pixels (default: 1024)"},
                        "steps": {"type": "integer", "description": "Diffusion steps — schnell default 4, dev default 50"}
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image
            },

            # ── Stable Diffusion 3.5 image generation ─────────────────────────
            "generate_image_sd35": {
                "description": "Generate an image using Stable Diffusion 3.5 Large via NVIDIA NIM. Higher quality, different style than FLUX. Returns path to saved image.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":          {"type": "string",  "description": "Image description / prompt"},
                        "negative_prompt": {"type": "string",  "description": "Things to avoid in the image (optional)"},
                        "width":           {"type": "integer", "description": "Image width in pixels (default: 1024, max: 1536)"},
                        "height":          {"type": "integer", "description": "Image height in pixels (default: 1024, max: 1536)"},
                        "steps":           {"type": "integer", "description": "Diffusion steps (default: 40, range: 10-100)"},
                        "cfg_scale":       {"type": "number",  "description": "Guidance scale (default: 5.0, range: 1-20)"},
                        "seed":            {"type": "integer", "description": "Random seed (0 = random)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image_sd35
            },

            # ── Google Imagen image generation ────────────────────────────────
            "generate_image_imagen": {
                "description": "Generate an image using Google Imagen 4 via the Google Generative AI API. High-quality photorealistic images. Returns path to saved image.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":         {"type": "string",  "description": "Image description / prompt"},
                        "model":          {"type": "string",  "description": "Imagen model to use: imagen-4-ultra (best), imagen-4 (standard), imagen-4-fast (quick). Default: imagen-4"},
                        "aspect_ratio":   {"type": "string",  "description": "Aspect ratio: 1:1, 16:9, 9:16, 4:3, 3:4. Default: 1:1"},
                        "number_of_images": {"type": "integer", "description": "Number of images to generate (1-4, default: 1)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image_imagen
            },
            "generate_video": {
                "description": "Generate a short video clip using Google Veo AI. Returns the path to the saved MP4 file. Supports text-to-video with configurable duration, resolution, and aspect ratio.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":          {"type": "string",  "description": "Scene description for the video"},
                        "duration":        {"type": "integer", "description": "Video duration in seconds: 4, 6, or 8 (default: 8)"},
                        "aspect_ratio":    {"type": "string",  "description": "Aspect ratio: 16:9 or 9:16 (default: 16:9)"},
                        "resolution":      {"type": "string",  "description": "Resolution: 720p, 1080p, or 4k (default: 1080p)"},
                        "negative_prompt": {"type": "string",  "description": "Things to avoid in the video (optional)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_video
            },
            "generate_video_from_image": {
                "description": "Animate a still image into a short video clip using Google Veo. Takes an image (from Imagen, FLUX, or SD3.5) and turns it into motion video. Returns path to saved MP4.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":       {"type": "string",  "description": "Description of the motion/animation to apply"},
                        "image_path":   {"type": "string",  "description": "Path to the source image file"},
                        "duration":     {"type": "integer", "description": "Video duration in seconds: 4, 6, or 8 (default: 8)"},
                        "aspect_ratio": {"type": "string",  "description": "Aspect ratio: 16:9 or 9:16 (default: 16:9)"},
                    },
                    "required": ["prompt", "image_path"]
                },
                "fn": self.tool_generate_video_from_image
            },

            # ── File & system utilities ────────────────────────────────────────
            "list_dir": {
                "description": "List files and directories at a path with sizes, dates, and types. ALWAYS use absolute paths (e.g. 'C:/Users/name/folder' or 'F:/My Folder') — relative paths resolve to the server working directory and will return errors for user folders. If the result starts with [ERROR], report that error to the user verbatim.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":    {"type": "string",  "description": "ABSOLUTE directory path to list (e.g. 'F:/Galactic AI Media'). Relative paths resolve to server CWD, not user folders."},
                        "pattern": {"type": "string",  "description": "Optional glob pattern to filter, e.g. '*.py' or '*.mp4'"},
                        "recurse": {"type": "boolean", "description": "Recurse into subdirectories (default: false)"},
                    },
                    "required": []
                },
                "fn": self.tool_list_dir
            },
            "find_files": {
                "description": "Find files matching a name pattern recursively under a directory. Faster and safer than exec_shell find/dir.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":    {"type": "string", "description": "Root directory to search from (default: current working directory)"},
                        "pattern": {"type": "string", "description": "Glob pattern, e.g. '*.log', '**/*.py', 'config.*'"},
                        "limit":   {"type": "integer", "description": "Maximum results to return (default: 100)"},
                    },
                    "required": ["pattern"]
                },
                "fn": self.tool_find_files
            },
            "hash_file": {
                "description": "Compute SHA256 (default), MD5, or SHA1 checksum of a file. Useful for verifying downloads or detecting changes.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":      {"type": "string", "description": "Path to the file"},
                        "algorithm": {"type": "string", "description": "Hash algorithm: sha256 (default), md5, sha1"},
                    },
                    "required": ["path"]
                },
                "fn": self.tool_hash_file
            },
            "diff_files": {
                "description": "Show a unified diff between two text files, or between a file and a string. Great for reviewing changes before overwriting.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path_a":  {"type": "string", "description": "Path to first file"},
                        "path_b":  {"type": "string", "description": "Path to second file (if comparing two files)"},
                        "text_b":  {"type": "string", "description": "String content to compare against path_a (if not comparing two files)"},
                        "context": {"type": "integer", "description": "Lines of context around changes (default: 3)"},
                    },
                    "required": ["path_a"]
                },
                "fn": self.tool_diff_files
            },
            "zip_create": {
                "description": "Create a ZIP archive from a file or directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source":      {"type": "string", "description": "File or directory path to archive"},
                        "destination": {"type": "string", "description": "Output .zip file path (default: source + '.zip')"},
                    },
                    "required": ["source"]
                },
                "fn": self.tool_zip_create
            },
            "zip_extract": {
                "description": "Extract a ZIP archive to a directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source":      {"type": "string", "description": "Path to the .zip file"},
                        "destination": {"type": "string", "description": "Directory to extract into (default: same directory as zip)"},
                    },
                    "required": ["source"]
                },
                "fn": self.tool_zip_extract
            },
            "image_info": {
                "description": "Get metadata about an image file: dimensions, format, file size, color mode. Does NOT send the image to any AI — pure local metadata.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to image file"},
                    },
                    "required": ["path"]
                },
                "fn": self.tool_image_info
            },

            # ── Clipboard ─────────────────────────────────────────────────────
            "clipboard_get": {
                "description": "Read the current text content of the OS clipboard.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_clipboard_get
            },
            "clipboard_set": {
                "description": "Write text to the OS clipboard so the user can paste it anywhere.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text": {"type": "string", "description": "Text to put on the clipboard"},
                    },
                    "required": ["text"]
                },
                "fn": self.tool_clipboard_set
            },

            # ── Desktop notifications ─────────────────────────────────────────
            "notify": {
                "description": "Send a desktop notification (toast/balloon) to the user's screen. Works on Windows, macOS, and Linux.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title":   {"type": "string", "description": "Notification title"},
                        "message": {"type": "string", "description": "Notification body text"},
                        "sound":   {"type": "boolean", "description": "Play a sound (default: false)"},
                    },
                    "required": ["title", "message"]
                },
                "fn": self.tool_notify
            },

            # ── Window management ─────────────────────────────────────────────
            "window_list": {
                "description": "List all currently open application windows with their titles, process names, and window IDs.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_window_list
            },
            "window_focus": {
                "description": "Bring a window to the foreground and focus it by title substring or window ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {"type": "string", "description": "Partial window title to match (case-insensitive)"},
                        "hwnd":  {"type": "integer", "description": "Exact window handle/ID from window_list"},
                    },
                    "required": []
                },
                "fn": self.tool_window_focus
            },
            "window_resize": {
                "description": "Resize and/or move an application window by title or window ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title":  {"type": "string",  "description": "Partial window title to match"},
                        "hwnd":   {"type": "integer", "description": "Window handle from window_list"},
                        "x":      {"type": "integer", "description": "Left position (pixels from screen left)"},
                        "y":      {"type": "integer", "description": "Top position (pixels from screen top)"},
                        "width":  {"type": "integer", "description": "Window width in pixels"},
                        "height": {"type": "integer", "description": "Window height in pixels"},
                    },
                    "required": []
                },
                "fn": self.tool_window_resize
            },

            # ── HTTP / API ────────────────────────────────────────────────────
            "http_request": {
                "description": "Make a raw HTTP request (GET, POST, PUT, DELETE, PATCH) to any URL. Supports custom headers, JSON body, and form data. Great for calling REST APIs directly.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "method":  {"type": "string", "description": "HTTP method: GET, POST, PUT, DELETE, PATCH (default: GET)"},
                        "url":     {"type": "string", "description": "Full URL including protocol"},
                        "headers": {"type": "object", "description": "Request headers as key-value pairs"},
                        "json":    {"type": "object", "description": "JSON body (sets Content-Type: application/json automatically)"},
                        "data":    {"type": "string", "description": "Raw string body"},
                        "params":  {"type": "object", "description": "URL query parameters as key-value pairs"},
                        "timeout": {"type": "integer", "description": "Timeout in seconds (default: 30)"},
                    },
                    "required": ["url"]
                },
                "fn": self.tool_http_request
            },

            # ── QR code ───────────────────────────────────────────────────────
            "qr_generate": {
                "description": "Generate a QR code image from any text or URL. Saves to logs/ and returns the file path.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text":       {"type": "string",  "description": "Text or URL to encode in the QR code"},
                        "size":       {"type": "integer", "description": "Box size in pixels (default: 10)"},
                        "border":     {"type": "integer", "description": "Border width in boxes (default: 4)"},
                        "error_correction": {"type": "string", "description": "Error correction level: L, M, Q, H (default: M)"},
                    },
                    "required": ["text"]
                },
                "fn": self.tool_qr_generate
            },

            # ── Environment variables ─────────────────────────────────────────
            "env_get": {
                "description": "Read an environment variable value. Returns all env vars if no name specified (filtered list, excludes secrets).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Environment variable name (e.g. PATH, HOME). Omit to list all."},
                    },
                    "required": []
                },
                "fn": self.tool_env_get
            },
            "env_set": {
                "description": "Set an environment variable for the current process (affects subprocesses spawned from this session).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":  {"type": "string", "description": "Environment variable name"},
                        "value": {"type": "string", "description": "Value to set"},
                    },
                    "required": ["name", "value"]
                },
                "fn": self.tool_env_set
            },

            # ── System info ────────────────────────────────────────────────────
            "system_info": {
                "description": "Get detailed system information: CPU, RAM, disk usage, OS version, uptime, Python version, and running process count.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_system_info
            },
            "kill_process_by_name": {
                "description": "Kill all running processes matching a name or partial name (e.g. 'chrome', 'notepad'). More convenient than process_kill which needs an ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":  {"type": "string",  "description": "Process name or partial name to match (case-insensitive)"},
                        "force": {"type": "boolean", "description": "Force kill (SIGKILL/taskkill /F). Default: false (graceful SIGTERM)"},
                    },
                    "required": ["name"]
                },
                "fn": self.tool_kill_process_by_name
            },

            # ── Color picker ───────────────────────────────────────────────────
            "color_pick": {
                "description": "Sample the pixel color at exact desktop screen coordinates. Returns hex, RGB, and HSL values. Useful for UI automation color verification.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "x": {"type": "integer", "description": "X coordinate (pixels from left edge of screen)"},
                        "y": {"type": "integer", "description": "Y coordinate (pixels from top edge of screen)"},
                    },
                    "required": ["x", "y"]
                },
                "fn": self.tool_color_pick
            },

            # ── Text / data utilities ──────────────────────────────────────────
            "text_transform": {
                "description": "Transform text: convert case, encode/decode base64, URL-encode/decode, count words/lines/chars, reverse, strip, wrap, or extract regex matches.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text":      {"type": "string", "description": "Input text to transform"},
                        "operation": {"type": "string", "description": "Operation: upper, lower, title, snake_case, camel_case, base64_encode, base64_decode, url_encode, url_decode, reverse, count, strip, regex_extract, json_format, csv_to_json"},
                        "pattern":   {"type": "string", "description": "Regex pattern (for regex_extract operation)"},
                    },
                    "required": ["text", "operation"]
                },
                "fn": self.tool_text_transform
            },

            # ── New v0.9.2 tools ─────────────────────────────────────────
            "execute_python": {
                "description": "Execute Python code in a subprocess and return stdout/stderr. Use for data processing, calculations, CSV/JSON manipulation, or quick scripts. Timeout: 60s default.",
                "parameters": {"type": "object", "properties": {
                    "code": {"type": "string", "description": "Python code to execute"},
                    "timeout": {"type": "integer", "description": "Timeout seconds (default: 60, max: 300)"},
                }, "required": ["code"]},
                "fn": self.tool_execute_python
            },
            "wait": {
                "description": "Pause execution for a specified number of seconds. Use between actions that need settling time, or to wait before retrying something.",
                "parameters": {"type": "object", "properties": {
                    "seconds": {"type": "number", "description": "Seconds to wait (max: 300)"},
                }, "required": ["seconds"]},
                "fn": self.tool_wait
            },
            "send_telegram": {
                "description": "Send a proactive message to a Telegram chat (defaults to admin). Useful for alerts, task completion notifications, and automation reports.",
                "parameters": {"type": "object", "properties": {
                    "message": {"type": "string", "description": "Message text (Markdown supported)"},
                    "chat_id": {"type": "string", "description": "Chat ID (default: admin_chat_id from config)"},
                    "image_path": {"type": "string", "description": "Optional path to image to attach"},
                }, "required": ["message"]},
                "fn": self.tool_send_telegram
            },
            "read_pdf": {
                "description": "Extract text content from a PDF file. Returns plain text of all pages (or specific page range).",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to PDF file"},
                    "pages": {"type": "string", "description": "Page range: '1-5', '3', 'all' (default: all)"},
                }, "required": ["path"]},
                "fn": self.tool_read_pdf
            },
            "read_csv": {
                "description": "Read CSV file and return contents as JSON rows with headers. Great for data analysis.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to CSV file"},
                    "limit": {"type": "integer", "description": "Max rows to return (default: 200)"},
                    "delimiter": {"type": "string", "description": "Delimiter character (default: comma)"},
                }, "required": ["path"]},
                "fn": self.tool_read_csv
            },
            "write_csv": {
                "description": "Write JSON rows to a CSV file. Takes a list of dictionaries as rows.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Output CSV file path"},
                    "rows": {"type": "array", "description": "Array of {key: value} objects"},
                    "append": {"type": "boolean", "description": "Append to existing file (default: false)"},
                }, "required": ["path", "rows"]},
                "fn": self.tool_write_csv
            },
            "read_excel": {
                "description": "Read Excel file (.xlsx) and return contents as JSON rows. Requires openpyxl.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to .xlsx file"},
                    "sheet": {"type": "string", "description": "Sheet name (default: first sheet)"},
                    "limit": {"type": "integer", "description": "Max rows to return (default: 100)"},
                }, "required": ["path"]},
                "fn": self.tool_read_excel
            },
            "regex_search": {
                "description": "Search file contents using regex. Returns matching lines with file paths and line numbers. Faster and safer than exec_shell grep.",
                "parameters": {"type": "object", "properties": {
                    "pattern": {"type": "string", "description": "Regex pattern to search for"},
                    "path": {"type": "string", "description": "File or directory to search in"},
                    "file_pattern": {"type": "string", "description": "Glob to filter files, e.g. '*.py' (default: all)"},
                    "limit": {"type": "integer", "description": "Max results (default: 50)"},
                }, "required": ["pattern", "path"]},
                "fn": self.tool_regex_search
            },
            "image_resize": {
                "description": "Resize an image to specified dimensions. Supports PNG, JPEG, WebP, BMP.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to source image"},
                    "width": {"type": "integer", "description": "Target width in pixels"},
                    "height": {"type": "integer", "description": "Target height in pixels"},
                    "output_path": {"type": "string", "description": "Output path (default: adds _resized suffix)"},
                }, "required": ["path"]},
                "fn": self.tool_image_resize
            },
            "image_convert": {
                "description": "Convert image between formats (PNG, JPEG, WebP, BMP, TIFF).",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to source image"},
                    "format": {"type": "string", "description": "Target format: png, jpeg, webp, bmp"},
                    "output_path": {"type": "string", "description": "Output path (default: same name, new extension)"},
                    "quality": {"type": "integer", "description": "JPEG/WebP quality 1-100 (default: 85)"},
                }, "required": ["path", "format"]},
                "fn": self.tool_image_convert
            },
            "git_status": {
                "description": "Run 'git status' in a directory. Returns working tree status.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                }, "required": []},
                "fn": self.tool_git_status
            },
            "git_diff": {
                "description": "Run 'git diff' to show changes. Returns unified diff output.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "staged": {"type": "boolean", "description": "Show staged changes only (default: false)"},
                }, "required": []},
                "fn": self.tool_git_diff
            },
            "git_log": {
                "description": "Show recent git commit history.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "count": {"type": "integer", "description": "Number of commits (default: 10)"},
                }, "required": []},
                "fn": self.tool_git_log
            },
            "git_commit": {
                "description": "Stage files and create a git commit.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "message": {"type": "string", "description": "Commit message"},
                    "files": {"type": "array", "description": "Files to stage (default: all changed files)"},
                }, "required": ["message"]},
                "fn": self.tool_git_commit
            },
            # spawn_subagent, check_subagent  — Migrated to skills/core/subagent_manager.py
            # chrome_* tools (16)             — Migrated to skills/core/chrome_bridge.py
            # post_tweet, read_mentions, read_dms, post_reddit,
            # read_reddit_inbox, reply_reddit  — Migrated to skills/core/social_media.py

            "create_skill": {
                "description": (
                    "Create a new Galactic AI skill. Writes a .py file to skills/community/ "
                    "and loads it immediately. The skill must subclass GalacticSkill and "
                    "implement get_tools(). Use list_skills first to check what already exists. "
                    "CRITICAL: You MUST use 'from skills.base import GalacticSkill' at the top of the file."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":        {"type": "string", "description": "Skill name in snake_case (e.g. 'weather_lookup'). Used as the filename."},
                        "code":        {"type": "string", "description": "Full Python source code. MUST import from skills.base (NOT tools.base) and subclass GalacticSkill."},
                        "description": {"type": "string", "description": "One-line description of what this skill does."}
                    },
                    "required": ["name", "code", "description"]
                },
                "fn": self.tool_create_skill
            },
            "list_skills": {
                "description": "List all loaded skills with their metadata and tools. Shows both core and community skills.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                    "required": []
                },
                "fn": self.tool_list_skills
            },
            "remove_skill": {
                "description": (
                    "Remove a community skill by name. Core skills cannot be removed. "
                    "Unloads the skill and deletes its file from skills/community/."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "The skill_name to remove (e.g. 'weather_lookup')."}
                    },
                    "required": ["name"]
                },
                "fn": self.tool_remove_skill
            },
            "resume_workflow": {
                "description": "Resume an interrupted background workflow or task from a saved checkpoint.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "uuid": {"type": "string", "description": "The UUID of the checkpoint to load."}
                    },
                    "required": ["uuid"]
                },
                "fn": self.tool_resume_workflow
            },
        }

    def register_skill_tools(self, skills):
        """Merge tools from all loaded skills into self.tools.
        Called by GalacticCore.load_skills() after all skills are instantiated.
        Skill tools will OVERWRITE core gateway tools if names match, ensuring
        upgraded skill versions take priority.
        """
        count = 0
        overwritten = []
        for skill in skills:
            if not skill.enabled:
                continue
            skill_tools = skill.get_tools()
            for tool_name, tool_def in skill_tools.items():
                if tool_name in self.tools:
                    overwritten.append(tool_name)
                self.tools[tool_name] = tool_def
                count += 1
        
        if count:
            print(f"[Skills] Registered {count} tool(s) from skills.")
        if overwritten:
            print(f"[Skills] Upgraded core tools: {', '.join(set(overwritten))}")

    # --- Tool Implementations ---
    async def tool_resume_workflow(self, args):
        """Tool handler to restore state from a checkpoint and continue."""
        uuid_val = args.get('uuid')
        if not uuid_val:
            return "Error: 'uuid' is required."
            
        try:
            state = await self.load_checkpoint(uuid_val)
            # The actual resuming is tricky to do fully inside a tool call since it interrupts the current flow, 
            # but setting the state means the NEXT turn uses it. 
            # We will return a system prompt instructing the model to proceed.
            return f"[OK] Restored checkpoint {uuid_val}. State restored: turn {state.get('turn_count')}. You MUST now continue the interrupted task."
        except Exception as e:
            return f"[Error] Failed to resume workflow: {str(e)}"

    async def tool_read_file(self, args):
        path = args.get('path')
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            return f"Error reading file: {e}"

    # Core files that should never be overwritten by the AI agent
    _PROTECTED_FILES = {
        'gateway_v2.py', 'galactic_core_v2.py', 'web_deck.py', 'model_manager.py',
        'remote_access.py', 'personality.py', 'memory_module_v2.py', 'scheduler.py',
        'nvidia_gateway.py', 'splash.py', 'telegram_bridge.py', 'discord_bridge.py',
        'whatsapp_bridge.py', 'gmail_bridge.py', 'imprint_engine.py', 'ollama_manager.py',
        'requirements.txt', 'config.yaml', 'personality.yaml',
        'install.ps1', 'install.sh', 'update.ps1', 'update.sh',
        'launch.ps1', 'launch.sh', '.gitignore', 'LICENSE',
    }

    async def tool_write_file(self, args):
        path = args.get('path')
        content = args.get('content')
        try:
            # Guard: prevent overwriting core system files
            filename = os.path.basename(path)
            if filename in self._PROTECTED_FILES:
                return (
                    f"[BLOCKED] Cannot overwrite protected core file '{filename}'. "
                    f"Create a new file with a different name instead."
                )
            with open(path, 'w', encoding='utf-8') as f:
                f.write(content)
            return f"Successfully wrote to {path}"
        except Exception as e:
            return f"Error writing file: {e}"

    async def tool_web_search(self, args):
        """Web search using DuckDuckGo — returns parsed, ranked results (no API key needed)."""
        query = args.get('query', '')
        if not query:
            return "[ERROR] No search query provided."
        try:
            import urllib.parse
            from bs4 import BeautifulSoup

            encoded_q = urllib.parse.quote_plus(query)
            search_url = f"https://duckduckgo.com/html/?q={encoded_q}"

            async with httpx.AsyncClient(
                timeout=15.0,
                follow_redirects=True,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept-Language': 'en-US,en;q=0.9',
                }
            ) as client:
                response = await client.get(search_url)

            soup = BeautifulSoup(response.text, 'html.parser')
            results = []

            for result in soup.select('.result__body, .result')[:10]:
                title_el  = result.select_one('.result__title, .result__a')
                snippet_el = result.select_one('.result__snippet')
                url_el    = result.select_one('.result__url')

                title   = title_el.get_text(strip=True)   if title_el   else ''
                snippet = snippet_el.get_text(strip=True) if snippet_el else ''
                url     = url_el.get_text(strip=True)     if url_el     else ''

                if title and (snippet or url):
                    results.append({"title": title, "snippet": snippet, "url": url})

            if not results:
                return f"No results found for: '{query}'. Try rephrasing or use web_fetch on a specific URL."

            lines = [f"🔍 Web results for **'{query}'**:\n"]
            for i, r in enumerate(results[:8], 1):
                lines.append(f"{i}. **{r['title']}**")
                if r['snippet']:
                    lines.append(f"   {r['snippet']}")
                if r['url']:
                    lines.append(f"   🔗 {r['url']}")
                lines.append("")

            return "\n".join(lines)

        except ImportError:
            # bs4 not available: fall back to raw fetch
            return f"[Web Search] Query: {query} — Install beautifulsoup4 for parsed results."
        except Exception as e:
            return f"Web search error: {e}"
    
    async def tool_open_browser(self, args):
        """Open a URL in Playwright browser."""
        url = args.get('url')
        try:
            browser_plugin = next(
                (p for p in self.core.plugins
                 if "BrowserExecutorPro" in p.__class__.__name__
                 or getattr(p, 'skill_name', '') == 'browser_pro'),
                None
            )
            if not browser_plugin:
                return "[ERROR] BrowserExecutorPro plugin not loaded."
            
            result = await browser_plugin.navigate(url)
            
            if result['status'] == 'success':
                return f"[BROWSER] Navigated to: {url}"
            else:
                return f"[ERROR] Navigation failed: {result.get('message', 'Unknown error')}"
        except Exception as e:
            return f"[ERROR] Browser navigation: {e}"
    
    async def tool_screenshot(self, args):
        """Take a screenshot of the current browser page."""
        path = args.get('path')
        try:
            browser_plugin = next(
                (p for p in self.core.plugins
                 if "BrowserExecutorPro" in p.__class__.__name__
                 or getattr(p, 'skill_name', '') == 'browser_pro'),
                None
            )
            if not browser_plugin:
                return "[ERROR] BrowserExecutorPro plugin not loaded."
            result = await browser_plugin.screenshot(path=path, full_page=True)
            if result['status'] == 'success':
                return f"[BROWSER] Screenshot saved: {result['path']}"
            else:
                return f"[ERROR] Screenshot failed: {result.get('message', 'Unknown error')}"
        except Exception as e:
            return f"[ERROR] Browser screenshot: {e}"

    # ── Skills meta-tools ──────────────────────────────────────────────────

    async def tool_list_skills(self, args):
        """List all loaded skills and their tools."""
        if not self.core.skills:
            return "No skills loaded. Core skills are still running as legacy plugins during migration."
        lines = []
        for skill in self.core.skills:
            tool_names = list(skill.get_tools().keys())
            core_tag = " [core]" if skill.is_core else " [community]"
            enabled_tag = "" if skill.enabled else " (DISABLED)"
            lines.append(
                f"{skill.icon} **{skill.skill_name}** v{skill.version}{core_tag}{enabled_tag}\n"
                f"   {skill.description}\n"
                f"   Tools ({len(tool_names)}): {', '.join(tool_names) if tool_names else '(none)'}"
            )
        return "\n\n".join(lines)

    async def tool_create_skill(self, args):
        """Create a new community skill at runtime."""
        import importlib
        import ast as _ast

        name = args.get('name', '').strip()
        code = args.get('code', '')
        desc = args.get('description', '')

        if not name or not code:
            return "[ERROR] Both 'name' and 'code' are required."

        # Validate name is safe for use as a Python module name
        if not all(c.isalnum() or c == '_' for c in name) or name[0].isdigit():
            return "[ERROR] Skill name must be snake_case (letters, digits, underscores; cannot start with a digit)."

        # Validate code contains required elements
        if 'GalacticSkill' not in code:
            return "[ERROR] Code must contain a class that inherits from GalacticSkill."
        if 'get_tools' not in code:
            return "[ERROR] Skill class must implement get_tools()."

        # Find the skill class name via AST parsing
        skill_class_names = []
        try:
            tree = _ast.parse(code)
            for node in _ast.walk(tree):
                if isinstance(node, _ast.ClassDef):
                    for base in node.bases:
                        base_name = ''
                        if isinstance(base, _ast.Name):
                            base_name = base.id
                        elif isinstance(base, _ast.Attribute):
                            base_name = base.attr
                        if base_name == 'GalacticSkill':
                            skill_class_names.append(node.name)
                            break
        except SyntaxError as e:
            return f"[ERROR] Syntax error in skill code: {e}"

        if not skill_class_names:
            return "[ERROR] Could not find a class inheriting from GalacticSkill in the provided code."
        if len(skill_class_names) > 1:
            return f"[ERROR] Found multiple GalacticSkill subclasses: {', '.join(skill_class_names)}. Only one is allowed per skill file."
        skill_class_name = skill_class_names[0]

        # Check for duplicate skill names
        for existing in self.core.skills:
            if existing.skill_name == name:
                return f"[ERROR] Skill '{name}' already loaded. Use remove_skill first to replace it."

        # Write to community/
        skills_dir = os.path.join(os.path.dirname(os.path.abspath(self.core.config_path)), 'skills', 'community')
        os.makedirs(skills_dir, exist_ok=True)
        skill_path = os.path.join(skills_dir, f'{name}.py')

        try:
            with open(skill_path, 'w', encoding='utf-8') as f:
                f.write(code)
        except Exception as e:
            return f"[ERROR] Failed to write skill file: {e}"

        # Dynamic import
        try:
            module_name = f'skills.community.{name}'
            if module_name in sys.modules:
                del sys.modules[module_name]

            mod = importlib.import_module(module_name)
            cls = getattr(mod, skill_class_name)
            skill = cls(self.core)
            skill.is_core = False

            await skill.on_load()
            self.core.skills.append(skill)

            # Register its tools
            new_tools = skill.get_tools()
            registered = []
            for tool_name, tool_def in new_tools.items():
                if tool_name not in self.tools:
                    self.tools[tool_name] = tool_def
                    registered.append(tool_name)

            asyncio.create_task(skill.run())

            # Update registry
            from datetime import datetime as _dt
            registry = self.core._read_registry()
            if not registry.get('installed') and any(not s.is_core for s in self.core.skills if s.skill_name != name):
                await self.core.log("[Skills] Warning: registry.json was empty/missing — existing community skill entries may have been lost. Check skills/registry.json.", priority=1)
            registry['installed'].append({
                'module': name,
                'class': skill_class_name,
                'file': f'{name}.py',
                'installed_at': _dt.now().isoformat(),
                'source': 'ai_authored',
                'description': desc
            })
            self.core._write_registry(registry)

            return (
                f"[OK] Skill '{name}' created and loaded.\n"
                f"  Class: {skill_class_name}\n"
                f"  Tools registered: {', '.join(registered) if registered else '(none)'}\n"
                f"  File: {skill_path}"
            )

        except Exception as e:
            try:
                os.remove(skill_path)
            except OSError:
                pass
            sys.modules.pop(f'skills.community.{name}', None)
            return f"[ERROR] Failed to load skill '{name}': {e}"

    async def tool_remove_skill(self, args):
        """Remove a community skill by name."""
        name = args.get('name', '').strip()
        if not name:
            return "[ERROR] 'name' is required."

        target = next((s for s in self.core.skills if s.skill_name == name), None)
        if not target:
            return f"[ERROR] Skill '{name}' not found. Use list_skills to see loaded skills."

        if target.is_core:
            return f"[ERROR] '{name}' is a core skill and cannot be removed."

        try:
            await target.on_unload()
        except Exception:
            pass

        # Unregister tools
        tool_names = list(target.get_tools().keys())
        for tn in tool_names:
            self.tools.pop(tn, None)

        self.core.skills.remove(target)

        # Delete file
        skills_dir = os.path.join(os.path.dirname(os.path.abspath(self.core.config_path)), 'skills', 'community')
        skill_path = os.path.join(skills_dir, f'{name}.py')
        try:
            os.remove(skill_path)
        except OSError:
            pass

        # Update registry
        registry = self.core._read_registry()
        if not registry.get('installed') and any(not s.is_core for s in self.core.skills if s.skill_name != name):
            await self.core.log("[Skills] Warning: registry.json was empty/missing — existing community skill entries may have been lost. Check skills/registry.json.", priority=1)
        registry['installed'] = [e for e in registry['installed'] if e.get('module') != name]
        self.core._write_registry(registry)

        return f"[OK] Skill '{name}' removed. Tools unregistered: {', '.join(tool_names) if tool_names else '(none)'}"

    # ChromeBridge helpers & handlers    — Migrated to skills/core/chrome_bridge.py
    # Social Media handlers              — Migrated to skills/core/social_media.py

    async def tool_generate_image(self, args):
        """Generate an image using FLUX via NVIDIA's GenAI API."""
        import base64 as _b64, time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_image requires a 'prompt' argument."
        model = args.get('model', 'black-forest-labs/flux.1-schnell')
        # Strip nvidia/ prefix if user passed the alias path
        if model.startswith('nvidia/'):
            model = model[len('nvidia/'):]
        width = int(args.get('width', 1024))
        height = int(args.get('height', 1024))
        is_schnell = 'schnell' in model
        steps = int(args.get('steps', 4 if is_schnell else 50))

        # FLUX schnell and dev each have their own API key.
        # fluxDevApiKey → flux.1-dev, fluxApiKey → flux.1-schnell, apiKey → fallback.
        nvidia_cfg = self.core.config.get('providers', {}).get('nvidia', {})
        if not is_schnell:
            nvidia_key = (
                nvidia_cfg.get('fluxDevApiKey') or
                nvidia_cfg.get('fluxApiKey') or
                nvidia_cfg.get('apiKey') or ''
            )
        else:
            nvidia_key = (
                nvidia_cfg.get('fluxApiKey') or
                nvidia_cfg.get('apiKey') or ''
            )
        if not nvidia_key:
            return "[ERROR] NVIDIA FLUX key not found — add providers.nvidia.fluxApiKey (schnell) or fluxDevApiKey (dev) to config.yaml"

        url = f"https://ai.api.nvidia.com/v1/genai/{model}"
        headers = {
            "Authorization": f"Bearer {nvidia_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }
        # Build payload — schnell doesn't support cfg_scale or mode fields
        payload = {"prompt": prompt, "width": width, "height": height, "seed": 0, "steps": steps}
        if not is_schnell:
            payload["mode"] = "base"
            payload["cfg_scale"] = 5  # dev default per NVIDIA docs (1-9 range)

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                r = await client.post(url, headers=headers, json=payload)
                if r.status_code == 401:
                    return f"[ERROR] NVIDIA GenAI 401 Unauthorized — key used: nvapi-...{nvidia_key[-8:]}. Check that your NVIDIA API key has access to the FLUX model at ai.api.nvidia.com."
                if r.status_code == 500:
                    return f"[ERROR] NVIDIA GenAI HTTP 500 — their inference server is down right now. Do NOT retry. Report this to the user and suggest trying again in a few minutes or switching to flux.1-schnell."
                if r.status_code != 200:
                    return f"[ERROR] NVIDIA GenAI HTTP {r.status_code}: {r.text[:500]}"
                data = r.json()

            artifact = data.get('artifacts', [{}])[0]
            finish = artifact.get('finishReason', '')
            if finish == 'CONTENT_FILTERED':
                return "⚠️ Image generation blocked by content filter. Try a different prompt."
            b64 = artifact.get('base64', '')
            if not b64:
                return f"[ERROR] Image generation failed: {json.dumps(data)}"

            # API returns JPEG data
            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'flux')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"flux_{int(_time.time())}.jpg"
            path = os.path.join(img_subdir, fname)
            with open(path, 'wb') as f:
                f.write(_b64.b64decode(b64))
            # Signal Telegram bridge and web deck to deliver the image directly
            self.last_image_file = path
            return f"✅ Image generated and saved to: {path}\nModel: {model}\nPrompt: {prompt}"
        except Exception as e:
            return f"[ERROR] generate_image: {str(e)}"

    async def tool_schedule_task(self, args):
        """Schedule a task/reminder using the scheduler plugin."""
        name = args.get('name')
        delay_seconds = args.get('delay_seconds')
        message = args.get('message')
        
        try:
            # Check if scheduler plugin is available
            scheduler_plugin = next((p for p in self.core.plugins if "Scheduler" in p.__class__.__name__), None)
            if scheduler_plugin:
                await scheduler_plugin.schedule_task(name, delay_seconds, message)
                return f"Task '{name}' scheduled to fire in {delay_seconds} seconds."
            else:
                return "Scheduler plugin not available. Task not scheduled."
        except Exception as e:
            return f"Error scheduling task: {e}"
    
    async def tool_list_tasks(self, args):
        """List all scheduled tasks."""
        try:
            scheduler_plugin = next((p for p in self.core.plugins if "Scheduler" in p.__class__.__name__), None)
            if scheduler_plugin:
                tasks = await scheduler_plugin.list_tasks()
                if tasks:
                    return json.dumps(tasks, indent=2)
                else:
                    return "No scheduled tasks."
            else:
                return "Scheduler plugin not available."
        except Exception as e:
            return f"Error listing tasks: {e}"
    
    async def tool_edit_file(self, args):
        """Edit a file by finding and replacing exact text."""
        path = args.get('path')

        # Normalize alternative parameter formats LLMs sometimes use
        old_text = args.get('old_text')
        if old_text is None:
            old_text = args.get('old')
        new_text = args.get('new_text')
        if new_text is None:
            new_text = args.get('new')

        # Handle replacements array: {replacements: [{old/old_text, new/new_text}]}
        if old_text is None and 'replacements' in args:
            replacements = args['replacements']
            if isinstance(replacements, list) and len(replacements) > 0:
                first = replacements[0]
                old_text = first.get('old_text') or first.get('old')
                new_text = first.get('new_text') or first.get('new')

        if not path:
            return "Error: 'path' parameter is required."
        if old_text is None or new_text is None:
            return ("Error: 'old_text' and 'new_text' parameters are required. "
                    "Accepted formats: {old_text, new_text} or {old, new} or "
                    "{replacements: [{old, new}]}.")

        try:
            # Read current content
            with open(path, 'r', encoding='utf-8') as f:
                content = f.read()

            # Check if old_text exists
            if old_text not in content:
                return f"Error: Could not find exact text in {path}. No changes made."

            # Count occurrences
            count = content.count(old_text)
            if count > 1:
                return f"Warning: Found {count} occurrences of text. Please be more specific. No changes made."

            # Replace
            new_content = content.replace(old_text, new_text, 1)

            # Write back
            with open(path, 'w', encoding='utf-8') as f:
                f.write(new_content)

            return f"[OK] Successfully edited {path} (replaced 1 occurrence)"
        except Exception as e:
            return f"Error editing file: {e}"
    
    async def tool_web_fetch(self, args):
        """Fetch and extract readable content from a URL."""
        url = args.get('url')
        mode = args.get('mode', 'markdown')
        
        try:
            import httpx
            from bs4 import BeautifulSoup
            
            async with httpx.AsyncClient(follow_redirects=True, timeout=30.0, verify=False) as client:
                response = await client.get(url, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                })
                response.raise_for_status()
                
                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Remove script and style elements
                for script in soup(["script", "style", "nav", "footer", "header"]):
                    script.decompose()
                
                # Get text
                if mode == 'text':
                    text = soup.get_text(separator='\n', strip=True)
                else:  # markdown mode
                    # Basic markdown conversion
                    title = soup.find('title')
                    title_text = f"# {title.string}\n\n" if title else ""
                    
                    body = soup.find('body') or soup
                    paragraphs = body.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'li'])
                    
                    text_parts = [title_text]
                    for p in paragraphs:
                        tag_name = p.name
                        text_content = p.get_text(strip=True)
                        
                        if tag_name == 'h1':
                            text_parts.append(f"\n# {text_content}\n")
                        elif tag_name == 'h2':
                            text_parts.append(f"\n## {text_content}\n")
                        elif tag_name == 'h3':
                            text_parts.append(f"\n### {text_content}\n")
                        elif tag_name == 'li':
                            text_parts.append(f"- {text_content}")
                        else:
                            text_parts.append(text_content)
                    
                    text = '\n'.join(text_parts)
                
                # Limit length
                max_chars = 8000
                if len(text) > max_chars:
                    text = text[:max_chars] + "\n\n[... content truncated]"
                
                return f"[DOC] Content from {url}:\n\n{text}"
        except Exception as e:
            return f"Error fetching URL: {e}"
    
    async def tool_process_start(self, args):
        """Start a background process."""
        command = args.get('command')
        session_id = args.get('session_id', f"proc_{int(asyncio.get_event_loop().time())}")
        
        try:
            # Store processes in core if not exists
            if not hasattr(self.core, 'processes'):
                self.core.processes = {}
            
            # Start process
            process = await asyncio.create_subprocess_shell(
                command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            self.core.processes[session_id] = {
                'process': process,
                'command': command,
                'started': asyncio.get_event_loop().time(),
                'stdout': [],
                'stderr': []
            }
            
            # Start log collection task
            asyncio.create_task(self._collect_process_output(session_id))
            
            return f"[RUN] Process started: {session_id}\nCommand: {command}\nPID: {process.pid}"
        except Exception as e:
            return f"Error starting process: {e}"
    
    async def _collect_process_output(self, session_id):
        """Collect output from a running process."""
        try:
            proc_info = self.core.processes.get(session_id)
            if not proc_info:
                return
            
            process = proc_info['process']
            
            # Read stdout
            if process.stdout:
                async for line in process.stdout:
                    proc_info['stdout'].append(line.decode())
            
            # Wait for completion
            await process.wait()
            proc_info['exit_code'] = process.returncode
            proc_info['finished'] = asyncio.get_event_loop().time()
            
        except Exception as e:
            await self.core.log(f"Process output collection error: {e}", priority=1)
    
    async def tool_process_status(self, args):
        """Check status of a background process."""
        session_id = args.get('session_id')
        
        try:
            if not hasattr(self.core, 'processes') or session_id not in self.core.processes:
                return f"[ERR] Process not found: {session_id}"
            
            proc_info = self.core.processes[session_id]
            process = proc_info['process']
            
            status = "running" if process.returncode is None else f"exited ({process.returncode})"
            runtime = asyncio.get_event_loop().time() - proc_info['started']
            
            stdout_preview = ''.join(proc_info['stdout'][-10:])[:500]
            
            return (
                f"[STATUS] Process Status: {session_id}\n"
                f"Command: {proc_info['command']}\n"
                f"PID: {process.pid}\n"
                f"Status: {status}\n"
                f"Runtime: {runtime:.1f}s\n"
                f"Recent output:\n{stdout_preview}"
            )
        except Exception as e:
            return f"Error checking process: {e}"
    
    async def tool_process_kill(self, args):
        """Kill a background process."""
        session_id = args.get('session_id')
        
        try:
            if not hasattr(self.core, 'processes') or session_id not in self.core.processes:
                return f"[ERR] Process not found: {session_id}"
            
            proc_info = self.core.processes[session_id]
            process = proc_info['process']
            
            if process.returncode is None:
                process.kill()
                await process.wait()
                return f"[KILL] Process killed: {session_id}"
            else:
                return f"Process already exited: {session_id} (code {process.returncode})"
        except Exception as e:
            return f"Error killing process: {e}"
    
    async def tool_analyze_image(self, args):
        """Analyze an image — routes to the active provider's vision endpoint."""
        path = args.get('path')
        prompt = args.get('prompt', 'Describe this image in detail. Include any text you see.')

        import base64
        from pathlib import Path

        if not path or not Path(path).exists():
            return f"[ERR] Image not found: {path}"

        # Read file and detect MIME type (no hardcoded JPEG)
        with open(path, 'rb') as f:
            raw = f.read()
        image_b64 = base64.b64encode(raw).decode('utf-8')
        suffix = Path(path).suffix.lower()
        mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                    '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp'}
        mime_type = mime_map.get(suffix, 'image/jpeg')

        return await self._analyze_image_b64(image_b64, mime_type, prompt)

    async def _analyze_image_gemini(self, path, prompt):
        """Analyze image using Google Gemini Vision."""
        import base64
        from pathlib import Path
        try:
            with open(path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')

            suffix = Path(path).suffix.lower()
            mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                        '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp'}
            mime_type = mime_map.get(suffix, 'image/jpeg')

            api_key = self.config.get('api_key') or self.core.config.get('providers', {}).get('google', {}).get('apiKey')
            if not api_key:
                return "[ERR] Google API key not configured for image analysis."

            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash-exp:generateContent?key={api_key}"
            payload = {"contents": [{"parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": mime_type, "data": image_data}}
            ]}]}

            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(url, json=payload)
                data = response.json()
                if 'candidates' in data and data['candidates']:
                    result = data['candidates'][0]['content']['parts'][0]['text']
                    return f"[VISION/Gemini] {Path(path).name}:\n\n{result}"
                return f"[ERR] Gemini vision error: {data}"
        except Exception as e:
            return f"Error analyzing image (Gemini): {e}"

    async def _analyze_image_ollama(self, path, prompt):
        """Analyze image using an Ollama vision model (llava, moondream, etc)."""
        import base64
        from pathlib import Path
        try:
            with open(path, 'rb') as f:
                image_data = base64.b64encode(f.read()).decode('utf-8')

            ollama_base = self.core.config.get('providers', {}).get('ollama', {}).get('baseUrl', 'http://127.0.0.1:11434/v1')
            if not ollama_base.rstrip('/').endswith('/v1'):
                ollama_base = ollama_base.rstrip('/') + '/v1'
            url = f"{ollama_base}/chat/completions"

            payload = {
                "model": self.llm.model,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{image_data}"}}
                    ]
                }]
            }

            async with httpx.AsyncClient(timeout=90.0, verify=False) as client:
                response = await client.post(url, json=payload)
                data = response.json()
                if 'choices' in data and data['choices']:
                    result = data['choices'][0]['message']['content']
                    return f"[VISION/Ollama] {Path(path).name}:\n\n{result}"
                return f"[ERR] Ollama vision error: {data}\n(Ensure you're using a vision-capable model like llava or moondream)"
        except Exception as e:
            return f"Error analyzing image (Ollama): {e}"

    # ── Vision routing (base64 pipeline) ─────────────────────────────────────
    # These methods accept pre-encoded base64 + MIME type, eliminating the
    # temp-file race condition from _handle_photo.

    async def _analyze_image_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Route image analysis to the best available provider (base64 input)."""
        provider = self.llm.provider
        if provider == "google":
            return await self._analyze_image_gemini_b64(image_b64, mime_type, prompt)
        elif provider == "anthropic":
            return await self._analyze_image_anthropic_b64(image_b64, mime_type, prompt)
        elif provider == "nvidia":
            return await self._analyze_image_nvidia_b64(image_b64, mime_type, prompt)
        elif provider == "ollama":
            return await self._analyze_image_ollama_b64(image_b64, mime_type, prompt)
        else:
            # xai, groq, openai, openrouter, etc. — try Google first, then OpenAI-compat
            google_key = self.core.config.get('providers', {}).get('google', {}).get('apiKey')
            if google_key:
                return await self._analyze_image_gemini_b64(image_b64, mime_type, prompt)
            return await self._analyze_image_openai_b64(image_b64, mime_type, prompt)

    async def _analyze_image_gemini_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Analyze image using Google Gemini Vision (base64 input)."""
        try:
            api_key = self.config.get('api_key') or self.core.config.get('providers', {}).get('google', {}).get('apiKey')
            if not api_key:
                return "[ERR] Google API key not configured for image analysis."

            # Use active model if Google, else fall back to gemini-2.5-flash
            vision_model = self.llm.model if self.llm.provider == "google" else "gemini-2.5-flash"

            url = f"https://generativelanguage.googleapis.com/v1beta/models/{vision_model}:generateContent?key={api_key}"
            payload = {"contents": [{"parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": mime_type, "data": image_b64}}
            ]}]}

            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(url, json=payload)
                data = response.json()
                if 'candidates' in data and data['candidates']:
                    result = data['candidates'][0]['content']['parts'][0]['text']
                    return f"[VISION/Gemini/{vision_model}]\n\n{result}"
                return f"[ERR] Gemini vision error: {data}"
        except Exception as e:
            return f"Error analyzing image (Gemini): {e}"

    async def _analyze_image_anthropic_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Analyze image using Anthropic Claude vision (native multimodal format)."""
        try:
            api_key = self._get_provider_api_key("anthropic")
            if not api_key:
                return "[ERR] Anthropic API key not configured."

            url = "https://api.anthropic.com/v1/messages"
            if api_key.startswith("sk-ant-oat"):
                headers = {
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {api_key}",
                    "anthropic-version": "2023-06-01",
                    "anthropic-beta": "claude-code-20250219,oauth-2025-04-20",
                }
            else:
                headers = {
                    "Content-Type": "application/json",
                    "x-api-key": api_key,
                    "anthropic-version": "2023-06-01",
                }

            vision_model = self.llm.model if self.llm.provider == "anthropic" else "claude-sonnet-4-6"

            payload = {
                "model": vision_model,
                "max_tokens": 1024,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image", "source": {
                            "type": "base64",
                            "media_type": mime_type,
                            "data": image_b64,
                        }},
                        {"type": "text", "text": prompt}
                    ]
                }]
            }

            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()
                if "content" in data and data["content"]:
                    text_blocks = [b["text"] for b in data["content"] if b.get("type") == "text"]
                    return f"[VISION/Anthropic/{vision_model}]\n\n" + "\n".join(text_blocks)
                return f"[ERR] Anthropic vision error: {data}"
        except Exception as e:
            return f"Error analyzing image (Anthropic): {e}"

    async def _analyze_image_nvidia_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Analyze image using NVIDIA vision endpoint (phi-3.5-vision-instruct)."""
        try:
            api_key = self._get_provider_api_key("nvidia")
            if not api_key:
                return "[ERR] NVIDIA API key not configured."

            vision_model = "microsoft/phi-3.5-vision-instruct"
            url = "https://integrate.api.nvidia.com/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {api_key}",
                "Content-Type": "application/json",
            }
            payload = {
                "model": vision_model,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}}
                    ]
                }],
                "max_tokens": 1024,
                "temperature": 0.2,
            }

            async with httpx.AsyncClient(timeout=60.0) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()
                if 'choices' in data and data['choices']:
                    result = data['choices'][0]['message']['content']
                    return f"[VISION/NVIDIA/{vision_model}]\n\n{result}"
                return f"[ERR] NVIDIA vision error: {data}"
        except Exception as e:
            return f"Error analyzing image (NVIDIA): {e}"

    async def _analyze_image_ollama_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Analyze image using Ollama vision model (correct MIME type)."""
        try:
            ollama_base = self.core.config.get('providers', {}).get('ollama', {}).get('baseUrl', 'http://127.0.0.1:11434/v1')
            if not ollama_base.rstrip('/').endswith('/v1'):
                ollama_base = ollama_base.rstrip('/') + '/v1'
            url = f"{ollama_base}/chat/completions"

            payload = {
                "model": self.llm.model,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}}
                    ]
                }]
            }

            async with httpx.AsyncClient(timeout=90.0, verify=False) as client:
                response = await client.post(url, json=payload)
                data = response.json()
                if 'choices' in data and data['choices']:
                    result = data['choices'][0]['message']['content']
                    return f"[VISION/Ollama]\n\n{result}"
                return f"[ERR] Ollama vision error: {data}\n(Ensure you're using a vision-capable model like llava or moondream)"
        except Exception as e:
            return f"Error analyzing image (Ollama): {e}"

    async def _analyze_image_openai_b64(self, image_b64: str, mime_type: str, prompt: str) -> str:
        """Analyze image via OpenAI-compatible multimodal format (xai, groq, openai, etc.)."""
        try:
            provider = self.llm.provider
            url = f"{self._get_provider_base_url(provider)}/chat/completions"
            api_key = self._get_provider_api_key(provider)
            headers = {"Content-Type": "application/json", "Authorization": f"Bearer {api_key}"}

            payload = {
                "model": self.llm.model,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "text", "text": prompt},
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}}
                    ]
                }],
                "max_tokens": 1024,
            }

            async with httpx.AsyncClient(timeout=60.0, verify=False) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()
                if 'choices' in data and data['choices']:
                    result = data['choices'][0]['message']['content']
                    return f"[VISION/{provider.upper()}]\n\n{result}"
                return f"[ERR] {provider} vision error: {data}"
        except Exception as e:
            return f"Error analyzing image ({self.llm.provider}): {e}"

    async def tool_memory_search(self, args):
        """Search semantic memory for relevant context."""
        query = args.get('query')
        top_k = int(args.get('top_k', 5))
        
        try:
            # Access core memory (could be semantic or keyword-based)
            if hasattr(self.core, 'memory'):
                results = await self.core.memory.recall(query, top_k=top_k)
                
                if not results:
                    return f"[MEMORY] No relevant memories found for: {query}"
                
                # Format results
                formatted = [f"[MEMORY] Found {len(results)} relevant memories:\n"]
                for i, mem in enumerate(results, 1):
                    score = mem.get('relevance_score', 'N/A')
                    content_preview = mem['content'][:200] + "..." if len(mem['content']) > 200 else mem['content']
                    source = mem.get('metadata', {}).get('source', 'unknown')
                    formatted.append(f"\n{i}. [Score: {score}] ({source})\n{content_preview}\n")
                
                return "".join(formatted)
            else:
                return "[ERR] Memory system not available."
        except Exception as e:
            return f"Error searching memory: {e}"
    
    async def tool_memory_imprint(self, args):
        """Save information to long-term memory and persist to MEMORY.md."""
        content = args.get('content')
        tags = args.get('tags', '')

        try:
            if hasattr(self.core, 'memory'):
                metadata = {
                    "source": "manual_imprint",
                    "tags": tags
                }
                await self.core.memory.imprint(content, metadata)

                # Also write to MEMORY.md so it appears in every future system prompt
                try:
                    workspace = self.core.config.get('paths', {}).get('workspace', '')
                    if workspace:
                        memory_path = os.path.join(workspace, 'MEMORY.md')
                        from datetime import datetime
                        timestamp = datetime.now().strftime('%Y-%m-%d')
                        tag_str = f" [{tags}]" if tags else ""
                        entry = f"\n- {timestamp}{tag_str}: {content}"
                        # Create file with header if it doesn't exist
                        if not os.path.exists(memory_path):
                            with open(memory_path, 'w', encoding='utf-8') as f:
                                f.write("# Memory\n")
                        with open(memory_path, 'a', encoding='utf-8') as f:
                            f.write(entry)
                        # Reload personality so next prompt includes this memory
                        if hasattr(self, 'personality') and hasattr(self.personality, 'reload_memory'):
                            self.personality.reload_memory()
                except Exception:
                    pass  # MEMORY.md write is best-effort; imprint already succeeded

                return f"[MEMORY] Saved to long-term memory. Tags: {tags or 'none'}"
            else:
                return "[ERR] Memory system not available."
        except Exception as e:
            return f"Error saving to memory: {e}"
    
    async def tool_text_to_speech(self, args):
        """Convert text to speech using ElevenLabs, edge-tts (free male), or gTTS fallback."""
        text = args.get('text')
        # Voice options:
        #   'Nova'  → ElevenLabs Rachel (female, premium)
        #   'Byte'  → ElevenLabs Adam (male, premium)
        #   'Guy'   → edge-tts en-US-GuyNeural (male, FREE, no key needed)
        #   'Aria'  → edge-tts en-US-AriaNeural (female, FREE, no key needed)
        #   'gtts'  → Google TTS (female, FREE, no key needed)
        # Default pulled from config.yaml elevenlabs.voice, fallback to 'Guy'
        cfg_voice = self.core.config.get('elevenlabs', {}).get('voice', 'Guy')
        voice = args.get('voice', cfg_voice)

        try:
            import hashlib as _hashlib

            text_hash = _hashlib.md5(text.encode()).hexdigest()[:8]
            logs_dir = self.config.get('paths', {}).get('logs', './logs')
            os.makedirs(logs_dir, exist_ok=True)

            # ── ElevenLabs (premium) ─────────────────────────────────────────
            el_key = self.core.config.get('elevenlabs', {}).get('api_key', '')
            if el_key and voice in ('Nova', 'Byte', 'Default'):
                try:
                    from elevenlabs import generate, save
                    voice_map = {
                        'Nova':    '21m00Tcm4TlvDq8ikWAM',  # Rachel
                        'Byte':    'pNInz6obpgDQGcFmaJgB',  # Adam
                        'Default': '21m00Tcm4TlvDq8ikWAM',
                    }
                    output_path = os.path.join(logs_dir, f'tts_{text_hash}.mp3')
                    audio = generate(text=text, voice=voice_map.get(voice, voice_map['Default']), api_key=el_key)
                    save(audio, output_path)
                    return f"[VOICE] Generated speech: {output_path}"
                except Exception as e:
                    pass  # Fall through to free options

            # ── edge-tts (FREE — Microsoft neural voices, no key needed) ─────
            # Voices: Guy = male, Aria = female, Jenny = female, Davis = male
            edge_voice_map = {
                'Guy':   'en-US-GuyNeural',    # Natural male voice
                'Davis': 'en-US-DavisNeural',  # Expressive male voice
                'Aria':  'en-US-AriaNeural',   # Natural female voice
                'Jenny': 'en-US-JennyNeural',  # Friendly female voice
                'Byte':  'en-US-GuyNeural',    # Byte defaults to Guy when no EL key
                'Nova':  'en-US-AriaNeural',   # Nova defaults to Aria when no EL key
            }
            edge_voice_name = edge_voice_map.get(voice, 'en-US-GuyNeural')
            try:
                import edge_tts
                output_path = os.path.join(logs_dir, f'tts_{text_hash}.mp3')
                communicate = edge_tts.Communicate(text, edge_voice_name)
                await communicate.save(output_path)
                return f"[VOICE] Generated speech: {output_path}"
            except ImportError:
                pass  # edge-tts not installed, fall through to gTTS

            # ── gTTS (FREE fallback — female only) ───────────────────────────
            try:
                from gtts import gTTS
                output_path = os.path.join(logs_dir, f'tts_{text_hash}.mp3')
                gTTS(text=text, lang='en', slow=False).save(output_path)
                return f"[VOICE] Generated speech: {output_path}"
            except ImportError:
                pass

            return "[ERR] No TTS engine available. Run: pip install edge-tts"

        except Exception as e:
            return f"Error generating speech: {e}"

    # --- LLM Interaction ---

    def _extract_tool_call(self, response_text):
        """
        Robustly extract a tool-call JSON object from an LLM response.

        Handles all the messy ways local models (Qwen, Llama, Mistral, etc.) output JSON:
          • Bare JSON:                  {"tool": "...", "args": {...}}
          • Markdown fenced:            ```json\n{"tool":...}\n```
          • Inline wrapped:             "I'll use the tool: {"tool":...}"
          • Think-tag wrapped (Qwen3):  <think>...</think>{"tool":...}
          • action/action_input schema: {"action":"tool","action_input":{...}}
          • Nested tool schema:         {"name":"tool","parameters":{...}}

        Returns (tool_name, tool_args) tuple or (None, None) if no valid tool call found.
        """
        if not response_text or "{" not in response_text:
            return None, None

        # Step 1: Strip <think>...</think> blocks (Qwen3, DeepSeek-R1 emit these)
        cleaned = re.sub(r'<think>.*?</think>', '', response_text, flags=re.DOTALL).strip()

        # Step 2: Try to pull JSON from markdown code fences first (highest confidence)
        fence_match = re.search(r'```(?:json)?\s*(\{.*?\})\s*```', cleaned, re.DOTALL)
        candidates = []
        if fence_match:
            candidates.append(fence_match.group(1))

        # Step 3: Find ALL {...} spans in the cleaned text in O(N)
        # We use a stack to find all balanced brace pairs.
        stack = []
        for i, ch in enumerate(cleaned):
            if ch == '{':
                stack.append(i)
            elif ch == '}':
                if stack:
                    start_idx = stack.pop()
                    candidates.append(cleaned[start_idx:i+1])

        # Try outermost/largest blocks first (they were completed last)
        candidates.reverse()

        # Step 4: Try each candidate JSON blob
        for json_str in candidates:
            try:
                obj = json.loads(json_str)
                if not isinstance(obj, dict):
                    continue

                # Schema A: standard Galactic format {"tool": "name", "args": {...}}
                if "tool" in obj:
                    args = obj.get("args")
                    if args is None:
                        # Synthesize args from remaining keys if model put them at root level
                        args = {k: v for k, v in obj.items() if k != "tool"}
                    elif not isinstance(args, dict):
                        args = {}
                    return obj["tool"], args

                # Schema B: LangChain-style {"action": "name", "action_input": {...}}
                if "action" in obj and "action_input" in obj:
                    return obj["action"], obj["action_input"]

                # Schema C: OpenAI function-call style {"name": "name", "parameters": {...}}
                if "name" in obj and "parameters" in obj and obj["name"] in self.tools:
                    return obj["name"], obj["parameters"]

                # Schema D: {"function": "name", "arguments": {...}}
                if "function" in obj and "arguments" in obj:
                    args = obj["arguments"]
                    if isinstance(args, str):
                        try:
                            args = json.loads(args)
                        except Exception:
                            args = {}
                    return obj["function"], args

            except (json.JSONDecodeError, TypeError):
                continue

        return None, None

    def format_plan(self, plan: dict) -> str:
        """Formats the active plan into a readable string for the AI."""
        if not plan or not plan.get('steps'):
            return "No active plan."

        formatted_steps = []
        for i, step in enumerate(plan['steps']):
            prefix = "-> " if i == plan['current_step'] else "   "
            formatted_steps.append(f"{prefix}{i+1}. {step}")
        return "\n".join(formatted_steps)

    def _build_system_prompt(self, context="", is_ollama=False):
        """
        Build the system prompt. We inject:
          - Full parameter schemas so the model knows exact argument names and types
          - Concrete few-shot examples of correct tool-call JSON
          - Explicit instruction to output ONLY raw JSON
        """
        personality_prompt = self.personality.get_system_prompt()

        # Full schema for every tool so models know what args to pass
        tool_schemas = {}
        for name, tool in self.tools.items():
            tool_schemas[name] = {
                "description": tool.get("description", ""),
                "parameters": tool.get("parameters", {})
            }
        tool_block = json.dumps(tool_schemas, indent=2)

        few_shot = (
            'EXAMPLES OF CORRECT TOOL CALLS:\n'
            '  Read a file:\n'
            '  {"tool": "read_file", "args": {"path": "C:\\\\data\\\\notes.txt"}}\n\n'
            '  Run a shell command:\n'
            '  {"tool": "exec_shell", "args": {"command": "dir C:\\\\Users"}}\n\n'
            '  Navigate browser to URL:\n'
            '  {"tool": "browser_navigate", "args": {"url": "https://example.com"}}\n\n'
            '  Take a screenshot:\n'
            '  {"tool": "browser_screenshot", "args": {}}\n\n'
            '  Search the web:\n'
            '  {"tool": "web_search", "args": {"query": "python asyncio tutorial"}}\n'
        )

        protocol = (
            "TOOL USAGE RULES — FOLLOW EXACTLY:\n"
            "1. To use a tool output ONLY a raw JSON object. NO markdown. NO prose. NO code fences.\n"
            "   CORRECT:   {\"tool\": \"read_file\", \"args\": {\"path\": \"/tmp/a.txt\"}}\n"
            "   WRONG:     ```json\\n{...}\\n```   (never use fences)\n"
            "   WRONG:     'I will read the file: {...}'  (never wrap in prose)\n"
            "2. After a tool output appears as 'Tool Output: ...' give your FINAL answer in plain text.\n"
            "3. For simple tasks: use 1 tool then answer immediately.\n"
            "4. For complex tasks: chain up to 10 tool calls, then answer.\n"
            "5. NEVER repeat a tool call with the same args — trust the output.\n"
            "6. If you don't need a tool, just answer in plain text — no JSON.\n"
            "7. If a tool fails or times out: do NOT retry the same approach. Explain the failure and try a different strategy.\n"
            "8. If stuck after 3+ failed attempts: STOP. Tell the user what you tried, what went wrong, and ask for guidance.\n"
            "9. BEFORE writing scripts: read config.yaml for real credentials. NEVER use placeholder values.\n"
            "10. NEVER overwrite requirements.txt, config.yaml, or core .py files. Create NEW files with unique names.\n"
            "11. NEVER run scripts with while True loops or sleep() via exec_shell — they timeout. Tell the user how to launch them.\n"
            "12. CRITICAL: NEVER use `write_file` or `edit_file` on an existing file without using `read_file` first to see its content.\n"
            "13. CRITICAL: If you spawn a background task or subagent, DO NOT hallucinate that it finished successfully. You MUST check its status or read the output files before claiming success.\n"
        )

        system_prompt = (
            f"{personality_prompt}\n\n"
            f"AVAILABLE TOOLS (with parameter schemas):\n{tool_block}\n\n"
            f"{few_shot}\n"
            f"{protocol}\n"
            f"Context: {context}"
        )

        return system_prompt

    async def _send_telegram_typing_ping(self, chat_id):
        """Helper to send a typing indicator to Telegram if the bridge is active."""
        if hasattr(self.core, 'telegram_bridge'):
            try:
                await self.core.telegram_bridge.send_typing(chat_id)
            except Exception as e:
                await self.core.log(f"Telegram typing ping error: {e}", priority=1)

    async def _emit_trace(self, phase, turn, **kwargs):
        """Emit a structured agent_trace event to all connected WS clients."""
        payload = {"phase": phase, "turn": turn, "ts": time.time()}
        payload.update(kwargs)
        await self.core.relay.emit(3, "agent_trace", payload)

    async def checkpoint(self, uuid_str):
        """Save the current agent state to a JSON file."""
        if not uuid_str:
            return
        run_dir = os.path.join(self.runs_dir, uuid_str)
        os.makedirs(run_dir, exist_ok=True)
        checkpoint_path = os.path.join(run_dir, 'checkpoint.json')
        
        # Mask API key for security
        import copy
        api_key_masked = "NONE"
        if self.llm.api_key and len(self.llm.api_key) > 8:
            api_key_masked = "***" + self.llm.api_key[-8:]
            
        state = {
            'uuid': uuid_str,
            'history': copy.deepcopy(self.history),
            'active_plan': copy.deepcopy(self.active_plan),
            'turn_count': self._tool_count_since_cp,
            'llm_state': {
                'provider': self.llm.provider,
                'model': self.llm.model,
                'api_key_mask': api_key_masked
            },
            'trace_sid': self._trace_sid,
            'recent_tools': copy.deepcopy(self._recent_tools),
            'consecutive_failures': self._consecutive_failures
        }
        
        try:
            with open(checkpoint_path, 'w', encoding='utf-8') as f:
                json.dump(state, f, indent=2)
            await self.core.log(f"💾 Checkpoint saved: {uuid_str}", priority=3)
        except Exception as e:
            await self.core.log(f"⚠️ Failed to save checkpoint {uuid_str}: {e}", priority=1)

    async def load_checkpoint(self, uuid_str):
        """Load a previously saved agent state."""
        checkpoint_path = os.path.join(self.runs_dir, uuid_str, 'checkpoint.json')
        if not os.path.exists(checkpoint_path):
            return False
            
        try:
            with open(checkpoint_path, 'r', encoding='utf-8') as f:
                state = json.load(f)
                
            self.checkpoint_uuid = uuid_str
            self.history = state.get('history', [])
            self.active_plan = state.get('active_plan')
            self._tool_count_since_cp = state.get('turn_count', 0)
            self._trace_sid = state.get('trace_sid')
            self._recent_tools = state.get('recent_tools', [])
            self._consecutive_failures = state.get('consecutive_failures', 0)
            
            # Note: We do not restore the raw API key from state since it is masked.
            # The current initialized provider/model API key will be used.
            
            await self.core.log(f"🔄 Checkpoint restored: {uuid_str}", priority=2)
            return True
        except Exception as e:
            await self.core.log(f"⚠️ Failed to load checkpoint {uuid_str}: {e}", priority=1)
            return False

    async def checkpoint(self, turn_count, messages):
        """Save the current workflow state to disk for resumability."""
        if not self.checkpoint_uuid:
            self.checkpoint_uuid = str(uuid.uuid4())[:8]
            
        cp_dir = os.path.join(self.runs_dir, self.checkpoint_uuid)
        os.makedirs(cp_dir, exist_ok=True)
        
        # Mask API key before saving
        llm_state = {
            "provider": self.llm.provider,
            "model": self.llm.model,
            "api_key_mask": f"***{self.llm.api_key[-8:]}" if hasattr(self.llm, 'api_key') and self.llm.api_key else "NONE"
        }
        
        import copy
        state = {
            "uuid": self.checkpoint_uuid,
            "history": copy.deepcopy(self.history),
            "messages": copy.deepcopy(messages),
            "active_plan": copy.deepcopy(self.active_plan),
            "turn_count": turn_count,
            "llm_state": llm_state,
            "trace_sid": self._trace_sid,
            "recent_tools": copy.deepcopy(self._recent_tools),
            "consecutive_failures": self._consecutive_failures
        }
        
        cp_path = os.path.join(cp_dir, "checkpoint.json")
        try:
            with open(cp_path, 'w', encoding='utf-8') as f:
                json.dump(state, f, indent=2)
            await self.core.log(f"💾 Checkpoint saved: {self.checkpoint_uuid}", priority=3)
        except Exception as e:
            await self.core.log(f"⚠️ Failed to save checkpoint: {e}", priority=1)

    async def load_checkpoint(self, target_uuid):
        """Restore workflow state from a saved checkpoint."""
        cp_path = os.path.join(self.runs_dir, target_uuid, "checkpoint.json")
        if not os.path.exists(cp_path):
            raise FileNotFoundError(f"Checkpoint {target_uuid} not found")
            
        with open(cp_path, 'r', encoding='utf-8') as f:
            state = json.load(f)
            
        self.checkpoint_uuid = state.get('uuid')
        self.history = state.get('history', [])
        # We don't restore the API key, we keep the currently loaded config
        self.active_plan = state.get('active_plan')
        self._trace_sid = state.get('trace_sid')
        self._recent_tools = state.get('recent_tools', [])
        self._consecutive_failures = state.get('consecutive_failures', 0)
        self._tool_count_since_cp = 0  # reset for new run
        
        await self.core.log(f"🔄 Restored checkpoint: {target_uuid}", priority=2)
        return state

    async def _generate_plan(self, user_input):
        """Generates a step-by-step plan using an isolated planner agent that can scan the codebase."""
        planner_provider = self.core.config.get('models', {}).get('planner_provider')
        planner_model = self.core.config.get('models', {}).get('planner_model')

        planner_fallback_provider = self.core.config.get('models', {}).get('planner_fallback_provider')
        planner_fallback_model = self.core.config.get('models', {}).get('planner_fallback_model')

        # Fallback to simple gemini_code tool if no planner model is explicitly configured
        if not planner_provider or not planner_model:
            if "gemini_code" not in self.tools:
                await self.core.log("[Planner] Gemini Coder tool not available for planning.", priority=1)
                return None

            planning_prompt = _PLANNING_PROMPT_TEMPLATE.format(user_input=user_input)
            await self.core.log(f"[Planner] Generating plan for: {user_input[:80]}...")
            await self._emit_trace("planning_start", 0, session_id="planner", query=user_input[:500])

            try:
                # Use the gemini_code tool for planning
                plan_raw_output = await self.tools["gemini_code"]["fn"]({"prompt": planning_prompt, "model": "gemini-3-flash-preview"})
                
                # Extract the numbered list
                plan_steps = re.findall(r'^\s*\d\.\s*(.*)', plan_raw_output, re.MULTILINE)
                
                if plan_steps:
                    plan = { "steps": plan_steps, "current_step": 0, "original_query": user_input }
                    await self.core.log(f"[Planner] Generated plan with {len(plan_steps)} steps.", priority=2)
                    await self._emit_trace("plan_generated", 0, session_id="planner", plan=plan_steps)
                    
                    # Store the plan in long-term memory
                    if "store_memory" in self.tools:
                        plan_text = "\n".join([f"{i+1}. {step}" for i, step in enumerate(plan_steps)])
                        await self.tools["store_memory"]["fn"]({
                            "text": f"User request: {user_input}\nGenerated Plan:\n{plan_text}",
                            "metadata": { "type": "plan", "original_query": user_input[:200] }
                        })
                    return plan
                else:
                    await self.core.log("[Planner] Failed to extract plan steps from LLM response. Using raw output.", priority=1)
                    # Fallback: Treat the whole response as one big step if no numbers found
                    plan_steps = [s.strip() for s in plan_raw_output.split('\n') if s.strip()]
                    if plan_steps:
                         return { "steps": plan_steps, "current_step": 0, "original_query": user_input }
                    return None
            except Exception as e:
                await self.core.log(f"[Planner] Error generating plan: {e}", priority=1)
                return None

        # --- ADVANCED PLANNER LOOP ---
        planner_context = (
            "You are the Lead Architect and Strategic Planner.\n"
            "Your job is to thoroughly analyze the user's request, scan the necessary files and codebase to understand the context, "
            "and output a detailed step-by-step implementation plan.\n"
            "DO NOT execute the final changes (do not write the final code). Only explore, use tools like `list_dir`, `find_files`, `regex_search`, or `read_file` to gather information.\n"
            "Once you have fully investigated the problem and formulated a plan, output your final plan wrapped EXACTLY in <plan>...</plan> tags as a numbered list.\n"
            "Focus on identifying information gaps, research needs, and logical progression."
        )

        attempts = [
            (planner_provider, planner_model, "Primary"),
        ]
        if planner_fallback_provider and planner_fallback_model:
            attempts.append((planner_fallback_provider, planner_fallback_model, "Fallback"))

        for prov, mod, label in attempts:
            await self.core.log(f"[Planner] Spawning isolated planner agent ({label}: {prov}/{mod})...")
            await self._emit_trace("planning_start", 0, session_id="planner", query=user_input[:500])

            try:
                # Run the isolated ReAct loop using the planner model with a strict timeout
                result = await asyncio.wait_for(
                    self.speak_isolated(
                        user_input=f"Analyze and plan the following task:\n\n{user_input}",
                        context=planner_context,
                        override_provider=prov,
                        override_model=mod,
                        use_lock=False, # ALREADY LOCKED BY SPEAK()
                        skip_planning=True # PREVENT RECURSION
                    ),
                    timeout=300
                )

                if result and result.startswith("[ERROR]"):
                    raise Exception(result)

                # Extract the <plan> from the result
                match = re.search(r'<plan>(.*?)</plan>', result, re.DOTALL)
                plan_text = match.group(1).strip() if match else result
                
                # Extract the numbered list
                plan_steps = re.findall(r'^\s*\d\.\s*(.*)', plan_text, re.MULTILINE)
                
                if not plan_steps:
                    # Fallback if no numbers were used (split by lines)
                    plan_steps = [s.strip() for s in plan_text.split('\n') if s.strip()][:10]

                if plan_steps:
                    plan = { "steps": plan_steps, "current_step": 0, "original_query": user_input }
                    await self.core.log(f"[Planner] Generated plan with {len(plan_steps)} steps.", priority=2)
                    await self._emit_trace("plan_generated", 0, session_id="planner", plan=plan_steps)
                    
                    # Store the plan in long-term memory
                    if "store_memory" in self.tools:
                        formatted_plan = "\n".join([f"{i+1}. {step}" for i, step in enumerate(plan_steps)])
                        await self.tools["store_memory"]["fn"]({
                            "text": f"User request: {user_input}\nGenerated Plan:\n{formatted_plan}",
                            "metadata": { "type": "plan", "original_query": user_input[:200] }
                        })
                    return plan
                else:
                    await self.core.log(f"[Planner] Failed to extract plan steps from Planner Agent ({label}) output.", priority=1)
                    continue # Try fallback if extraction fails
                    
            except asyncio.TimeoutError:
                await self.core.log(f"[Planner] {label} model timed out after 300 seconds", priority=1)
                continue
            except Exception as e:
                await self.core.log(f"[Planner] Error generating plan via {label} agent: {e}", priority=1)
                continue
                
        await self.core.log("[Planner] All planner attempts failed.", priority=1)
        return None

    async def speak(self, user_input, context="", chat_id=None, images=None, skip_planning=False):
        """
        Main entry point for user interaction.
        Serialized via _speak_lock to prevent concurrent executions and duplicate planners.
        """
        async with self._speak_lock:
            try:
                return await self._speak_logic(user_input, context=context, chat_id=chat_id, images=images, skip_planning=skip_planning)
            except asyncio.CancelledError:
                # Catch the cancellation here at the top level to return a clean string
                # instead of letting the exception crash the request handler.
                await self._emit_trace("session_abort", 0, session_id=self._trace_sid,
                                       reason="user_cancelled")
                cancel_msg = "🛑 Task cancelled by user."
                self.history.append({"role": "assistant", "content": cancel_msg})
                self._log_chat("assistant", cancel_msg, source="telegram" if chat_id else "web")
                return cancel_msg

    async def _speak_logic(self, user_input, context="", chat_id=None, images=None, skip_planning=False):
        """
        Internal implementation of the ReAct loop.
        Expects caller to handle locking and state snapshots.

        images: optional list of {name, mime, b64} dicts for vision-capable models.
          When provided, the user message is built as a multimodal content array
          (text + base64 image parts) compatible with OpenAI/Anthropic/Google vision APIs.

        Ollama/local models get:
          - Full parameter schemas in system prompt
          - Few-shot tool-call examples
          - Robust multi-pattern JSON extraction (handles think-tags, fences, prose wrapping)
          - Full messages array passed directly (not collapsed to a string)
        """
        # Track input tokens (rough estimate: 1 token ~= 4 chars)
        self._estimated_input_tokens = len(user_input) // 4
        self.total_tokens_in += self._estimated_input_tokens

        # Initialize active_plan if not present
        if not hasattr(self, 'active_plan'):
            self.active_plan = None # { 'steps': [], 'current_step': 0, 'original_query_id': None }

        # Reset per-turn state
        self.last_voice_file = None

        # Build user message — multimodal content array if images are attached
        if images:
            content = []
            if user_input:
                content.append({"type": "text", "text": user_input})
            for img in images:
                content.append({
                    "type": "image_url",
                    "image_url": {
                        "url": f"data:{img['mime']};base64,{img['b64']}"
                    }
                })
            self.history.append({"role": "user", "content": content})
        else:
            self.history.append({"role": "user", "content": user_input})

        # Persist to JSONL
        source = "telegram" if chat_id else "web"
        self._log_chat("user", user_input, source=source)

        # Smart model routing — pick the best model for this task type (opt-in via config)
        model_mgr = getattr(self.core, 'model_manager', None)
        if model_mgr:
            await model_mgr.auto_route(user_input)

        # Determine if we're on a local/Ollama model
        is_ollama = (self.llm.provider == "ollama")

        # ── Planning Phase (for complex tasks) ──────────────────────────
        # Decide if a plan is needed. Trigger on explicit command or complex code tasks.
        needs_plan = False
        lower_input = user_input.lower()
        if not self.active_plan and not skip_planning:
            if lower_input.startswith("/plan ") or "plan out" in lower_input or "scan the codebase" in lower_input:
                needs_plan = True
                user_input = user_input.replace("/plan ", "").strip()
            elif any(kw in lower_input for kw in ["refactor", "build a ", "create a ", "write a script", "complex task", "update", "add", "fix", "change", "implement"]):
                needs_plan = True

        if needs_plan and not self.active_plan and not skip_planning:
            plan = await self._generate_plan(user_input)
            if plan:
                self.active_plan = plan
                # Add the plan to the context for the next turn
                context = f"You are currently executing a plan. Here is the plan:\n" \
                          f"{self.format_plan(self.active_plan)}\n\n" \
                          f"Focus on completing the current step before moving to the next.\n\n{context}"
                await self.core.log(f"[Planner] Activated plan for: {user_input[:80]}...")

        # 1. Build system prompt (Ollama gets full schemas + few-shot examples)
        system_prompt = self._build_system_prompt(context=context, is_ollama=is_ollama)
        messages = [{"role": "system", "content": system_prompt}] + self.history

        # 2. ReAct Loop (with wall-clock timeout)
        max_turns = int(self.config.get('models', {}).get('max_turns', 50))
        speak_timeout = float(self.core.config.get('models', {}).get('speak_timeout', 600))
        turn_count = 0
        last_tool_call = None  # Track last (tool_name, json_args_str) to prevent duplicate calls
        # Tools that are legitimately called repeatedly with same args (snapshots, reads, etc.)
        _DUPLICATE_EXEMPT = {'browser_snapshot', 'web_search', 'read_file', 'memory_search', 'generate_image'}

        # ── Anti-spin guardrails ──
        consecutive_failures = 0   # Consecutive tool errors/timeouts
        recent_tools = []          # Rolling window of last 6 tool names
        _nudge_half_sent = False   # Track whether 50% nudge was sent
        _nudge_80_sent = False     # Track whether 80% nudge was sent

        # Mark that the gateway is actively processing (prevents model switching mid-task)
        self._speaking = True

        # Unique session ID for tracing this speak() invocation
        trace_sid = self._trace_sid or str(uuid.uuid4())[:8]
        self._trace_sid = trace_sid
        
        if not self.checkpoint_uuid:
            self.checkpoint_uuid = str(uuid.uuid4())[:8]

        await self._emit_trace("session_start", 0, session_id=trace_sid,
                               query=user_input[:500])

        # ── Inner function: entire ReAct loop wrapped with wall-clock timeout ──
        async def _react_loop():
            nonlocal turn_count, last_tool_call, messages
            nonlocal consecutive_failures, recent_tools, _nudge_half_sent, _nudge_80_sent

            for _ in range(max_turns):
                turn_count += 1
                await self._emit_trace("turn_start", turn_count, session_id=trace_sid)

                # ── Progressive backpressure: nudge the AI to wrap up ──
                half_mark = max_turns // 2
                eighty_mark = int(max_turns * 0.8)
                if turn_count == half_mark and not _nudge_half_sent:
                    _nudge_half_sent = True
                    messages.append({
                        "role": "user",
                        "content": (
                            f"⚠️ You've used {turn_count} of {max_turns} tool turns. "
                            f"Start wrapping up — deliver what you have so far."
                        )
                    })
                    await self.core.log(
                        f"⚠️ Agent nudge: {turn_count}/{max_turns} turns used (50%)",
                        priority=2
                    )
                elif turn_count == eighty_mark and not _nudge_80_sent:
                    _nudge_80_sent = True
                    messages.append({
                        "role": "user",
                        "content": (
                            f"🛑 {turn_count}/{max_turns} turns used. "
                            f"Give your FINAL answer NOW. Summarize what you accomplished "
                            f"and what remains to be done."
                        )
                    })
                    await self.core.log(
                        f"🛑 Agent nudge: {turn_count}/{max_turns} turns used (80%)",
                        priority=1
                    )

                await self._send_telegram_typing_ping(chat_id)
                response_text = await self._call_llm_resilient(messages)

                # Capture think-tag content before stripping (for Thinking tab)
                think_match = re.search(r'<think>(.*?)</think>', response_text, re.DOTALL)
                if think_match:
                    await self._emit_trace("thinking", turn_count, session_id=trace_sid,
                                           content=think_match.group(1).strip()[:5000])

                # Emit raw LLM response
                await self._emit_trace("llm_response", turn_count, session_id=trace_sid,
                                       content=response_text[:3000])

                # Strip think-tags from final response text (Qwen3/DeepSeek-R1)
                display_text = re.sub(r'<think>.*?</think>', '', response_text, flags=re.DOTALL).strip()

                # Try to extract a tool call
                tool_name, tool_args = self._extract_tool_call(response_text)

                if tool_name is not None:
                    # Duplicate-call guard (prevents infinite loops with stubborn models)
                    # Exempt tools that are legitimately called repeatedly (snapshots, searches, etc.)
                    call_sig = f"{tool_name}:{json.dumps(tool_args, sort_keys=True)}"
                    if call_sig == last_tool_call and tool_name not in _DUPLICATE_EXEMPT:
                        await self.core.log(
                            f"⚠️ Duplicate tool call detected ({tool_name}), forcing final answer.",
                            priority=2
                        )
                        await self._emit_trace("duplicate_blocked", turn_count, session_id=trace_sid,
                                               tool=tool_name)
                        # Force the model to give a final answer
                        messages.append({"role": "assistant", "content": response_text})
                        messages.append({
                            "role": "user",
                            "content": (
                                "You already called that tool with those arguments. "
                                "Please give your FINAL answer now in plain text — no more tool calls."
                            )
                        })
                        last_tool_call = None
                        continue
                    last_tool_call = call_sig

                    # Fuzzy tool name match: handle "browser.navigate" → "browser_navigate" etc.
                    if tool_name not in self.tools:
                        normalized = tool_name.replace(".", "_").replace("-", "_").lower()
                        if normalized in self.tools:
                            tool_name = normalized
                        else:
                            # Try prefix match (e.g. model said "navigate" and we have "browser_navigate")
                            matches = [t for t in self.tools if t.endswith(f"_{normalized}") or t == normalized]
                            if len(matches) == 1:
                                tool_name = matches[0]

                    await self.core.log(f"🛠️ Executing: {tool_name} {tool_args}", priority=2)

                    if tool_name in self.tools:
                        # Emit tool_call trace before executing
                        await self._emit_trace("tool_call", turn_count, session_id=trace_sid,
                                               tool=tool_name,
                                               args=tool_args if isinstance(tool_args, dict) else str(tool_args)[:1000])
                        tool_timeout = self._get_tool_timeout(tool_name)
                        try:
                            result = await asyncio.wait_for(
                                self.tools[tool_name]["fn"](tool_args),
                                timeout=tool_timeout
                            )
                            await self._send_telegram_typing_ping(chat_id)
                            await self._emit_trace("tool_result", turn_count, session_id=trace_sid,
                                                   tool=tool_name, result=str(result)[:3000], success=True)
                        except asyncio.TimeoutError:
                            result = f"[Tool Timeout] {tool_name} exceeded {tool_timeout}s limit and was killed."
                            await self.core.log(f"⏱ Tool timeout: {tool_name} after {tool_timeout}s", priority=1)
                            await self._emit_trace("tool_result", turn_count, session_id=trace_sid,
                                                   tool=tool_name, result=result, success=False)
                        except Exception as e:
                            result = f"[Tool Error] {tool_name} raised: {type(e).__name__}: {e}"
                            await self._emit_trace("tool_result", turn_count, session_id=trace_sid,
                                                   tool=tool_name, result=str(result)[:3000], success=False)

                        # Track TTS output so callers (telegram_bridge) can send the audio file
                        if tool_name == "text_to_speech" and "[VOICE]" in str(result):
                            voice_match = re.search(r'Generated speech.*?:\s*(.+\.mp3)', str(result))
                            if voice_match:
                                self.last_voice_file = voice_match.group(1).strip()

                        # ── Anti-spin: track consecutive failures ──
                        result_str = str(result)
                        if result_str.startswith("[Tool Error]") or result_str.startswith("[Tool Timeout]"):
                            consecutive_failures += 1
                        else:
                            consecutive_failures = 0

                        # ── Anti-spin: track tool-type repetition ──
                        recent_tools.append(tool_name)
                        if len(recent_tools) > 6:
                            recent_tools.pop(0)
                            
                        # ── Resumable Checkpoints ──
                        self._tool_count_since_cp += 1
                        if self._tool_count_since_cp >= 5 or consecutive_failures > 0:
                            await self.checkpoint(turn_count, messages)
                            self._tool_count_since_cp = 0

                        messages.append({"role": "assistant", "content": response_text})

                        # Vision-capable tool result: screenshot or zoom returned an image dict
                        if isinstance(result, dict) and "__image_b64__" in result:
                            img_b64 = result.get("__image_b64__", "")
                            media_type = result.get("media_type", "image/jpeg")
                            caption = result.get("text", "[CHROME] Screenshot captured")
                            if img_b64 and isinstance(img_b64, str):
                                # Build multimodal user message so the LLM can visually see the screenshot
                                img_content = [
                                    {"type": "text", "text": f"Tool Output: {caption}"},
                                    {"type": "image_url", "image_url": {"url": f"data:{media_type};base64,{img_b64}"}}
                                ]
                                messages.append({"role": "user", "content": img_content})
                            else:
                                messages.append({"role": "user", "content": f"Tool Output: {caption}"})
                        else:
                            messages.append({"role": "user", "content": f"Tool Output: {result}"})

                        # ── Circuit breaker: 3+ consecutive failures ──
                        if consecutive_failures >= 3:
                            await self.core.log(
                                f"🔌 Circuit breaker: {consecutive_failures} consecutive tool failures",
                                priority=1
                            )
                            await self._emit_trace("circuit_breaker", turn_count, session_id=trace_sid,
                                                   failures=consecutive_failures)
                            messages.append({
                                "role": "user",
                                "content": (
                                    f"⚠️ {consecutive_failures} consecutive tool failures. "
                                    f"STOP calling tools. Explain to the user what you were trying to do, "
                                    f"what went wrong, and suggest next steps or ask for guidance."
                                )
                            })
                            consecutive_failures = 0  # Reset after intervention
                            break  # Hard break out of the ReAct loop

                        # ── Tool-type repetition guard ──
                        if len(recent_tools) >= 5:
                            from collections import Counter
                            tool_counts = Counter(recent_tools)
                            most_common_tool, most_common_count = tool_counts.most_common(1)[0]
                            if most_common_count >= 4 and most_common_tool not in _DUPLICATE_EXEMPT:
                                await self.core.log(
                                    f"🔄 Tool repetition guard: {most_common_tool} called "
                                    f"{most_common_count}x in last {len(recent_tools)} turns",
                                    priority=1
                                )
                                await self._emit_trace("repetition_guard", turn_count, session_id=trace_sid,
                                                       tool=most_common_tool, count=most_common_count)
                                messages.append({
                                    "role": "user",
                                    "content": (
                                        f"You've called {most_common_tool} {most_common_count} times in the "
                                        f"last {len(recent_tools)} turns without resolving the issue. "
                                        f"Try a completely different approach or explain the situation to the user."
                                    )
                                })
                                recent_tools.clear()  # Reset after intervention

                        continue  # Loop back to LLM with result
                    else:
                        await self._emit_trace("tool_not_found", turn_count, session_id=trace_sid,
                                               tool=tool_name)
                        tool_list_hint = ", ".join(list(self.tools.keys())[:20]) + "..."
                        messages.append({"role": "assistant", "content": response_text})
                        messages.append({
                            "role": "user",
                            "content": (
                                f"Error: Tool '{tool_name}' not found. "
                                f"Available tools include: {tool_list_hint} "
                                f"Please use the exact tool name from the list, then try again."
                            )
                        })
                        continue

                # No tool call detected → this is the final answer
                # Use display_text (think-tags stripped) for the history and relay
                await self._emit_trace("final_answer", turn_count, session_id=trace_sid,
                                       content=display_text[:3000])
                self.history.append({"role": "assistant", "content": display_text})
                # Only emit "thought" to the web UI if this is a web chat request.
                # Telegram calls are handled by process_and_respond which emits
                # "chat_from_telegram" — emitting "thought" here too causes duplicates.
                if not chat_id:
                    await self.core.relay.emit(2, "thought", display_text)

                self.total_tokens_out += len(display_text) // 4
                # Log cost with real token counts if available, otherwise estimates
                if hasattr(self.core, 'cost_tracker'):
                    real = self._last_usage
                    if real and (real.get('prompt_tokens') or real.get('completion_tokens')):
                        tin = real['prompt_tokens']
                        tout = real['completion_tokens']
                        # Update running totals with real counts (overwrite estimates)
                        self.total_tokens_in += tin - self._estimated_input_tokens
                        self.total_tokens_out += tout - (len(display_text) // 4)
                    else:
                        tin = self._estimated_input_tokens
                        tout = len(display_text) // 4
                    # Fetch actual cost from OpenRouter when available
                    actual_cost = None
                    gen_id = getattr(self, '_last_generation_id', None)
                    if self.llm.provider == 'openrouter' and gen_id:
                        actual_cost = await self._fetch_openrouter_generation_cost(gen_id)
                        self._last_generation_id = None

                    self.core.cost_tracker.log_usage(
                        model=self.llm.model,
                        provider=self.llm.provider,
                        tokens_in=tin,
                        tokens_out=tout,
                        actual_cost=actual_cost,
                    )

                # Persist to JSONL
                source = "telegram" if chat_id else "web"
                self._log_chat("assistant", display_text, source=source)

                return display_text

            # Hit max turns
            await self._emit_trace("session_abort", turn_count, session_id=trace_sid,
                                   reason="max_turns_exceeded")
            error_msg = (
                f"[ABORT] Hit maximum tool call limit ({max_turns} turns). "
                f"Used {turn_count} tool calls but couldn't form a final answer. "
                f"Try simplifying your query or asking for specific info."
            )
            self.total_tokens_out += len(error_msg) // 4
            self.history.append({"role": "assistant", "content": error_msg})
            self._log_chat("assistant", error_msg, source="telegram" if chat_id else "web")
            return error_msg

        # ── Execute the ReAct loop with wall-clock timeout ──
        t = asyncio.current_task()
        self._active_tasks.add(t)
        try:
            spinner.start()
            return await asyncio.wait_for(_react_loop(), timeout=speak_timeout)
        except asyncio.TimeoutError:
            await self._emit_trace("session_abort", turn_count, session_id=trace_sid,
                                   reason="speak_timeout")
            timeout_msg = (
                f"⏱ Task exceeded the maximum execution time ({int(speak_timeout)}s). "
                f"Completed {turn_count} turns before timeout. "
                f"Try breaking your request into smaller steps."
            )
            self.total_tokens_out += len(timeout_msg) // 4
            self.history.append({"role": "assistant", "content": timeout_msg})
            self._log_chat("assistant", timeout_msg, source="telegram" if chat_id else "web")
            return timeout_msg
        finally:
            self._active_tasks.discard(t)
            await spinner.stop()
            # ── Always clear speaking flag and restore smart routing ──
            self._speaking = False

            # Restore model if smart routing switched it for this request
            if model_mgr and getattr(model_mgr, '_routed', False):
                pre = getattr(model_mgr, '_pre_route_state', None)
                if pre:
                    self.llm.provider = pre['provider']
                    self.llm.model = pre['model']
                    self.llm.api_key = pre['api_key']
                    await self.core.log(
                        f"🔄 Smart routing restored: {pre['provider']}/{pre['model']}",
                        priority=3
                    )
                model_mgr._routed = False

            # Apply any queued model switch that arrived while we were speaking
            queued = getattr(self, '_queued_switch', None)
            if queued:
                q_provider, q_model = queued
                self._queued_switch = None
                if model_mgr:
                    model_mgr.primary_provider = q_provider
                    model_mgr.primary_model = q_model
                    model_mgr.current_mode = 'primary'
                    self.llm.provider = q_provider
                    self.llm.model = q_model
                    model_mgr._set_api_key(q_provider)
                    await model_mgr._save_config()
                    await self.core.log(
                        f"🔄 Queued model switch applied: {q_provider}/{q_model}",
                        priority=2
                    )

    # ── Isolated speak for sub-agents ─────────────────────────────────

    async def speak_isolated(self, user_input, context="", chat_id=None, images=None, override_provider=None, override_model=None, use_lock=True, skip_planning=False):
        """
        Run speak() with isolated state for sub-agents or planners.
        Saves and restores all mutable gateway state so concurrent calls
        don't corrupt the main agent's session.
        """
        if use_lock:
            async with self._speak_lock:
                return await self._speak_isolated_internal(user_input, context, chat_id, images, override_provider, override_model, skip_planning)
        else:
            return await self._speak_isolated_internal(user_input, context, chat_id, images, override_provider, override_model, skip_planning)

    async def _speak_isolated_internal(self, user_input, context, chat_id, images, override_provider, override_model, skip_planning):
        # Snapshot current state
        saved_speaking = self._speaking
        saved_history = self.history
        saved_llm_prov = self.llm.provider
        saved_llm_model = self.llm.model
        saved_llm_key = self.llm.api_key
        saved_queued = self._queued_switch

        # Snapshot model_manager routing state
        model_mgr = getattr(self.core, 'model_manager', None)
        saved_routed = getattr(model_mgr, '_routed', False) if model_mgr else False
        saved_pre_route = getattr(model_mgr, '_pre_route_state', None) if model_mgr else None

        try:
            # Apply overrides if provided
            if override_provider and override_model and model_mgr:
                self.llm.provider = override_provider
                self.llm.model = override_model
                model_mgr._set_api_key(override_provider)

            # Use isolated history (sub-agent has no prior conversation)
            self.history = []
            self._speaking = False
            self._queued_switch = None
            if model_mgr:
                model_mgr._routed = False
                model_mgr._pre_route_state = None

            return await self._speak_logic(user_input, context=context, chat_id=chat_id, images=images, skip_planning=skip_planning)
        finally:
            # Restore all state
            self._speaking = saved_speaking
            self.history = saved_history
            self.llm.provider = saved_llm_prov
            self.llm.model = saved_llm_model
            self.llm.api_key = saved_llm_key
            self._queued_switch = saved_queued
            if model_mgr:
                model_mgr._routed = saved_routed
                model_mgr._pre_route_state = saved_pre_route

    # ── Tool timeout defaults ────────────────────────────────────────
    _TOOL_TIMEOUTS = {
        'exec_shell': 120, 'execute_python': 60, 'open_browser': 60,
        'web_fetch': 30, 'web_search': 15, 'browser_click': 30,
        'browser_type': 15, 'browser_wait': 60, 'browser_extract': 30,
        'browser_snapshot': 30, 'browser_fill_form': 30,
        'browser_execute_js': 30, 'browser_pdf': 30,
        'desktop_screenshot': 60, 'desktop_click': 10, 'desktop_type': 15,
        'generate_image': 180, 'generate_image_sd35': 180,
        'generate_image_imagen': 180, 'analyze_image': 60,
        'text_to_speech': 30, 'spawn_subagent': 5, 'memory_search': 10,
        'memory_imprint': 10, 'wait': 310, 'read_file': 10, 'write_file': 10,
        'edit_file': 10, 'find_files': 30, 'list_dir': 10,
        'read_pdf': 30, 'read_csv': 15, 'read_excel': 15, 'write_csv': 15,
        'regex_search': 30, 'send_telegram': 15,
        'git_status': 15, 'git_diff': 15, 'git_commit': 30, 'git_log': 15,
        'image_resize': 15, 'image_convert': 15, 'http_request': 60,
        # Chrome Bridge tools
        'chrome_screenshot': 15, 'chrome_navigate': 30, 'chrome_read_page': 15,
        'chrome_find': 10, 'chrome_click': 10, 'chrome_type': 15,
        'chrome_scroll': 10, 'chrome_form_input': 10, 'chrome_execute_js': 30,
        'chrome_get_text': 15, 'chrome_tabs_list': 10, 'chrome_tabs_create': 10,
        'chrome_key_press': 10, 'chrome_read_console': 10, 'chrome_read_network': 10,
        'chrome_hover': 10,
        # Social Media tools
        'post_tweet': 30, 'read_mentions': 30, 'read_dms': 30,
        'post_reddit': 30, 'read_reddit_inbox': 30, 'reply_reddit': 30,
    }

    def _get_tool_timeout(self, tool_name):
        """Per-tool timeout: config override > built-in default > 60s."""
        overrides = self.core.config.get('tool_timeouts', {})
        return overrides.get(tool_name, self._TOOL_TIMEOUTS.get(tool_name, 60))

    # ── Resilient LLM call with fallback chain ───────────────────────

    async def _call_llm_resilient(self, messages):
        """
        Wrapper around _call_llm that detects [ERROR] responses and
        transparently retries / walks the fallback chain.
        On the happy path (no error), this adds zero overhead.
        """
        result = await self._call_llm(messages)

        # Happy path — no error
        if not isinstance(result, str) or not result.startswith("[ERROR]"):
            return result

        # Error detected — check if fallback is enabled
        model_mgr = getattr(self.core, 'model_manager', None)
        if not model_mgr or not model_mgr.auto_fallback_enabled:
            return result  # Fallback disabled — return error as-is

        error_type = model_mgr.classify_error(result)
        await self.core.log(
            f"⚠️ LLM error ({error_type}): {self.llm.provider}/{self.llm.model} — {result[:150]}",
            priority=1
        )

        # For transient errors, retry the SAME model once with a short delay
        if error_type in TRANSIENT_ERRORS:
            delay = 2.0 if error_type == ERROR_RATE_LIMIT else 1.0
            await asyncio.sleep(delay)
            retry = await self._call_llm(messages)
            if not isinstance(retry, str) or not retry.startswith("[ERROR]"):
                await self.core.log(f"✅ Retry succeeded for {self.llm.provider}", priority=2)
                return retry

        # Record the failure on the current provider
        model_mgr._record_provider_failure(self.llm.provider, error_type)
        await model_mgr.handle_api_error(result)

        # Walk the fallback chain
        return await self._walk_fallback_chain(messages, error_type)

    async def _walk_fallback_chain(self, messages, original_error_type):
        """
        Try each model in the fallback chain until one succeeds.
        Provider/model is restored to the original state after every attempt.
        """
        model_mgr = self.core.model_manager

        # Save current (user-selected) state — ALWAYS restored at end
        orig_provider = self.llm.provider
        orig_model    = self.llm.model
        orig_key      = self.llm.api_key

        chain = model_mgr.fallback_chain
        last_error = None

        async with model_mgr._fallback_lock:
            # Check shortcut cache — if a fallback worked recently, try it first
            if model_mgr._last_successful_fallback:
                fb_p, fb_m, fb_ts = model_mgr._last_successful_fallback
                if (datetime.now() - fb_ts).total_seconds() < 60:
                    # Try the cached fallback first
                    self.llm.provider = fb_p
                    self.llm.model = fb_m
                    model_mgr._set_api_key(fb_p)
                    try:
                        result = await self._call_llm(messages)
                        if not isinstance(result, str) or not result.startswith("[ERROR]"):
                            model_mgr._record_provider_success(fb_p)
                            model_mgr._last_successful_fallback = (fb_p, fb_m, datetime.now())
                            await self.core.log(
                                f"⚡ Fallback cache hit: {fb_p}/{fb_m} (orig: {orig_provider}/{orig_model})",
                                priority=2
                            )
                            await self.core.relay.emit(2, "model_fallback", {
                                "original": f"{orig_provider}/{orig_model}",
                                "fallback": f"{fb_p}/{fb_m}",
                                "reason": original_error_type,
                            })
                            # Restore original state
                            self.llm.provider = orig_provider
                            self.llm.model = orig_model
                            self.llm.api_key = orig_key
                            return result
                    except Exception as e:
                        await self.core.log(
                            f"Fallback cache miss ({fb_p}/{fb_m}): {type(e).__name__}: {e}",
                            priority=3
                        )

            # Walk the full chain
            for entry in chain:
                provider = entry['provider']
                model    = entry['model']

                # Skip the provider that just failed
                if provider == orig_provider and model == orig_model:
                    continue

                # Skip providers in cooldown
                if not model_mgr._is_provider_available(provider):
                    continue

                # Skip Ollama if offline (avoid 180s timeout on dead server)
                if provider == 'ollama':
                    ollama_mgr = getattr(self.core, 'ollama_manager', None)
                    if ollama_mgr:
                        healthy = await ollama_mgr.health_check()
                        if not healthy:
                            continue

                # Swap to fallback
                self.llm.provider = provider
                self.llm.model = model
                model_mgr._set_api_key(provider)

                await self.core.log(f"🔄 Fallback → trying {provider}/{model}...", priority=2)

                try:
                    result = await self._call_llm(messages)

                    if not isinstance(result, str) or not result.startswith("[ERROR]"):
                        # Success!
                        model_mgr._record_provider_success(provider)
                        model_mgr._last_successful_fallback = (provider, model, datetime.now())
                        await self.core.log(
                            f"✅ Fallback SUCCESS: {provider}/{model} "
                            f"(original: {orig_provider}/{orig_model})",
                            priority=1
                        )
                        await self.core.relay.emit(2, "model_fallback", {
                            "original": f"{orig_provider}/{orig_model}",
                            "fallback": f"{provider}/{model}",
                            "reason": original_error_type,
                        })
                        # Restore original model for next call
                        self.llm.provider = orig_provider
                        self.llm.model = orig_model
                        self.llm.api_key = orig_key
                        return result
                    else:
                        # This fallback also failed
                        fb_error = model_mgr.classify_error(result)
                        model_mgr._record_provider_failure(provider, fb_error)
                        last_error = result

                except Exception as e:
                    last_error = f"[ERROR] Fallback {provider}: {e}"
                    model_mgr._record_provider_failure(provider, "UNKNOWN")

            # All exhausted — restore and return failure
            self.llm.provider = orig_provider
            self.llm.model = orig_model
            self.llm.api_key = orig_key

        total_tried = len(chain) + 1  # +1 for original
        return (
            f"[Galactic] All {total_tried} models in the fallback chain failed. "
            f"Last error: {(last_error or 'unknown')[:200]}. "
            f"Check API keys and service status, or try again in a few minutes."
        )

    async def _call_llm(self, messages):
        """
        Route to the appropriate provider using CURRENT self.llm settings.

        Key behaviour per provider:
          • google    → collapse to prompt+context string → Gemini REST API
          • anthropic → collapse to system+messages → Anthropic Messages API
          • ollama    → pass raw messages[] array directly (preserves conversation structure)
          • nvidia/xai→ collapse to prompt+context string → OpenAI-compat REST API
        """

        # ── Ollama: pre-flight health check ──────────────────────────
        if self.llm.provider == "ollama":
            ollama_mgr = getattr(self.core, 'ollama_manager', None)
            if ollama_mgr:
                healthy = await ollama_mgr.health_check()
                if not healthy:
                    model_mgr = getattr(self.core, 'model_manager', None)
                    if model_mgr:
                        await model_mgr.handle_api_error("Ollama unreachable")
                    return (
                        "[Galactic] ⚠️ Ollama is offline or unreachable at "
                        f"{ollama_mgr.base_url}. Switched to fallback model. "
                        "Run `ollama serve` and I'll reconnect automatically."
                    )

        # ── Context-window trimming ────────────────────────────────────
        # For Ollama: auto-detect from Ollama, but per-model override wins.
        # For other providers: only trim if per-model/global context_window is set.
        ctx_override = self._get_context_window_for_model()
        if self.llm.provider == "ollama":
            ollama_mgr = getattr(self.core, 'ollama_manager', None)
            if ollama_mgr and self.core.config.get('models', {}).get('context_window_trim', True):
                ctx_limit = ctx_override or ollama_mgr.get_context_window(self.llm.model, default=32768)
                # Rough heuristic: 1 token ≈ 4 chars; leave 20% headroom for the response
                char_limit = int(ctx_limit * 4 * 0.8)
                total_chars = sum(len(m.get('content', '')) for m in messages)
                while total_chars > char_limit and len(messages) > 2:
                    messages.pop(1)  # drop oldest non-system message
                    total_chars = sum(len(m.get('content', '')) for m in messages)

        # Pre-process pseudo-providers (e.g. openrouter-frontier -> openrouter)
        orig_provider = self.llm.provider
        base_provider = orig_provider
        if base_provider.startswith("openrouter-"):
            base_provider = "openrouter"
        
        # Temporarily swap for the duration of the call to ensure 
        # internal logic matches the expected API provider
        self.llm.provider = base_provider

        try:
            # ── Route to provider ─────────────────────────────────────────
            if base_provider == "google":
                # Gemini uses a single text blob (system context + user prompt)
                prompt = messages[-1]['content']
                context_str = "\n".join(
                    [f"{m['role']}: {m['content']}" for m in messages[:-1]]
                )
                return await self._call_gemini(prompt, context_str)

            elif base_provider == "deepseek":
                return await self._call_deepseek_messages(messages)

            elif base_provider == "anthropic":
                # Anthropic Messages API: separate system field + messages array
                system_msg = ""
                msg_list = []
                for m in messages:
                    if m["role"] == "system":
                        system_msg = m["content"]
                    else:
                        msg_list.append(m)
                return await self._call_anthropic_messages(system_msg, msg_list)

            elif base_provider == "ollama":
                # Ollama supports the full OpenAI /chat/completions messages array
                return await self._call_openai_compatible_messages(messages)

            elif base_provider == "xai":
                # xAI: collapse to prompt+context (stateless one-shot)
                prompt = messages[-1]['content']
                context_str = "\n".join(
                    [f"{m['role']}: {m['content']}" for m in messages[:-1]]
                )
                return await self._call_openai_compatible(prompt, context_str)

            elif base_provider in ["nvidia", "openai", "groq", "mistral", "cerebras",
                                        "openrouter", "huggingface", "kimi", "zai", "minimax",
                                        "xiaomi", "moonshot", "qwen-portal", "qianfan", "together",
                                        "vllm", "doubao", "byteplus", "cloudflare-ai-gateway", "kilocode"]:
                # OpenAI-compatible providers: pass full messages array for proper multi-turn context
                return await self._call_openai_compatible_messages(messages)

            else:
                return f"[ERROR] Unknown provider: {orig_provider}"
        finally:
            # ALWAYS restore the original provider (including pseudo-suffix)
            self.llm.provider = orig_provider
    
    async def _call_gemini(self, prompt, context):
        """Google Gemini API call."""
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{self.llm.model}:generateContent?key={self.llm.api_key}"
        payload = {"contents": [{"parts": [{"text": f"SYSTEM CONTEXT: {context}\n\nUser: {prompt}"}]}]}
        try:
            self._last_usage = None
            async with httpx.AsyncClient(timeout=60.0, verify=False) as client:
                response = await client.post(url, json=payload)
                data = response.json()
                if 'candidates' not in data or not data['candidates']:
                    return f"[ERROR] Google API: {json.dumps(data)}"
                candidate = data['candidates'][0]
                # Gemini sometimes returns a candidate with finishReason but no content
                # (e.g. safety filter, recitation, or empty response)
                if 'content' not in candidate:
                    reason = candidate.get('finishReason', 'UNKNOWN')
                    return f"[ERROR] Google returned no content (finishReason: {reason}). Try rephrasing."
                # Extract real token counts from Google response
                um = data.get('usageMetadata', {})
                self._last_usage = {
                    "prompt_tokens": um.get('promptTokenCount', 0),
                    "completion_tokens": um.get('candidatesTokenCount', 0),
                }
                return candidate['content']['parts'][0]['text']
        except Exception as e:
            return f"[ERROR] Google: {str(e)}"
    
    async def _call_anthropic(self, prompt, context):
        """
        Anthropic Claude API call using the NATIVE Anthropic Messages API.
        This is NOT OpenAI-compatible — it requires x-api-key + anthropic-version headers
        and uses the /v1/messages endpoint with its own response schema.
        """
        api_key = self.llm.api_key
        if not api_key or api_key == "NONE":
            api_key = self.core.config.get('providers', {}).get('anthropic', {}).get('apiKey', '')
        if not api_key:
            return "[ERROR] Anthropic API key not configured. Set providers.anthropic.apiKey in config.yaml"

        url = "https://api.anthropic.com/v1/messages"
        # OAuth tokens (Claude Pro / Claude Code) require Bearer auth + special beta headers
        if api_key.startswith("sk-ant-oat"):
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "anthropic-version": "2023-06-01",
                "anthropic-beta": "claude-code-20250219,oauth-2025-04-20,fine-grained-tool-streaming-2025-05-14",
                "x-app": "cli",
                "user-agent": "claude-cli/2.1.2 (external, cli)",
            }
        else:
            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            }

        # Anthropic separates system prompt from messages
        payload = {
            "model": self.llm.model,
            "max_tokens": 8096,
            "system": context if context else "You are a helpful AI assistant.",
            "messages": [
                {"role": "user", "content": prompt}
            ]
        }

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()

                # Anthropic response: {"content": [{"type": "text", "text": "..."}], ...}
                if "content" in data and data["content"]:
                    text_blocks = [b["text"] for b in data["content"] if b.get("type") == "text"]
                    return "\n".join(text_blocks) if text_blocks else "[ERROR] Anthropic: Empty response"
                elif "error" in data:
                    err = data["error"]
                    return f"[ERROR] Anthropic ({err.get('type','unknown')}): {err.get('message','Unknown error')}"
                else:
                    return f"[ERROR] Anthropic: Unexpected response: {json.dumps(data)}"
        except Exception as e:
            return f"[ERROR] Anthropic: {str(e)}"

    async def _call_anthropic_messages(self, system_prompt, messages):
        """
        Anthropic Messages API with full conversation history.
        Used by _call_llm() for multi-turn Anthropic conversations (preserves tool-call context).
        """
        api_key = self.llm.api_key
        if not api_key or api_key == "NONE":
            api_key = self.core.config.get('providers', {}).get('anthropic', {}).get('apiKey', '')
        if not api_key:
            return "[ERROR] Anthropic API key not configured. Set providers.anthropic.apiKey in config.yaml"

        url = "https://api.anthropic.com/v1/messages"
        # OAuth tokens (Claude Pro / Claude Code) require Bearer auth + special beta headers
        if api_key.startswith("sk-ant-oat"):
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "anthropic-version": "2023-06-01",
                "anthropic-beta": "claude-code-20250219,oauth-2025-04-20,fine-grained-tool-streaming-2025-05-14",
                "x-app": "cli",
                "user-agent": "claude-cli/2.1.2 (external, cli)",
            }
        else:
            headers = {
                "Content-Type": "application/json",
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
            }

        # Ensure messages alternate user/assistant (Anthropic requirement)
        # Merge consecutive same-role messages
        merged = []
        for m in messages:
            if m.get("role") not in ("user", "assistant"):
                continue
            if merged and merged[-1]["role"] == m["role"]:
                merged[-1]["content"] += "\n" + m["content"]
            else:
                merged.append({"role": m["role"], "content": m["content"]})

        # Must start with user
        if not merged or merged[0]["role"] != "user":
            merged.insert(0, {"role": "user", "content": "(conversation start)"})

        payload = {
            "model": self.llm.model,
            "max_tokens": self._get_max_tokens(default=8192),
            "system": system_prompt if system_prompt else "You are a helpful AI assistant.",
            "messages": merged,
        }

        try:
            self._last_usage = None
            async with httpx.AsyncClient(timeout=120.0) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()
                # Extract real token counts from Anthropic response
                usage = data.get('usage', {})
                self._last_usage = {
                    "prompt_tokens": usage.get('input_tokens', 0),
                    "completion_tokens": usage.get('output_tokens', 0),
                }
                if "content" in data and data["content"]:
                    text_blocks = [b["text"] for b in data["content"] if b.get("type") == "text"]
                    return "\n".join(text_blocks) if text_blocks else "[ERROR] Anthropic: Empty response"
                elif "error" in data:
                    err = data["error"]
                    return f"[ERROR] Anthropic ({err.get('type','unknown')}): {err.get('message','Unknown error')}"
                else:
                    return f"[ERROR] Anthropic: Unexpected response: {json.dumps(data)}"
        except Exception as e:
            return f"[ERROR] Anthropic: {str(e)}"

    def _get_provider_base_url(self, provider):
        """Return the base URL for an OpenAI-compatible provider from config."""
        if provider and provider.startswith("openrouter-"):
            provider = "openrouter"
        providers_cfg = self.core.config.get('providers', {})
        default_urls = {
            "openai":       "https://api.openai.com/v1",
            "groq":         "https://api.groq.com/openai/v1",
            "mistral":      "https://api.mistral.ai/v1",
            "cerebras":     "https://api.cerebras.ai/v1",
            "openrouter":   "https://openrouter.ai/api/v1",
            "huggingface":  "https://router.huggingface.co/v1",
            "kimi":         "https://api.kimi.com/v1",
            "zai":          "https://api.z.ai/api/paas/v4",
            "minimax":      "https://api.minimax.io/v1",
            "nvidia":       "https://integrate.api.nvidia.com/v1",
            "xai":          "https://api.x.ai/v1",
            "ollama":       "http://127.0.0.1:11434/v1",
        }
        configured = providers_cfg.get(provider, {}).get('baseUrl', '')
        base = configured or default_urls.get(provider, '')
        # Normalize Ollama URL — ensure it ends with /v1
        if provider == "ollama" and not base.rstrip('/').endswith('/v1'):
            base = base.rstrip('/') + '/v1'
        return base.rstrip('/')

    def _get_provider_api_key(self, provider):
        """Return the API key for a provider, falling back to config providers section."""
        if provider and provider.startswith("openrouter-"):
            provider = "openrouter"
        # Use the live llm.api_key if it's set and not placeholder
        key = self.llm.api_key
        if key and key not in ("NONE", ""):
            return key
        providers_cfg = self.core.config.get('providers', {})
        provider_cfg = providers_cfg.get(provider, {})

        # NVIDIA: prefer the unified apiKey (works for all 500+ models on build.nvidia.com).
        # Fall back to the legacy per-model keys: sub-dict for backwards compatibility
        # with installs that have the old multi-key format.
        if provider == 'nvidia':
            # 1. Unified single key (new setup wizard path)
            single_key = provider_cfg.get('apiKey', '') or provider_cfg.get('api_key', '')
            if single_key:
                return single_key
            # 2. Legacy keys: sub-dict — match nickname against active model name
            model_str = (getattr(self.llm, 'model', '') or '').lower()
            nvidia_keys = provider_cfg.get('keys', {}) or {}
            for nickname, nvapi_key in nvidia_keys.items():
                if nvapi_key and nickname.lower() in model_str:
                    return nvapi_key
            # 3. Fall back to first non-empty legacy key
            for nvapi_key in nvidia_keys.values():
                if nvapi_key:
                    return nvapi_key

        return provider_cfg.get('apiKey', '') or provider_cfg.get('api_key', '')

    def _get_model_override(self, key, default=None):
        """Return a per-model override value for the active model, falling back to global config."""
        model_id = getattr(self.llm, 'model', '') or ''
        overrides = self.core.config.get('model_overrides', {}) or {}
        # Check exact model match first
        if model_id in overrides and key in (overrides[model_id] or {}):
            try:
                val = int(overrides[model_id][key])
                if val > 0:
                    return val
            except (TypeError, ValueError):
                pass
        # Check aliases — if model_id matches an alias value, also check by alias name
        aliases = self.core.config.get('aliases', {}) or {}
        for alias, aliased_model in aliases.items():
            # aliased_model might be "provider/model" form; strip provider prefix
            stripped = aliased_model.split('/', 1)[-1] if '/' in aliased_model else aliased_model
            if (aliased_model == model_id or stripped == model_id) and alias in overrides:
                try:
                    val = int((overrides[alias] or {}).get(key, 0))
                    if val > 0:
                        return val
                except (TypeError, ValueError):
                    pass
        return default

    def _get_max_tokens(self, default=None):
        """Return max_tokens: per-model override first, then global config, then default."""
        # Per-model override
        per_model = self._get_model_override('max_tokens')
        if per_model:
            return per_model
        # Global config
        val = self.core.config.get('models', {}).get('max_tokens', 0)
        try:
            val = int(val)
        except (TypeError, ValueError):
            val = 0
        return val if val > 0 else default

    def _get_context_window_for_model(self, default=None):
        """Return context_window: per-model override first, then global config, then default."""
        per_model = self._get_model_override('context_window')
        if per_model:
            return per_model
        val = self.core.config.get('models', {}).get('context_window', 0)
        try:
            val = int(val)
        except (TypeError, ValueError):
            val = 0
        return val if val > 0 else default

    async def _call_openai_compatible(self, prompt, context):
        """OpenAI-compatible API call (NVIDIA, XAI, Ollama). All URLs are config-driven."""

        # FLUX models are image-generation only — they don't support chat/completions.
        # Auto-invoke generate_image with the user's prompt instead of erroring.
        if self.llm.provider == "nvidia" and "flux" in self.llm.model.lower():
            return await self.tool_generate_image({
                "prompt": prompt,
                "model": self.llm.model,
            })

        url = f"{self._get_provider_base_url(self.llm.provider)}/chat/completions"

        # Ollama doesn't need auth header
        headers = {"Content-Type": "application/json"}
        if self.llm.provider not in ("ollama",):
            headers["Authorization"] = f"Bearer {self._get_provider_api_key(self.llm.provider)}"

        # Use streaming for Ollama when configured (faster feel on local hardware)
        use_streaming = (
            self.llm.provider == "ollama"
            and self.core.config.get('models', {}).get('streaming', True)
        )
        if use_streaming:
            return await self._call_openai_compatible_streaming(prompt, context, url, headers)

        payload = {
            "model": self.llm.model,
            "messages": [
                {"role": "system", "content": context},
                {"role": "user", "content": prompt}
            ]
        }
        max_tokens = self._get_max_tokens()
        if max_tokens:
            payload["max_tokens"] = max_tokens

        # Inject thinking/reasoning params for NVIDIA models that require them
        if self.llm.provider == "nvidia":
            extra = _NVIDIA_THINKING_MODELS.get(self.llm.model, {})
            if extra:
                payload.update(extra)

        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False) as client:
                response = await client.post(url, headers=headers, json=payload)
                data = response.json()
                if 'choices' not in data:
                    return f"[ERROR] {self.llm.provider}: {json.dumps(data)}"
                msg = data['choices'][0]['message']
                content = (msg.get('content') or '').strip()
                reasoning = (msg.get('reasoning_content') or '').strip()
                # Handle native tool_calls
                if not content and not reasoning and msg.get('tool_calls'):
                    tc_list = msg['tool_calls']
                    if tc_list:
                        fn = tc_list[0].get('function', {})
                        fn_name = fn.get('name', '')
                        fn_args_str = fn.get('arguments', '{}')
                        try:
                            fn_args = json.loads(fn_args_str) if fn_args_str else {}
                        except json.JSONDecodeError:
                            fn_args = {}
                        return json.dumps({"tool": fn_name, "args": fn_args})
                if content:
                    return content
                elif reasoning:
                    return f"[Reasoning]\n{reasoning}"
                else:
                    return f"[ERROR] {self.llm.provider}: empty content in response"
        except Exception as e:
            return f"[ERROR] {self.llm.provider}: {str(e)}"

    async def _call_openai_compatible_streaming(self, prompt, context, url, headers):
        """Streaming variant – returns full text but streams internally for real-time web UI updates."""
        payload = {
            "model": self.llm.model,
            "messages": [
                {"role": "system", "content": context},
                {"role": "user", "content": prompt}
            ],
            "stream": True
        }
        full_response = []
        try:
            async with httpx.AsyncClient(timeout=120.0, verify=False) as client:
                async with client.stream("POST", url, headers=headers, json=payload) as response:
                    token_buf = []
                    async for line in response.aiter_lines():
                        if not line.startswith("data: "):
                            continue
                        data_str = line[6:].strip()
                        if data_str == "[DONE]":
                            break
                        try:
                            chunk = json.loads(data_str)
                            delta = chunk.get('choices', [{}])[0].get('delta', {}).get('content', '')
                            if delta:
                                full_response.append(delta)
                                token_buf.append(delta)
                                if len(token_buf) >= 8:
                                    await self.core.relay.emit(3, "stream_chunk", "".join(token_buf))
                                    token_buf = []
                                    await asyncio.sleep(0)
                        except json.JSONDecodeError:
                            continue
                    if token_buf:
                        await self.core.relay.emit(3, "stream_chunk", "".join(token_buf))
            res = "".join(full_response)
            if not res.strip():
                return f"[ERROR] {self.llm.provider}: empty stream content"
            return res
        except Exception as e:
            return f"[ERROR] {self.llm.provider} (streaming): {str(e)}"

    async def _fetch_openrouter_generation_cost(self, generation_id):
        """Query OpenRouter's generation API for the actual cost charged.

        Returns the total cost as a float, or None on failure.
        This is a lightweight GET request (no streaming, fast timeout).
        """
        api_key = self._get_provider_api_key('openrouter')
        if not api_key:
            return None
        try:
            url = f"https://openrouter.ai/api/v1/generation?id={generation_id}"
            async with httpx.AsyncClient(timeout=5.0) as client:
                resp = await client.get(url, headers={
                    "Authorization": f"Bearer {api_key}",
                })
                if resp.status_code == 200:
                    data = resp.json().get('data', {})
                    cost = data.get('total_cost')
                    if cost is not None:
                        return float(cost)
        except Exception:
            pass  # Non-fatal — fall back to estimated cost
        return None

    async def _call_openai_compatible_messages(self, messages):
        """
        OpenAI-compatible call that passes the FULL messages array.

        Used for: Ollama (local), OpenAI, Groq, Mistral, Cerebras, OpenRouter,
                  HuggingFace, Kimi, ZAI/GLM, MiniMax — any provider using
                  the standard /chat/completions messages array format.

        Key features:
          • Passes messages[] directly (preserves multi-turn conversation context)
          • Supports streaming for Ollama and OpenAI-compatible providers
          • Reads base URL and API key from config for all providers
          • Injects max_tokens if configured
        """
        provider = self.llm.provider
        self._last_usage = None

        # FLUX models are image-generation only — auto-invoke generate_image
        if provider == "nvidia" and "flux" in self.llm.model.lower():
            prompt = messages[-1]['content'] if messages else ''
            return await self.tool_generate_image({
                "prompt": prompt,
                "model": self.llm.model,
            })

        url = f"{self._get_provider_base_url(provider)}/chat/completions"

        headers = {"Content-Type": "application/json"}
        # Ollama doesn't use auth; all other providers use Bearer token
        if provider != "ollama":
            api_key = self._get_provider_api_key(provider)
            if api_key:
                headers["Authorization"] = f"Bearer {api_key}"
            # OpenRouter requires an extra header
            if provider == "openrouter":
                headers["HTTP-Referer"] = "https://galactic-ai.local"
                headers["X-Title"] = "Galactic AI"

        use_streaming = (
            provider in ("ollama", "openai", "groq", "mistral", "cerebras", "openrouter", "nvidia")
            and self.core.config.get('models', {}).get('streaming', True)
            and not (provider == "nvidia" and self.llm.model in _NVIDIA_NO_STREAM)
            and not (provider == "openrouter" and self.llm.model in _OPENROUTER_NO_STREAM)
        )

        max_tokens = self._get_max_tokens()

        if use_streaming:
            payload = {
                "model": self.llm.model,
                "messages": messages,
                "stream": True,
                "stream_options": {"include_usage": True},
            }
            # Ollama benefits from explicit temperature in options
            if provider == "ollama":
                payload["options"] = {"temperature": 0.3}
            # NVIDIA thinking/reasoning models need extra body params
            if provider == "nvidia":
                extra = _NVIDIA_THINKING_MODELS.get(self.llm.model, {})
                if extra:
                    payload.update(extra)
            if max_tokens:
                payload["max_tokens"] = max_tokens
            full_response = []
            try:
                # Granular timeout: fast connect (30s) but long read (600s) for
                # large models (Qwen 397B, GLM5 744B) with slow first-token latency
                _timeout = httpx.Timeout(connect=30.0, read=600.0, write=30.0, pool=30.0)
                async with httpx.AsyncClient(timeout=_timeout, verify=False) as client:
                    async with client.stream("POST", url, headers=headers, json=payload) as response:
                        # Check HTTP status before parsing SSE stream
                        if response.status_code != 200:
                            body = await response.aread()
                            try:
                                err_data = json.loads(body)
                                err_msg = err_data.get('error', {}).get('message', '') or err_data.get('detail', '') or body.decode()[:500]
                            except Exception:
                                err_msg = body.decode('utf-8', errors='replace')[:500]
                            return f"[ERROR] {provider} HTTP {response.status_code}: {err_msg}"
                        token_buf = []
                        # Accumulator for streamed native tool_calls (arguments arrive
                        # incrementally across multiple chunks)
                        _tc_name = ''
                        _tc_args_parts = []
                        async for line in response.aiter_lines():
                            if not line.startswith("data: "):
                                continue
                            data_str = line[6:].strip()
                            if data_str == "[DONE]":
                                break
                            try:
                                chunk = json.loads(data_str)
                                choices = chunk.get('choices', [])
                                if not choices:
                                    # Empty choices — could be error, heartbeat, or model loading
                                    # Check for error payload
                                    if 'error' in chunk:
                                        err_msg = chunk['error'].get('message', str(chunk['error']))
                                        return f"[ERROR] {provider}: {err_msg}"
                                    continue
                                choice = choices[0]
                                if not choice:
                                    continue
                                    
                                delta_obj = choice.get('delta', {})
                                if not delta_obj:
                                    continue
                                    
                                delta = delta_obj.get('content', '') or ''
                                # NVIDIA thinking models may also stream reasoning_content
                                if not delta:
                                    delta = delta_obj.get('reasoning_content', '') or ''

                                # ── Accumulate native tool_calls from delta ──
                                tc_deltas = delta_obj.get('tool_calls', [])
                                for tc in (tc_deltas or []):
                                    if not tc: continue
                                    fn = tc.get('function', {})
                                    if not fn: continue
                                    if fn.get('name'):
                                        _tc_name = fn['name']
                                    if fn.get('arguments'):
                                        _tc_args_parts.append(fn['arguments'])

                                # ── Capture finish_reason for diagnostics ──
                                finish_reason = choice.get('finish_reason')
                                if finish_reason and finish_reason not in ('stop', None):
                                    await self.core.log(
                                        f"⚠️ Stream finish_reason={finish_reason} "
                                        f"(provider={provider}, model={self.llm.model})",
                                        priority=2
                                    )

                                if delta:
                                    full_response.append(delta)
                                    token_buf.append(delta)
                                    # Batch emit every 8 tokens to reduce event loop pressure
                                    if len(token_buf) >= 8:
                                        await self.core.relay.emit(3, "stream_chunk", "".join(token_buf))
                                        token_buf = []
                                        await asyncio.sleep(0)  # yield to other tasks (typing, etc.)

                                # Capture usage from final streaming chunk (OpenAI/OpenRouter)
                                usage = chunk.get('usage')
                                if usage:
                                    self._last_usage = {
                                        "prompt_tokens": usage.get('prompt_tokens', 0),
                                        "completion_tokens": usage.get('completion_tokens', 0),
                                    }
                                # Capture OpenRouter generation ID for actual cost lookup
                                if 'id' in chunk and provider == 'openrouter':
                                    self._last_generation_id = chunk['id']
                            except json.JSONDecodeError:
                                continue
                        # Flush remaining buffer
                        if token_buf:
                            await self.core.relay.emit(3, "stream_chunk", "".join(token_buf))
                        # ── Flush accumulated native tool_call ──
                        if _tc_name and not full_response:
                            args_str = "".join(_tc_args_parts) or '{}'
                            try:
                                fn_args = json.loads(args_str)
                            except json.JSONDecodeError:
                                fn_args = {}
                            synthesized = json.dumps({"tool": _tc_name, "args": fn_args})
                            full_response.append(synthesized)
                            await self.core.log(
                                f"🔧 Native tool_call intercepted (stream): "
                                f"{_tc_name} → converted to text",
                                priority=2
                            )
                result = "".join(full_response)
                # ── Diagnostic: log when streaming produced empty result ──
                if not result.strip():
                    await self.core.log(
                        f"⚠️ [DIAG] Streaming returned empty content "
                        f"(provider={provider}, model={self.llm.model}, "
                        f"chunks_processed={len(full_response)})",
                        priority=1
                    )
                if not result.strip():
                    # Streaming returned no content — fall through to non-streaming for all providers.
                    await self.core.log(
                        f"⚠️ {provider}/{self.llm.model} streaming returned empty — "
                        f"retrying non-streaming",
                        priority=2
                    )
                else:
                    return result
            except Exception as e:
                if provider == "nvidia":
                    # Streaming failed — fall through to non-streaming as fallback
                    await self.core.log(
                        f"⚠️ {provider}/{self.llm.model} streaming error: {e} — "
                        f"retrying non-streaming",
                        priority=2
                    )
                else:
                    return f"[ERROR] {provider} (streaming): {str(e)}"

        # ── Non-streaming path (also serves as NVIDIA streaming fallback) ──
        payload = {
            "model": self.llm.model,
            "messages": messages,
            "stream": False,
        }
        if provider == "ollama":
            payload["options"] = {"temperature": 0.3}
        # NVIDIA thinking/reasoning models need extra body params
        if provider == "nvidia":
            extra = _NVIDIA_THINKING_MODELS.get(self.llm.model, {})
            if extra:
                payload.update(extra)
        if max_tokens:
            payload["max_tokens"] = max_tokens
        # NVIDIA NIM is serverless — large models (397B, 744B) get unloaded when
        # idle and cold-start can exceed NVIDIA's 5-min gateway timeout (HTTP 504).
        # Retry up to 2 times on 504 to ride out the cold-start window.
        _max_retries = 2 if provider == "nvidia" else 0
        for _attempt in range(_max_retries + 1):
            try:
                _timeout = httpx.Timeout(connect=30.0, read=600.0, write=30.0, pool=30.0)
                async with httpx.AsyncClient(timeout=_timeout, verify=False) as client:
                    response = await client.post(url, headers=headers, json=payload)
                    # Check HTTP status before parsing JSON
                    if response.status_code != 200:
                        body_text = response.text[:500]
                        # NVIDIA 504 = model cold-starting — retry
                        if response.status_code in (502, 503, 504) and _attempt < _max_retries:
                            await self.core.log(
                                f"⏳ NVIDIA model loading (HTTP {response.status_code}) — "
                                f"retry {_attempt + 1}/{_max_retries}, waiting for cold-start...",
                                priority=2
                            )
                            await asyncio.sleep(10)  # brief pause before retry
                            continue
                        try:
                            err_data = json.loads(body_text)
                            err_msg = (err_data.get('error', {}).get('message', '')
                                       or err_data.get('error', '')
                                       or err_data.get('detail', '')
                                       or body_text)
                        except Exception:
                            err_msg = body_text or f"HTTP {response.status_code} (empty body)"
                        if response.status_code in (502, 503, 504):
                            err_msg = (f"NVIDIA model cold-start timeout after "
                                       f"{_max_retries + 1} attempts — model may be "
                                       f"unavailable. Try again in a few minutes.")
                        return f"[ERROR] {provider} HTTP {response.status_code}: {err_msg}"
                    # Safe JSON parse — guard against empty body
                    body_text = response.text.strip()
                    if not body_text:
                        return f"[ERROR] {provider}: empty response body (HTTP 200)"
                    try:
                        data = json.loads(body_text)
                    except json.JSONDecodeError as je:
                        return f"[ERROR] {provider}: invalid JSON — {je} — body: {body_text[:200]}"
                    # Extract real token counts from OpenAI-compatible response
                    usage = data.get('usage', {})
                    if usage:
                        self._last_usage = {
                            "prompt_tokens": usage.get('prompt_tokens', 0),
                            "completion_tokens": usage.get('completion_tokens', 0),
                        }
                    # Capture OpenRouter generation ID for actual cost lookup
                    if provider == 'openrouter' and 'id' in data:
                        self._last_generation_id = data['id']
                    if 'choices' not in data:
                        return f"[ERROR] {provider}: {json.dumps(data)[:500]}"
                    msg = data['choices'][0]['message']
                    content = (msg.get('content') or '').strip()
                    reasoning = (msg.get('reasoning_content') or msg.get('reasoning') or '').strip()
                    refusal = (msg.get('refusal') or '').strip()
                    # Handle native tool_calls (Gemini/GPT via OpenRouter may
                    # use this instead of putting JSON in content text)
                    if not content and not reasoning and not refusal and msg.get('tool_calls'):
                        tc_list = msg['tool_calls']
                        if tc_list:
                            fn = tc_list[0].get('function', {})
                            fn_name = fn.get('name', '')
                            fn_args_str = fn.get('arguments', '{}')
                            try:
                                fn_args = json.loads(fn_args_str) if fn_args_str else {}
                            except json.JSONDecodeError:
                                fn_args = {}
                            await self.core.log(
                                f"🔧 Native tool_call intercepted (non-stream): {fn_name}",
                                priority=3
                            )
                            return json.dumps({"tool": fn_name, "args": fn_args})
                    if content:
                        return content
                    elif reasoning:
                        return f"[Reasoning]\n{reasoning}"
                    elif refusal:
                        return f"[Refusal] {refusal}"
                    else:
                        return f"[ERROR] {provider}: empty content in response (possible safety filter or invalid model name)"
            except Exception as e:
                if _attempt < _max_retries and provider == "nvidia":
                    await self.core.log(
                        f"⏳ NVIDIA request error ({e.__class__.__name__}) — "
                        f"retry {_attempt + 1}/{_max_retries}...",
                        priority=2
                    )
                    await asyncio.sleep(10)
                    continue
                return f"[ERROR] {provider}: {str(e)}"


    # ═══════════════════════════════════════════════════════════════════════════
    # ── v0.8.0 NEW TOOLS — Ultimate Automation Suite ──────────────────────────
    # ═══════════════════════════════════════════════════════════════════════════

    async def tool_generate_image_sd35(self, args):
        """Generate an image using Stable Diffusion 3.5 Large via NVIDIA NIM."""
        import base64 as _b64, time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_image_sd35 requires a 'prompt' argument."
        negative_prompt = args.get('negative_prompt', '')
        width    = int(args.get('width', 1024))
        height   = int(args.get('height', 1024))
        steps    = int(args.get('steps', 40))
        cfg_scale = float(args.get('cfg_scale', 5.0))
        seed     = int(args.get('seed', 0))

        nvidia_cfg = self.core.config.get('providers', {}).get('nvidia', {})
        nvidia_key = nvidia_cfg.get('apiKey', '')
        if not nvidia_key:
            return "[ERROR] No nvidia.apiKey found in config.yaml"

        url = "https://ai.api.nvidia.com/v1/genai/stabilityai/stable-diffusion-3.5-large"
        headers = {
            "Authorization": f"Bearer {nvidia_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }
        payload = {
            "prompt": prompt,
            "mode": "base",
            "width": width,
            "height": height,
            "steps": steps,
            "cfg_scale": cfg_scale,
            "seed": seed,
        }
        if negative_prompt:
            payload["negative_prompt"] = negative_prompt

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                r = await client.post(url, headers=headers, json=payload)
                if r.status_code == 401:
                    return f"[ERROR] NVIDIA SD3.5 401 Unauthorized — check your apiKey in config.yaml"
                if r.status_code == 500:
                    return "[ERROR] NVIDIA SD3.5 HTTP 500 — inference server error. Try again in a few minutes."
                if r.status_code != 200:
                    return f"[ERROR] NVIDIA SD3.5 HTTP {r.status_code}: {r.text[:500]}"
                data = r.json()

            artifact = data.get('artifacts', [{}])[0]
            finish = artifact.get('finishReason', '')
            if finish == 'CONTENT_FILTERED':
                return "⚠️ Image blocked by content filter. Try a different prompt."
            b64 = artifact.get('base64', '')
            if not b64:
                return f"[ERROR] SD3.5 generation failed: {json.dumps(data)}"

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'sd35')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"sd35_{int(_time.time())}.jpg"
            path = os.path.join(img_subdir, fname)
            with open(path, 'wb') as f:
                f.write(_b64.b64decode(b64))
            self.last_image_file = path
            return f"✅ SD3.5 image generated: {path}\nModel: stable-diffusion-3.5-large\nPrompt: {prompt}"
        except Exception as e:
            return f"[ERROR] generate_image_sd35: {e}"

    async def tool_generate_image_imagen(self, args):
        """Generate an image using Google Imagen 4 via the google-genai SDK."""
        import time as _time
        prompt       = args.get('prompt', '')
        model        = args.get('model', 'imagen-4')
        aspect_ratio = args.get('aspect_ratio', '1:1')
        n_images     = int(args.get('number_of_images', 1))

        if not prompt:
            return "[ERROR] generate_image_imagen: 'prompt' is required."

        # Map user-friendly names to SDK model identifiers
        model_map = {
            'imagen-4':       'imagen-4.0-generate-001',
            'imagen-4-ultra': 'imagen-4.0-ultra-generate-001',
            'imagen-4-fast':  'imagen-4.0-fast-generate-001',
        }
        sdk_model = model_map.get(model, 'imagen-4.0-generate-001')

        google_cfg = self.core.config.get('providers', {}).get('google', {})
        api_key = google_cfg.get('apiKey', '')
        if not api_key:
            return "[ERROR] generate_image_imagen: No Google API key configured. Add it at providers.google.apiKey in config.yaml."

        try:
            from google import genai
            from google.genai import types

            client = genai.Client(api_key=api_key)

            result = client.models.generate_images(
                model=sdk_model,
                prompt=prompt,
                config=types.GenerateImagesConfig(
                    number_of_images=max(1, min(4, n_images)),
                    aspect_ratio=aspect_ratio,
                    safety_filter_level="BLOCK_LOW_AND_ABOVE",
                    person_generation="ALLOW_ADULT",
                ),
            )

            if not result.generated_images:
                return "[ERROR] Imagen returned no images. Check your prompt for policy violations."

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'imagen')
            os.makedirs(img_subdir, exist_ok=True)

            saved = []
            for i, gen_img in enumerate(result.generated_images):
                fname = f"imagen_{int(_time.time())}_{i}.png"
                path = os.path.join(img_subdir, fname)
                # Use .save() if available, else write raw bytes
                if hasattr(gen_img.image, 'save'):
                    gen_img.image.save(path)
                else:
                    with open(path, 'wb') as f:
                        f.write(gen_img.image.image_bytes)
                saved.append(path)

            # Deliver the first image inline via Control Deck / Telegram
            self.last_image_file = saved[0]
            paths_str = '\n'.join(f"  {p}" for p in saved)
            return f"✅ Imagen image(s) generated ({model}):\n{paths_str}\nPrompt: {prompt}"
        except ImportError:
            return "[ERROR] google-genai not installed. Run: pip install google-genai"
        except Exception as e:
            return f"[ERROR] generate_image_imagen: {e}"

    async def tool_generate_video(self, args):
        """Generate a video using Google Veo via the google-genai SDK."""
        import time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_video requires a 'prompt' argument."

        video_cfg = self.core.config.get('video', {}).get('google', {})
        duration = str(args.get('duration', video_cfg.get('default_duration', 8)))
        aspect_ratio = args.get('aspect_ratio', video_cfg.get('default_aspect_ratio', '16:9'))
        resolution = args.get('resolution', video_cfg.get('default_resolution', '1080p'))
        negative_prompt = args.get('negative_prompt', '')
        model_name = video_cfg.get('model', 'veo-3.1')

        model_map = {
            'veo-2': 'veo-2-generate-preview',
            'veo-3': 'veo-3.0-generate-preview',
            'veo-3.1': 'veo-3.1-generate-preview',
        }
        model_id = model_map.get(model_name, model_name)

        google_key = self.core.config.get('providers', {}).get('google', {}).get('apiKey', '')
        if not google_key:
            return "[ERROR] No google.apiKey found in config.yaml"

        try:
            from google import genai
            from google.genai import types

            client = genai.Client(api_key=google_key)

            await self.core.log(f"🎬 Generating video with {model_id}...", priority=2)

            gen_config = types.GenerateVideosConfig(
                aspect_ratio=aspect_ratio,
                resolution=resolution,
                duration_seconds=duration,
            )
            if negative_prompt:
                gen_config.negative_prompt = negative_prompt

            operation = client.models.generate_videos(
                model=model_id,
                prompt=prompt,
                config=gen_config,
            )

            poll_count = 0
            while not operation.done:
                poll_count += 1
                if poll_count % 6 == 0:
                    await self.core.log(
                        f"🎬 Video still generating... ({poll_count * 10}s elapsed)",
                        priority=3
                    )
                await asyncio.sleep(10)
                operation = client.operations.get(operation)

            if not operation.response or not operation.response.generated_videos:
                return "[ERROR] Video generation returned no results."

            video = operation.response.generated_videos[0]
            client.files.download(file=video.video)

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            vid_subdir = os.path.join(images_dir, 'video')
            os.makedirs(vid_subdir, exist_ok=True)
            fname = f"veo_{int(_time.time())}.mp4"
            path = os.path.join(vid_subdir, fname)
            video.video.save(path)

            self.last_video_file = path
            return (
                f"✅ Video generated: {path}\n"
                f"Model: {model_id}\n"
                f"Duration: {duration}s | Resolution: {resolution} | Aspect: {aspect_ratio}\n"
                f"Prompt: {prompt}"
            )
        except Exception as e:
            return f"[ERROR] generate_video: {e}"

    async def tool_generate_video_from_image(self, args):
        """Animate a still image into video using Google Veo."""
        import time as _time
        prompt = args.get('prompt', '')
        image_path = args.get('image_path', '')
        if not prompt:
            return "[ERROR] generate_video_from_image requires a 'prompt' argument."
        if not image_path or not os.path.exists(image_path):
            return f"[ERROR] Image not found: {image_path}"

        video_cfg = self.core.config.get('video', {}).get('google', {})
        duration = str(args.get('duration', video_cfg.get('default_duration', 8)))
        aspect_ratio = args.get('aspect_ratio', video_cfg.get('default_aspect_ratio', '16:9'))
        model_name = video_cfg.get('model', 'veo-3.1')

        model_map = {
            'veo-2': 'veo-2-generate-preview',
            'veo-3': 'veo-3.0-generate-preview',
            'veo-3.1': 'veo-3.1-generate-preview',
        }
        model_id = model_map.get(model_name, model_name)

        google_key = self.core.config.get('providers', {}).get('google', {}).get('apiKey', '')
        if not google_key:
            return "[ERROR] No google.apiKey found in config.yaml"

        try:
            from google import genai
            from google.genai import types
            from PIL import Image as _PILImage

            client = genai.Client(api_key=google_key)

            await self.core.log(f"🎬 Animating image to video with {model_id}...", priority=2)

            img = _PILImage.open(image_path)

            operation = client.models.generate_videos(
                model=model_id,
                prompt=prompt,
                image=img,
                config=types.GenerateVideosConfig(
                    aspect_ratio=aspect_ratio,
                    duration_seconds=duration,
                ),
            )

            poll_count = 0
            while not operation.done:
                poll_count += 1
                if poll_count % 6 == 0:
                    await self.core.log(
                        f"🎬 Video still generating... ({poll_count * 10}s elapsed)",
                        priority=3
                    )
                await asyncio.sleep(10)
                operation = client.operations.get(operation)

            if not operation.response or not operation.response.generated_videos:
                return "[ERROR] Video generation returned no results."

            video = operation.response.generated_videos[0]
            client.files.download(file=video.video)

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            vid_subdir = os.path.join(images_dir, 'video')
            os.makedirs(vid_subdir, exist_ok=True)
            fname = f"veo_{int(_time.time())}.mp4"
            path = os.path.join(vid_subdir, fname)
            video.video.save(path)

            self.last_video_file = path
            return (
                f"✅ Image animated to video: {path}\n"
                f"Model: {model_id}\n"
                f"Source: {image_path}\n"
                f"Duration: {duration}s | Aspect: {aspect_ratio}\n"
                f"Prompt: {prompt}"
            )
        except Exception as e:
            return f"[ERROR] generate_video_from_image: {e}"

    async def tool_list_dir(self, args):
        """List directory contents with sizes and dates."""
        import glob as _glob, stat as _stat
        from datetime import datetime as _dt
        path    = args.get('path', '.') or '.'
        pattern = args.get('pattern', '*')
        recurse = bool(args.get('recurse', False))
        try:
            base = os.path.abspath(path)
            if not os.path.isdir(base):
                return (
                    f"[ERROR] list_dir FAILED — path does not exist or is not a directory.\n"
                    f"  Requested: {path!r}\n"
                    f"  Resolved to: {base!r}\n"
                    f"STOP — do not guess or invent file listings. Report this error to the user "
                    f"and ask them for the correct absolute path."
                )
            search = os.path.join(base, '**', pattern) if recurse else os.path.join(base, pattern)
            
            # Run blocking glob in a thread pool
            loop = asyncio.get_running_loop()
            entries = await loop.run_in_executor(None, lambda: _glob.glob(search, recursive=recurse))
            
            if not entries:
                return f"No files match '{pattern}' in {base}"
            lines = [f"{'TYPE':<5} {'SIZE':>10}  {'MODIFIED':<20}  NAME"]
            lines.append('-' * 70)
            for e in sorted(entries)[:500]:
                try:
                    st   = os.stat(e)
                    kind = 'DIR ' if os.path.isdir(e) else 'FILE'
                    size = '' if os.path.isdir(e) else f"{st.st_size:,}"
                    mtime = _dt.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    name = os.path.relpath(e, base)
                    lines.append(f"{kind:<5} {size:>10}  {mtime:<20}  {name}")
                except Exception:
                    pass
            if len(entries) > 500:
                lines.append(f"... (showing 500 of {len(entries)} matches)")
            return '\n'.join(lines)
        except Exception as e:
            return f"[ERROR] list_dir: {e}"

    async def tool_find_files(self, args):
        """Find files matching a glob pattern recursively."""
        import glob as _glob
        path    = args.get('path', '.') or '.'
        pattern = args.get('pattern', '*')
        limit   = int(args.get('limit', 100))
        try:
            base = os.path.abspath(path)
            if '**' in pattern or '/' in pattern or '\\' in pattern:
                search = os.path.join(base, pattern)
            else:
                search = os.path.join(base, '**', pattern)
            
            # Run blocking glob in a thread pool
            loop = asyncio.get_running_loop()
            results = await loop.run_in_executor(None, lambda: _glob.glob(search, recursive=True))
            
            results = [os.path.relpath(r, base) for r in sorted(results)]
            total = len(results)
            results = results[:limit]
            if not results:
                return f"No files found matching '{pattern}' under {base}"
            out = '\n'.join(results)
            if total > limit:
                out += f"\n... ({total - limit} more results — increase limit to see all)"
            return f"Found {total} file(s):\n{out}"
        except Exception as e:
            return f"[ERROR] find_files: {e}"

    async def tool_hash_file(self, args):
        """Compute a file's hash checksum."""
        import hashlib as _hl
        path = args.get('path', '')
        algo = args.get('algorithm', 'sha256').lower()
        algos = {'sha256': _hl.sha256, 'md5': _hl.md5, 'sha1': _hl.sha1}
        if algo not in algos:
            return f"[ERROR] Unsupported algorithm '{algo}'. Choose: sha256, md5, sha1"
        try:
            h = algos[algo]()
            with open(path, 'rb') as f:
                for chunk in iter(lambda: f.read(65536), b''):
                    h.update(chunk)
            size = os.path.getsize(path)
            return f"{algo.upper()}: {h.hexdigest()}\nFile: {path}\nSize: {size:,} bytes"
        except Exception as e:
            return f"[ERROR] hash_file: {e}"

    async def tool_diff_files(self, args):
        """Show unified diff between two files or a file and a string."""
        import difflib as _diff
        path_a  = args.get('path_a', '')
        path_b  = args.get('path_b', '')
        text_b  = args.get('text_b', None)
        context = int(args.get('context', 3))
        try:
            with open(path_a, 'r', encoding='utf-8', errors='replace') as f:
                lines_a = f.readlines()
            if path_b:
                with open(path_b, 'r', encoding='utf-8', errors='replace') as f:
                    lines_b = f.readlines()
                label_b = path_b
            elif text_b is not None:
                lines_b = [l if l.endswith('\n') else l + '\n' for l in text_b.splitlines()]
                label_b = '<new content>'
            else:
                return "[ERROR] Provide path_b or text_b to compare against."
            diff = list(_diff.unified_diff(lines_a, lines_b, fromfile=path_a, tofile=label_b, n=context))
            if not diff:
                return "✅ Files are identical — no differences found."
            return ''.join(diff)
        except Exception as e:
            return f"[ERROR] diff_files: {e}"

    async def tool_zip_create(self, args):
        """Create a ZIP archive from a file or directory."""
        import zipfile as _zip, time as _time
        source = args.get('source', '')
        dest   = args.get('destination', '') or source.rstrip('/\\') + '.zip'
        try:
            source = os.path.abspath(source)
            dest   = os.path.abspath(dest)
            if not os.path.exists(source):
                return f"[ERROR] Source does not exist: {source}"
            with _zip.ZipFile(dest, 'w', compression=_zip.ZIP_DEFLATED) as zf:
                if os.path.isdir(source):
                    for root, dirs, files in os.walk(source):
                        for file in files:
                            fp = os.path.join(root, file)
                            zf.write(fp, os.path.relpath(fp, os.path.dirname(source)))
                else:
                    zf.write(source, os.path.basename(source))
            size = os.path.getsize(dest)
            return f"✅ Created: {dest}\nSize: {size:,} bytes"
        except Exception as e:
            return f"[ERROR] zip_create: {e}"

    async def tool_zip_extract(self, args):
        """Extract a ZIP archive."""
        import zipfile as _zip
        source = args.get('source', '')
        dest   = args.get('destination', '') or os.path.dirname(os.path.abspath(source))
        try:
            source = os.path.abspath(source)
            dest   = os.path.abspath(dest)
            os.makedirs(dest, exist_ok=True)
            with _zip.ZipFile(source, 'r') as zf:
                names = zf.namelist()
                zf.extractall(dest)
            return f"✅ Extracted {len(names)} files to: {dest}"
        except Exception as e:
            return f"[ERROR] zip_extract: {e}"

    async def tool_image_info(self, args):
        """Get image metadata without loading to AI."""
        path = args.get('path', '')
        try:
            from PIL import Image as _Image
            size = os.path.getsize(path)
            with _Image.open(path) as img:
                w, h   = img.size
                fmt    = img.format or 'UNKNOWN'
                mode   = img.mode
                info   = img.info
            exif_str = ''
            if 'exif' in info:
                exif_str = ' (EXIF data present)'
            return (
                f"File:       {path}\n"
                f"Format:     {fmt}\n"
                f"Dimensions: {w} x {h} px\n"
                f"Color mode: {mode}\n"
                f"File size:  {size:,} bytes ({size/1024:.1f} KB){exif_str}"
            )
        except ImportError:
            # Fallback without PIL — just file size + extension
            ext = os.path.splitext(path)[1].upper().lstrip('.')
            size = os.path.getsize(path) if os.path.exists(path) else 0
            return f"File: {path}\nFormat: {ext}\nFile size: {size:,} bytes\n(Install Pillow for full metadata: pip install Pillow)"
        except Exception as e:
            return f"[ERROR] image_info: {e}"

    async def tool_clipboard_get(self, args):
        """Read text from the OS clipboard."""
        try:
            import subprocess as _sp, sys as _sys
            if _sys.platform == 'win32':
                result = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', 'Get-Clipboard',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await result.communicate()
                text = stdout.decode('utf-8', errors='replace').strip()
            elif _sys.platform == 'darwin':
                result = await asyncio.create_subprocess_exec(
                    'pbpaste', stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await result.communicate()
                text = stdout.decode('utf-8', errors='replace').strip()
            else:
                # Linux — try xclip then xsel
                try:
                    result = await asyncio.create_subprocess_exec(
                        'xclip', '-selection', 'clipboard', '-o',
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    stdout, _ = await result.communicate()
                    text = stdout.decode('utf-8', errors='replace').strip()
                except FileNotFoundError:
                    result = await asyncio.create_subprocess_exec(
                        'xsel', '--clipboard', '--output',
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    stdout, _ = await result.communicate()
                    text = stdout.decode('utf-8', errors='replace').strip()
            if not text:
                return "(Clipboard is empty)"
            return f"Clipboard content ({len(text)} chars):\n{text}"
        except Exception as e:
            return f"[ERROR] clipboard_get: {e}"

    async def tool_clipboard_set(self, args):
        """Write text to the OS clipboard."""
        text = args.get('text', '')
        try:
            import subprocess as _sp, sys as _sys
            if _sys.platform == 'win32':
                proc = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', f'Set-Clipboard -Value @"\n{text}\n"@',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            elif _sys.platform == 'darwin':
                proc = await asyncio.create_subprocess_exec(
                    'pbcopy', stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate(input=text.encode('utf-8'))
            else:
                try:
                    proc = await asyncio.create_subprocess_exec(
                        'xclip', '-selection', 'clipboard',
                        stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    await proc.communicate(input=text.encode('utf-8'))
                except FileNotFoundError:
                    proc = await asyncio.create_subprocess_exec(
                        'xsel', '--clipboard', '--input',
                        stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    await proc.communicate(input=text.encode('utf-8'))
            return f"✅ Copied {len(text)} characters to clipboard."
        except Exception as e:
            return f"[ERROR] clipboard_set: {e}"

    async def tool_notify(self, args):
        """Send a desktop notification."""
        import sys as _sys
        title   = args.get('title', 'Galactic AI')
        message = args.get('message', '')
        sound   = bool(args.get('sound', False))
        try:
            if _sys.platform == 'win32':
                # Use PowerShell toast on Windows 10/11
                ps_script = f"""
Add-Type -AssemblyName System.Windows.Forms
$notify = New-Object System.Windows.Forms.NotifyIcon
$notify.Icon = [System.Drawing.SystemIcons]::Information
$notify.Visible = $true
$notify.ShowBalloonTip(5000, '{title.replace("'", "")}', '{message.replace("'", "")}', [System.Windows.Forms.ToolTipIcon]::Info)
Start-Sleep -Milliseconds 5500
$notify.Dispose()
""".strip()
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', ps_script,
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            elif _sys.platform == 'darwin':
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'osascript', '-e',
                    f'display notification "{message}" with title "{title}"',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            else:
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'notify-send', title, message,
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            return f"✅ Notification sent: '{title}' — {message}"
        except Exception as e:
            return f"[ERROR] notify: {e}"

    async def tool_window_list(self, args):
        """List all open windows."""
        import sys as _sys
        try:
            if _sys.platform == 'win32':
                import ctypes, ctypes.wintypes as _wt
                EnumWindows        = ctypes.windll.user32.EnumWindows
                GetWindowTextW     = ctypes.windll.user32.GetWindowTextW
                GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                IsWindowVisible    = ctypes.windll.user32.IsWindowVisible
                GetWindowThreadProcessId = ctypes.windll.user32.GetWindowThreadProcessId
                EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                windows = []
                def callback(hwnd, lParam):
                    if IsWindowVisible(hwnd):
                        length = GetWindowTextLengthW(hwnd)
                        if length > 0:
                            buf = ctypes.create_unicode_buffer(length + 1)
                            GetWindowTextW(hwnd, buf, length + 1)
                            pid = ctypes.c_ulong()
                            GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
                            windows.append((int(hwnd), buf.value, pid.value))
                    return True
                EnumWindows(EnumWindowsProc(callback), 0)
                if not windows:
                    return "No visible windows found."
                lines = [f"{'HWND':>10}  {'PID':>7}  TITLE"]
                lines.append('-' * 70)
                for hwnd, title, pid in sorted(windows, key=lambda x: x[1].lower()):
                    lines.append(f"{hwnd:>10}  {pid:>7}  {title[:60]}")
                return '\n'.join(lines)
            else:
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'wmctrl', '-l', stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await proc.communicate()
                return stdout.decode('utf-8', errors='replace').strip() or "No windows found (wmctrl output was empty)"
        except Exception as e:
            return f"[ERROR] window_list: {e}"

    async def tool_window_focus(self, args):
        """Bring a window to the foreground."""
        import sys as _sys
        title = args.get('title', '')
        hwnd  = args.get('hwnd', None)
        try:
            if _sys.platform == 'win32':
                import ctypes
                if hwnd:
                    target_hwnd = int(hwnd)
                else:
                    # Find by title substring
                    EnumWindows = ctypes.windll.user32.EnumWindows
                    GetWindowTextW = ctypes.windll.user32.GetWindowTextW
                    GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
                    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                    found = []
                    def callback(h, lParam):
                        if IsWindowVisible(h):
                            length = GetWindowTextLengthW(h)
                            if length > 0:
                                buf = ctypes.create_unicode_buffer(length + 1)
                                GetWindowTextW(h, buf, length + 1)
                                if title.lower() in buf.value.lower():
                                    found.append((int(h), buf.value))
                        return True
                    EnumWindows(EnumWindowsProc(callback), 0)
                    if not found:
                        return f"[ERROR] No window found matching '{title}'"
                    target_hwnd = found[0][0]
                # Restore if minimized, then set foreground
                ctypes.windll.user32.ShowWindow(target_hwnd, 9)  # SW_RESTORE
                ctypes.windll.user32.SetForegroundWindow(target_hwnd)
                return f"✅ Focused window HWND={target_hwnd}"
            else:
                import subprocess as _sp
                cmd = ['wmctrl', '-a', title] if title else ['wmctrl', '-ia', str(hwnd)]
                proc = await asyncio.create_subprocess_exec(*cmd, stdout=_sp.PIPE, stderr=_sp.PIPE)
                _, stderr = await proc.communicate()
                if proc.returncode != 0:
                    return f"[ERROR] wmctrl: {stderr.decode().strip()}"
                return f"✅ Window focused"
        except Exception as e:
            return f"[ERROR] window_focus: {e}"

    async def tool_window_resize(self, args):
        """Resize and/or move a window."""
        import sys as _sys
        title  = args.get('title', '')
        hwnd   = args.get('hwnd', None)
        x      = args.get('x', None)
        y      = args.get('y', None)
        width  = args.get('width', None)
        height = args.get('height', None)
        try:
            if _sys.platform == 'win32':
                import ctypes
                if hwnd:
                    target_hwnd = int(hwnd)
                else:
                    EnumWindows = ctypes.windll.user32.EnumWindows
                    GetWindowTextW = ctypes.windll.user32.GetWindowTextW
                    GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
                    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                    found = []
                    def callback(h, lParam):
                        if IsWindowVisible(h):
                            length = GetWindowTextLengthW(h)
                            if length > 0:
                                buf = ctypes.create_unicode_buffer(length + 1)
                                GetWindowTextW(h, buf, length + 1)
                                if title.lower() in buf.value.lower():
                                    found.append(int(h))
                        return True
                    EnumWindows(EnumWindowsProc(callback), 0)
                    if not found:
                        return f"[ERROR] No window found matching '{title}'"
                    target_hwnd = found[0]
                # Get current rect
                import ctypes.wintypes as _wt
                rect = _wt.RECT()
                ctypes.windll.user32.GetWindowRect(target_hwnd, ctypes.byref(rect))
                nx = x      if x      is not None else rect.left
                ny = y      if y      is not None else rect.top
                nw = width  if width  is not None else (rect.right - rect.left)
                nh = height if height is not None else (rect.bottom - rect.top)
                ctypes.windll.user32.MoveWindow(target_hwnd, int(nx), int(ny), int(nw), int(nh), True)
                return f"✅ Window moved/resized: pos=({nx},{ny}) size={nw}x{nh}"
            else:
                import subprocess as _sp
                if title:
                    geo = ''
                    if width and height:
                        geo = f"{width}x{height}"
                        if x is not None and y is not None:
                            geo += f"+{x}+{y}"
                    proc = await asyncio.create_subprocess_exec(
                        'wmctrl', '-r', title, '-e', f"0,{x or -1},{y or -1},{width or -1},{height or -1}",
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    _, stderr = await proc.communicate()
                    if proc.returncode != 0:
                        return f"[ERROR] wmctrl: {stderr.decode().strip()}"
                    return "✅ Window resized"
                return "[ERROR] Provide title or hwnd"
        except Exception as e:
            return f"[ERROR] window_resize: {e}"

    async def tool_http_request(self, args):
        """Make a raw HTTP request to any URL."""
        method  = args.get('method', 'GET').upper()
        url     = args.get('url', '')
        headers = args.get('headers', {})
        body_json = args.get('json', None)
        body_data = args.get('data', None)
        params  = args.get('params', None)
        timeout = int(args.get('timeout', 30))
        if not url:
            return "[ERROR] http_request requires a 'url' argument."
        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                kwargs = {'headers': headers or {}}
                if params:
                    kwargs['params'] = params
                if body_json is not None:
                    kwargs['json'] = body_json
                elif body_data is not None:
                    kwargs['content'] = body_data.encode() if isinstance(body_data, str) else body_data
                r = await client.request(method, url, **kwargs)
            ct = r.headers.get('content-type', '')
            if 'application/json' in ct:
                try:
                    body = json.dumps(r.json(), indent=2)[:8000]
                except Exception:
                    body = r.text[:8000]
            else:
                body = r.text[:8000]
            return (
                f"HTTP {r.status_code} {r.reason_phrase}\n"
                f"Content-Type: {ct}\n"
                f"Headers: {dict(r.headers)}\n\n"
                f"{body}"
            )
        except Exception as e:
            return f"[ERROR] http_request: {e}"

    async def tool_qr_generate(self, args):
        """Generate a QR code and save it as a PNG image."""
        text  = args.get('text', '')
        size  = int(args.get('size', 10))
        border = int(args.get('border', 4))
        ec_map = {'L': 1, 'M': 0, 'Q': 3, 'H': 2}
        ec    = ec_map.get(args.get('error_correction', 'M').upper(), 0)
        if not text:
            return "[ERROR] qr_generate requires 'text' argument."
        try:
            import qrcode as _qr
            import time as _time
            qr = _qr.QRCode(
                version=None,
                error_correction=ec,
                box_size=size,
                border=border,
            )
            qr.add_data(text)
            qr.make(fit=True)
            img = qr.make_image(fill_color='black', back_color='white')
            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'qr')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"qr_{int(_time.time())}.png"
            path  = os.path.join(img_subdir, fname)
            img.save(path)
            self.last_image_file = path
            return f"✅ QR code saved to: {path}\nContent: {text[:80]}"
        except ImportError:
            return "[ERROR] qrcode library not installed. Run: pip install qrcode[pil]"
        except Exception as e:
            return f"[ERROR] qr_generate: {e}"

    async def tool_env_get(self, args):
        """Read environment variable(s)."""
        name = args.get('name', '')
        _SKIP = {'PATH', 'PYTHONPATH', 'APPDATA', 'LOCALAPPDATA', 'PROGRAMDATA',
                 'COMSPEC', 'PROCESSOR_ARCHITECTURE', 'NUMBER_OF_PROCESSORS'}
        if name:
            val = os.environ.get(name)
            if val is None:
                return f"Environment variable '{name}' is not set."
            return f"{name}={val}"
        else:
            lines = []
            for k, v in sorted(os.environ.items()):
                if any(secret in k.upper() for secret in ['KEY', 'SECRET', 'PASSWORD', 'TOKEN', 'PASS']):
                    lines.append(f"{k}=<hidden>")
                else:
                    lines.append(f"{k}={v[:120]}")
            return '\n'.join(lines)

    async def tool_env_set(self, args):
        """Set an environment variable for this session."""
        name  = args.get('name', '')
        value = args.get('value', '')
        if not name:
            return "[ERROR] env_set requires 'name' argument."
        os.environ[name] = value
        return f"✅ Set {name}={value}"

    async def tool_system_info(self, args):
        """Return detailed system hardware and OS information."""
        import platform as _pl, time as _tm
        try:
            import psutil as _ps
            cpu_count  = _ps.cpu_count(logical=True)
            cpu_phys   = _ps.cpu_count(logical=False)
            cpu_pct    = _ps.cpu_percent(interval=0.3)
            mem        = _ps.virtual_memory()
            disk       = _ps.disk_usage('/')
            boot_time  = _ps.boot_time()
            uptime_s   = int(_tm.time() - boot_time)
            uptime_str = f"{uptime_s//3600}h {(uptime_s%3600)//60}m"
            proc_count = len(_ps.pids())
            ram_total  = f"{mem.total / (1024**3):.1f} GB"
            ram_used   = f"{mem.used / (1024**3):.1f} GB ({mem.percent:.0f}%)"
            disk_total = f"{disk.total / (1024**3):.1f} GB"
            disk_used  = f"{disk.used / (1024**3):.1f} GB ({disk.percent:.0f}%)"
            psutil_info = (
                f"CPU:          {cpu_phys} physical / {cpu_count} logical cores @ {cpu_pct:.1f}% usage\n"
                f"RAM:          {ram_used} / {ram_total}\n"
                f"Disk (/):     {disk_used} / {disk_total}\n"
                f"Uptime:       {uptime_str}\n"
                f"Processes:    {proc_count} running\n"
            )
        except ImportError:
            psutil_info = "(Install psutil for CPU/RAM stats: pip install psutil)\n"
        import sys as _sys
        return (
            f"OS:           {_pl.system()} {_pl.release()} ({_pl.version()[:60]})\n"
            f"Machine:      {_pl.machine()} / {_pl.processor()[:60]}\n"
            f"Python:       {_sys.version.split()[0]} ({_sys.executable})\n"
            + psutil_info
        )

    async def tool_kill_process_by_name(self, args):
        """Kill processes by name substring."""
        name  = args.get('name', '').lower()
        force = bool(args.get('force', False))
        if not name:
            return "[ERROR] kill_process_by_name requires 'name' argument."
        try:
            import psutil as _ps
            killed = []
            for proc in _ps.process_iter(['pid', 'name', 'cmdline']):
                try:
                    pname = (proc.info['name'] or '').lower()
                    if name in pname:
                        if force:
                            proc.kill()
                        else:
                            proc.terminate()
                        killed.append(f"PID {proc.pid}: {proc.info['name']}")
                except (_ps.NoSuchProcess, _ps.AccessDenied):
                    pass
            if not killed:
                return f"No processes found matching '{name}'"
            return f"✅ Terminated {len(killed)} process(es):\n" + '\n'.join(killed)
        except ImportError:
            # Fallback to taskkill / kill
            import subprocess as _sp
            import sys as _sys
            if _sys.platform == 'win32':
                flag = '/F' if force else ''
                cmd = ['taskkill', '/IM', f'*{name}*', flag] if flag else ['taskkill', '/IM', f'*{name}*']
                proc = await asyncio.create_subprocess_exec(*[c for c in cmd if c], stdout=_sp.PIPE, stderr=_sp.PIPE)
                stdout, stderr = await proc.communicate()
                return stdout.decode('utf-8', errors='replace').strip() or stderr.decode('utf-8', errors='replace').strip()
            else:
                sig = '-9' if force else '-15'
                proc = await asyncio.create_subprocess_exec('pkill', sig, '-f', name, stdout=_sp.PIPE, stderr=_sp.PIPE)
                stdout, stderr = await proc.communicate()
                return f"pkill exit {proc.returncode}: {(stdout+stderr).decode(errors='replace').strip() or 'Done'}"
        except Exception as e:
            return f"[ERROR] kill_process_by_name: {e}"

    async def tool_color_pick(self, args):
        """Sample pixel color at screen coordinates."""
        x = int(args.get('x', 0))
        y = int(args.get('y', 0))
        try:
            import pyautogui as _pag
            import colorsys as _cs
            pixel = _pag.screenshot().getpixel((x, y))
            r, g, b = pixel[0], pixel[1], pixel[2]
            h, s, v = _cs.rgb_to_hsv(r/255, g/255, b/255)
            return (
                f"Pixel at ({x}, {y}):\n"
                f"  Hex:  #{r:02X}{g:02X}{b:02X}\n"
                f"  RGB:  rgb({r}, {g}, {b})\n"
                f"  HSV:  hsl({h*360:.0f}°, {s*100:.0f}%, {v*100:.0f}%)"
            )
        except Exception as e:
            return f"[ERROR] color_pick: {e}"

    async def tool_text_transform(self, args):
        """Transform text in various ways."""
        import re as _re, json as _json, urllib.parse as _up, base64 as _b64, csv as _csv, io as _io
        text      = args.get('text', '')
        operation = args.get('operation', '').lower().replace(' ', '_')
        pattern   = args.get('pattern', '')
        try:
            if operation == 'upper':
                return text.upper()
            elif operation == 'lower':
                return text.lower()
            elif operation == 'title':
                return text.title()
            elif operation == 'snake_case':
                return _re.sub(r'[\s\-]+', '_', _re.sub(r'(?<!^)(?=[A-Z])', '_', text)).lower()
            elif operation == 'camel_case':
                parts = _re.split(r'[\s_\-]+', text)
                return parts[0].lower() + ''.join(p.title() for p in parts[1:])
            elif operation == 'base64_encode':
                return _b64.b64encode(text.encode('utf-8')).decode('ascii')
            elif operation == 'base64_decode':
                return _b64.b64decode(text).decode('utf-8', errors='replace')
            elif operation == 'url_encode':
                return _up.quote(text, safe='')
            elif operation == 'url_decode':
                return _up.unquote(text)
            elif operation == 'reverse':
                return text[::-1]
            elif operation == 'count':
                lines = text.splitlines()
                words = text.split()
                non_space = len(text.replace(' ', '').replace('\n', ''))
                return (f"Characters: {len(text):,}\n"
                        f"Words:      {len(words):,}\n"
                        f"Lines:      {len(lines):,}\n"
                        f"Non-space:  {non_space:,}")
            elif operation == 'strip':
                return text.strip()
            elif operation == 'regex_extract':
                if not pattern:
                    return "[ERROR] regex_extract requires a 'pattern' argument."
                matches = _re.findall(pattern, text)
                if not matches:
                    return "No matches found."
                return f"Found {len(matches)} match(es):\n" + '\n'.join(str(m) for m in matches[:100])
            elif operation == 'json_format':
                try:
                    parsed = _json.loads(text)
                    return _json.dumps(parsed, indent=2, ensure_ascii=False)
                except Exception as e:
                    return f"[ERROR] Invalid JSON: {e}"
            elif operation == 'csv_to_json':
                reader = _csv.DictReader(_io.StringIO(text))
                rows = list(reader)
                return _json.dumps(rows, indent=2, ensure_ascii=False)
            else:
                ops = ['upper','lower','title','snake_case','camel_case','base64_encode','base64_decode',
                       'url_encode','url_decode','reverse','count','strip','regex_extract','json_format','csv_to_json']
                return f"[ERROR] Unknown operation '{operation}'. Available: {', '.join(ops)}"
        except Exception as e:
            return f"[ERROR] text_transform ({operation}): {e}"

    # ── New v0.9.2 Tool Implementations ──────────────────────────────

    async def tool_execute_python(self, args):
        """Execute Python code in a subprocess."""
        code = args.get('code', '')
        timeout = min(int(args.get('timeout', 60)), 300)
        if not code.strip():
            return "[ERROR] No code provided."
        import tempfile
        tmp = None
        try:
            tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False, encoding='utf-8')
            tmp.write(code)
            tmp.close()
            proc = await asyncio.create_subprocess_exec(
                'python', tmp.name,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            try:
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
            except asyncio.TimeoutError:
                proc.kill()
                await proc.wait()
                return f"[Timeout] Python script exceeded {timeout}s and was killed."
            out = stdout.decode('utf-8', errors='ignore').strip()
            err = stderr.decode('utf-8', errors='ignore').strip()
            result = ""
            if out:
                result += f"STDOUT:\n{out}\n"
            if err:
                result += f"STDERR:\n{err}\n"
            if proc.returncode != 0:
                result += f"Exit code: {proc.returncode}"
            return result or "Script completed with no output."
        except Exception as e:
            return f"[ERROR] execute_python: {e}"
        finally:
            if tmp:
                try:
                    os.unlink(tmp.name)
                except Exception:
                    pass

    async def tool_wait(self, args):
        """Pause execution."""
        seconds = min(float(args.get('seconds', 1)), 300)
        await asyncio.sleep(seconds)
        return f"Waited {seconds:.1f} seconds."

    async def tool_send_telegram(self, args):
        """Send a proactive Telegram message."""
        message = args.get('message', '')
        chat_id = args.get('chat_id', '') or str(self.core.config.get('telegram', {}).get('admin_chat_id', ''))
        image_path = args.get('image_path', '')
        if not chat_id:
            return "[ERROR] No chat_id provided and no admin_chat_id in config."
        if not message:
            return "[ERROR] No message provided."
        try:
            tg = getattr(self.core, 'telegram', None)
            if not tg:
                return "[ERROR] Telegram bridge not available."
            if image_path and os.path.exists(image_path):
                await tg.send_photo(int(chat_id), image_path, caption=message)
                return f"Sent photo + message to Telegram chat {chat_id}."
            else:
                await tg.send_message(int(chat_id), message)
                return f"Sent message to Telegram chat {chat_id}."
        except Exception as e:
            return f"[ERROR] send_telegram: {e}"

    async def tool_read_pdf(self, args):
        """Extract text from a PDF file."""
        path = args.get('path', '')
        pages_arg = args.get('pages', 'all')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            try:
                import pdfplumber
                with pdfplumber.open(path) as pdf:
                    total = len(pdf.pages)
                    page_range = self._parse_page_range(pages_arg, total)
                    texts = []
                    for i in page_range:
                        page = pdf.pages[i]
                        text = page.extract_text()
                        if text:
                            texts.append(f"--- Page {i+1} ---\n{text}")
                    return "\n\n".join(texts) if texts else "[INFO] No text content found in PDF."
            except ImportError:
                pass
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(path)
                total = len(reader.pages)
                page_range = self._parse_page_range(pages_arg, total)
                texts = []
                for i in page_range:
                    text = reader.pages[i].extract_text()
                    if text:
                        texts.append(f"--- Page {i+1} ---\n{text}")
                return "\n\n".join(texts) if texts else "[INFO] No text content found in PDF."
            except ImportError:
                return "[ERROR] Install pdfplumber or PyPDF2: pip install pdfplumber"
        except Exception as e:
            return f"[ERROR] read_pdf: {e}"

    def _parse_page_range(self, spec, total):
        """Parse page range like '1-5', '3', 'all'."""
        if not spec or spec.lower() == 'all':
            return range(total)
        if '-' in spec:
            parts = spec.split('-')
            start = max(0, int(parts[0]) - 1)
            end = min(total, int(parts[1]))
            return range(start, end)
        return [int(spec) - 1]

    async def tool_read_csv(self, args):
        """Read a CSV file and return as JSON rows."""
        import csv as _csv
        path = args.get('path', '')
        limit = int(args.get('limit', 200))
        delimiter = args.get('delimiter', ',')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            with open(path, 'r', encoding='utf-8-sig') as f:
                reader = _csv.DictReader(f, delimiter=delimiter)
                rows = []
                for i, row in enumerate(reader):
                    if i >= limit:
                        break
                    rows.append(dict(row))
            return json.dumps({"total_rows": len(rows), "columns": list(rows[0].keys()) if rows else [], "rows": rows}, indent=2)
        except Exception as e:
            return f"[ERROR] read_csv: {e}"

    async def tool_write_csv(self, args):
        """Write rows to a CSV file."""
        import csv as _csv
        path = args.get('path', '')
        rows = args.get('rows', [])
        append = args.get('append', False)
        if not path:
            return "[ERROR] No path provided."
        if not rows:
            return "[ERROR] No rows provided."
        try:
            mode = 'a' if append else 'w'
            file_exists = os.path.exists(path) and append
            with open(path, mode, newline='', encoding='utf-8') as f:
                writer = _csv.DictWriter(f, fieldnames=rows[0].keys())
                if not file_exists:
                    writer.writeheader()
                writer.writerows(rows)
            return f"Wrote {len(rows)} rows to {path}."
        except Exception as e:
            return f"[ERROR] write_csv: {e}"

    async def tool_read_excel(self, args):
        """Read an Excel (.xlsx) file."""
        path = args.get('path', '')
        sheet = args.get('sheet', None)
        limit = int(args.get('limit', 100))
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return "[INFO] Empty spreadsheet."
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:limit+1]:
                data.append({headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)})
            wb.close()
            return json.dumps({"sheets": wb.sheetnames if hasattr(wb, 'sheetnames') else [], "columns": headers, "rows": data, "total_rows": len(data)}, indent=2)
        except ImportError:
            return "[ERROR] Install openpyxl: pip install openpyxl"
        except Exception as e:
            return f"[ERROR] read_excel: {e}"

    async def tool_regex_search(self, args):
        """Search files with regex."""
        import fnmatch as _fn
        pattern = args.get('pattern', '')
        search_path = args.get('path', '.')
        file_pattern = args.get('file_pattern', '*')
        limit = int(args.get('limit', 50))
        if not pattern:
            return "[ERROR] No pattern provided."
        try:
            compiled = re.compile(pattern, re.IGNORECASE)
        except re.error as e:
            return f"[ERROR] Invalid regex: {e}"
        results = []
        try:
            if os.path.isfile(search_path):
                files = [search_path]
            else:
                files = []
                for root, dirs, fnames in os.walk(search_path):
                    for fn in fnames:
                        if _fn.fnmatch(fn, file_pattern):
                            files.append(os.path.join(root, fn))
                    if len(files) > 5000:
                        break
            for fpath in files:
                try:
                    with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                        for line_no, line in enumerate(f, 1):
                            if compiled.search(line):
                                results.append(f"{fpath}:{line_no}: {line.rstrip()[:200]}")
                                if len(results) >= limit:
                                    return f"Found {len(results)} matches (limit reached):\n" + "\n".join(results)
                except (PermissionError, IsADirectoryError):
                    continue
            return f"Found {len(results)} matches:\n" + "\n".join(results) if results else "No matches found."
        except Exception as e:
            return f"[ERROR] regex_search: {e}"

    async def tool_image_resize(self, args):
        """Resize an image."""
        path = args.get('path', '')
        width = args.get('width')
        height = args.get('height')
        output = args.get('output_path', '')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            from PIL import Image
            img = Image.open(path)
            orig_w, orig_h = img.size
            new_w = int(width) if width else orig_w
            new_h = int(height) if height else orig_h
            resized = img.resize((new_w, new_h), Image.LANCZOS)
            out_path = output or os.path.splitext(path)[0] + f"_{new_w}x{new_h}" + os.path.splitext(path)[1]
            resized.save(out_path)
            return f"Resized {orig_w}x{orig_h} → {new_w}x{new_h}. Saved to: {out_path}"
        except ImportError:
            return "[ERROR] Pillow not installed. Run: pip install Pillow"
        except Exception as e:
            return f"[ERROR] image_resize: {e}"

    async def tool_image_convert(self, args):
        """Convert image format."""
        path = args.get('path', '')
        fmt = args.get('format', 'png').lower()
        output = args.get('output_path', '')
        quality = int(args.get('quality', 85))
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            from PIL import Image
            img = Image.open(path)
            if fmt in ('jpeg', 'jpg') and img.mode in ('RGBA', 'LA', 'P'):
                img = img.convert('RGB')
            out_path = output or os.path.splitext(path)[0] + '.' + ('jpg' if fmt == 'jpeg' else fmt)
            save_kwargs = {}
            if fmt in ('jpeg', 'jpg', 'webp'):
                save_kwargs['quality'] = quality
            img.save(out_path, **save_kwargs)
            return f"Converted to {fmt.upper()}. Saved to: {out_path}"
        except ImportError:
            return "[ERROR] Pillow not installed. Run: pip install Pillow"
        except Exception as e:
            return f"[ERROR] image_convert: {e}"

    async def _git_exec(self, cmd_args, cwd=None):
        """Helper to run a git command."""
        cwd = cwd or self.core.config.get('paths', {}).get('workspace', '.')
        proc = await asyncio.create_subprocess_exec(
            'git', *cmd_args, cwd=cwd,
            stdout=asyncio.subprocess.PIPE, stderr=asyncio.subprocess.PIPE,
        )
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
        out = stdout.decode('utf-8', errors='ignore').strip()
        err = stderr.decode('utf-8', errors='ignore').strip()
        if proc.returncode != 0 and err:
            return f"[git error] {err}"
        return out or err or "(no output)"

    async def tool_git_status(self, args):
        path = args.get('path')
        return await self._git_exec(['status', '--short'], cwd=path)

    async def tool_git_diff(self, args):
        path = args.get('path')
        cmd = ['diff', '--stat']
        if args.get('staged'):
            cmd.append('--cached')
        return await self._git_exec(cmd, cwd=path)

    async def tool_git_log(self, args):
        path = args.get('path')
        count = str(int(args.get('count', 10)))
        return await self._git_exec(['log', f'--oneline', f'-{count}'], cwd=path)

    async def tool_git_commit(self, args):
        path = args.get('path')
        message = args.get('message', 'Auto-commit by Galactic AI')
        files = args.get('files', [])
        cwd = path or self.core.config.get('paths', {}).get('workspace', '.')
        if files:
            for f in files:
                await self._git_exec(['add', f], cwd=cwd)
        else:
            await self._git_exec(['add', '-A'], cwd=cwd)
        return await self._git_exec(['commit', '-m', message], cwd=cwd)

    # tool_spawn_subagent, tool_check_subagent — Migrated to skills/core/subagent_manager.py
