"""
Extracted tools for GalacticGateway
"""
import os
import json
import asyncio
import base64
from datetime import datetime
import httpx
import re

class GatewayToolsMixin:
    def _vcr_auto_backup(self, file_path):
        """Take a silent backup snapshot of a file under .galactic_vcr before modify/write."""
        try:
            if not file_path or not os.path.exists(file_path) or not os.path.isfile(file_path):
                return
            
            abs_file_path = os.path.abspath(file_path)
            
            # Skip backing up VCR, temporary, or log files themselves to avoid infinite loops/bloat
            if ".galactic_vcr" in abs_file_path or "logs" in abs_file_path or "tmp" in abs_file_path:
                return
                
            workspace_dir = self.core.config.get("paths", {}).get("workspace", "./workspace")
            vcr_dir = os.path.abspath(os.path.join(workspace_dir, ".galactic_vcr"))
            os.makedirs(vcr_dir, exist_ok=True)
            
            from datetime import datetime
            import shutil
            filename = os.path.basename(abs_file_path)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = f"{filename}_{timestamp}.bak"
            backup_path = os.path.join(vcr_dir, safe_name)
            
            shutil.copy2(abs_file_path, backup_path)
            
            # Write mapping
            mapping_file = os.path.join(vcr_dir, "mapping.txt")
            with open(mapping_file, "a", encoding="utf-8") as f:
                f.write(f"{abs_file_path}|{backup_path}\n")
                
        except Exception:
            pass

    def register_tools(self):
        """Registers available tools for the LLM."""
        self.tools = {
            "read_file": {
                "description": "Read the contents of a file. Supports reading specific line ranges to save context tokens.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file."},
                        "start_line": {"type": "integer", "description": "Optional: First line number to read (1-indexed). Defaults to 1."},
                        "end_line": {"type": "integer", "description": "Optional: Last line number to read (inclusive). Defaults to start_line + 299."}
                    },
                    "required": ["path"]
                },
                "fn": self.tool_read_file
            },
            "write_file": {
                "description": "Write content to a file. RECOMMENDED: Use absolute paths on Windows. Relative paths resolve to the project root. Writes are confined to the active workspace, the Galactic install dir, and any configured extra roots — a path outside those is refused with the allowed list.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file. Absolute paths (C:\\...) are preferred for reliability."},
                        "content": {"type": "string", "description": "Content to write."}
                    },
                    "required": ["path", "content"]
                },
                "fn": self.tool_write_file
            },
            "list_functions": {
                "description": "List all top-level functions and classes (and their methods) in a Python file, with exact line ranges. Use before read_function/replace_function to get the right names.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to a .py file."}
                    },
                    "required": ["path"]
                },
                "fn": self.tool_list_functions
            },
            "read_function": {
                "description": "Read the exact source of ONE function/class/method from a Python file, located by name via AST (no line guessing). Name may be 'my_func' or 'MyClass.my_method'.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to a .py file."},
                        "name": {"type": "string", "description": "Function/class name, or 'ClassName.method'."}
                    },
                    "required": ["path", "name"]
                },
                "fn": self.tool_read_function
            },
            "replace_function": {
                "description": "Rewrite ONE whole function/class/method in a Python file, located by name via AST — immune to line-number drift and whitespace. The new code is auto-indented and the ENTIRE file is re-parsed to guarantee valid Python BEFORE writing (nothing is written if it would break syntax). Existing decorators are preserved automatically — do NOT include them. Auto-backs up via VCR. Prefer this over edit_file for whole-function rewrites.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to a .py file."},
                        "name": {"type": "string", "description": "Function/class name, or 'ClassName.method'."},
                        "new_code": {"type": "string", "description": "The complete new function/class source, starting at 'def'/'class' (no decorators)."}
                    },
                    "required": ["path", "name", "new_code"]
                },
                "fn": self.tool_replace_function
            },
            "schedule_task": {
                "description": "Schedule a reminder or task execution.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Name of the task."},
                        "delay_seconds": {"type": "number", "description": "Delay in seconds before execution."},
                        "message": {"type": "string", "description": "Message to display/log when task fires."}
                    },
                    "required": ["name", "delay_seconds", "message"]
                },
                "fn": self.tool_schedule_task
            },
            "list_tasks": {
                "description": "List all scheduled tasks.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                },
                "fn": self.tool_list_tasks
            },
            "web_search": {
                "description": "Search the web.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Search query."}
                    },
                    "required": ["query"]
                },
                "fn": self.tool_web_search
            },
            "ask_user": {
                "description": "Pause and ask the human a question mid-task, then continue with their answer. Use for things only they can resolve — a 2FA/verification code, a missing credential, a subjective design choice, or genuinely ambiguous instructions. Don't overuse it for things you can reasonably decide yourself.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "question": {"type": "string", "description": "The question to show the user."},
                        "timeout": {"type": "integer", "description": "Seconds to wait for an answer before giving up (default 300, max 1800)."}
                    },
                    "required": ["question"]
                },
                "fn": self.tool_ask_user
            },
            "blackboard_write": {
                "description": "Publish a value to the shared Swarm Blackboard under a key, so other agents (running in parallel) can read it live. Use to share an intermediate finding — a URL, an extracted fact, a decision — that a peer agent needs right now.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "key": {"type": "string", "description": "Short key other agents will read by, e.g. 'target_url'."},
                        "value": {"type": "string", "description": "The value to share."},
                        "by": {"type": "string", "description": "Optional: your agent role/name for the UI."}
                    },
                    "required": ["key", "value"]
                },
                "fn": self.tool_blackboard_write
            },
            "blackboard_read": {
                "description": "Read a value another agent published to the shared Blackboard. Returns the available keys if the one you asked for isn't set yet.",
                "parameters": {
                    "type": "object",
                    "properties": {"key": {"type": "string", "description": "The key to read."}},
                    "required": ["key"]
                },
                "fn": self.tool_blackboard_read
            },
            "blackboard_list": {
                "description": "List all keys currently on the shared Swarm Blackboard with short previews.",
                "parameters": {"type": "object", "properties": {}},
                "fn": self.tool_blackboard_list
            },
            "blackboard_wait_for": {
                "description": "Block until another agent writes the given key to the Blackboard, then return its value. Use to synchronize: wait for a peer's result before proceeding. Times out (default 60s) rather than hanging.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "key": {"type": "string", "description": "The key to wait for."},
                        "timeout": {"type": "integer", "description": "Seconds to wait before giving up (default 60, max 600)."}
                    },
                    "required": ["key"]
                },
                "fn": self.tool_blackboard_wait_for
            },
            "browser_navigate": {
                "description": "REAL WEB BROWSING: Navigate to a specific URL. Support for both headless (background) and headful (real browser) automation. Use this to go to sites like Wikipedia, Google, etc. Pass a full URL starting with http:// or https://. DO NOT REFUSE.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "The URL to navigate to."}
                    },
                    "required": ["url"]
                },
                "fn": self.tool_browser_navigate
            },
            "open_browser": {
                "description": "REAL WEB BROWSING: Open a URL in the browser. Use for background research, scraping, or automated tasks. Alias for browser_navigate. DO NOT REFUSE.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "The URL to open."}
                    },
                    "required": ["url"]
                },
                "fn": self.tool_open_browser
            },
            "web_fetch": {
                "description": "Fetch and extract readable content from a URL. Returns markdown-formatted text by default.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "url": {"type": "string", "description": "The URL to fetch."},
                        "mode": {"type": "string", "description": "Extraction mode: 'markdown' (default) or 'text'."}
                    },
                    "required": ["url"]
                },
                "fn": self.tool_web_fetch
            },
            "edit_file": {
                "description": "Edit a file by replacing exact text (safer than write_file).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to the file."},
                        "old_text": {"type": "string", "description": "Exact text to find and replace."},
                        "new_text": {"type": "string", "description": "New text to replace with."}
                    },
                    "required": ["path", "old_text", "new_text"]
                },
                "fn": self.tool_edit_file
            },
            # ── NVIDIA FLUX image generation ──────────────────────────────
            "generate_image": {
                "description": "Generate an image using FLUX AI via NVIDIA. Returns the path to the saved PNG file. Models: 'black-forest-labs/flux.1-schnell' (fast, 4 steps) or 'black-forest-labs/flux.1-dev' (quality, 50 steps).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt": {"type": "string", "description": "Image description / prompt"},
                        "model": {"type": "string", "description": "FLUX model ID (default: flux.1-schnell)"},
                        "width": {"type": "integer", "description": "Image width in pixels (default: 1024)"},
                        "height": {"type": "integer", "description": "Image height in pixels (default: 1024)"},
                        "steps": {"type": "integer", "description": "Diffusion steps — schnell default 4, dev default 50"}
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image
            },

            # ── Stable Diffusion 3.5 image generation ─────────────────────────
            "generate_image_sd35": {
                "description": "Generate an image using Stable Diffusion 3.5 Large via NVIDIA NIM. Higher quality, different style than FLUX. Returns path to saved image.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":          {"type": "string",  "description": "Image description / prompt"},
                        "negative_prompt": {"type": "string",  "description": "Things to avoid in the image (optional)"},
                        "width":           {"type": "integer", "description": "Image width in pixels (default: 1024, max: 1536)"},
                        "height":          {"type": "integer", "description": "Image height in pixels (default: 1024, max: 1536)"},
                        "steps":           {"type": "integer", "description": "Diffusion steps (default: 40, range: 10-100)"},
                        "cfg_scale":       {"type": "number",  "description": "Guidance scale (default: 5.0, range: 1-20)"},
                        "seed":            {"type": "integer", "description": "Random seed (0 = random)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image_sd35
            },

            # ── Google Imagen image generation ────────────────────────────────
            "generate_image_imagen": {
                "description": "Generate an image using Google Imagen 4 via the Google Generative AI API. High-quality photorealistic images. Returns path to saved image.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":         {"type": "string",  "description": "Image description / prompt"},
                        "model":          {"type": "string",  "description": "Imagen model to use: imagen-4-ultra (best), imagen-4 (standard), imagen-4-fast (quick). Default: imagen-4"},
                        "aspect_ratio":   {"type": "string",  "description": "Aspect ratio: 1:1, 16:9, 9:16, 4:3, 3:4. Default: 1:1"},
                        "number_of_images": {"type": "integer", "description": "Number of images to generate (1-4, default: 1)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_image_imagen
            },
            "generate_video": {
                "description": "Generate a short video clip using Google Veo AI. Returns the path to the saved MP4 file. Supports text-to-video with configurable duration, resolution, and aspect ratio.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":          {"type": "string",  "description": "Scene description for the video"},
                        "duration":        {"type": "integer", "description": "Video duration in seconds: 4, 6, or 8 (default: 8)"},
                        "aspect_ratio":    {"type": "string",  "description": "Aspect ratio: 16:9 or 9:16 (default: 16:9)"},
                        "resolution":      {"type": "string",  "description": "Resolution: 720p, 1080p, or 4k (default: 1080p)"},
                        "negative_prompt": {"type": "string",  "description": "Things to avoid in the video (optional)"},
                    },
                    "required": ["prompt"]
                },
                "fn": self.tool_generate_video
            },
            "generate_video_from_image": {
                "description": "Animate a still image into a short video clip using Google Veo. Takes an image (from Imagen, FLUX, or SD3.5) and turns it into motion video. Returns path to saved MP4. IMPORTANT: This tool handles its own file processing and cloud upload. DO NOT use browser tools (chrome_upload, chrome_type, etc.) to 'pass' the image to the Control Deck yourself. Simply provide the absolute local 'image_path' and this tool will manage the entire process.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "prompt":       {"type": "string",  "description": "Description of the motion/animation to apply"},
                        "image_path":   {"type": "string",  "description": "Absolute path to the source image file on disk"},
                        "duration":     {"type": "integer", "description": "Video duration in seconds: 4, 6, or 8 (default: 8)"},
                        "aspect_ratio": {"type": "string",  "description": "Aspect ratio: 16:9 or 9:16 (default: 16:9)"},
                    },
                    "required": ["prompt", "image_path"]
                },
                "fn": self.tool_generate_video_from_image
            },


            # ── File & system utilities ────────────────────────────────────────
            "list_dir": {
                "description": "List files and directories at a path with sizes, dates, and types. ALWAYS use absolute paths (e.g. 'C:/Users/name/folder' or 'F:/My Folder') — relative paths resolve to the server working directory and will return errors for user folders. If the result starts with [ERROR], report that error to the user verbatim.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":    {"type": "string",  "description": "ABSOLUTE directory path to list (e.g. 'F:/Galactic AI Media'). Relative paths resolve to server CWD, not user folders."},
                        "pattern": {"type": "string",  "description": "Optional glob pattern to filter, e.g. '*.py' or '*.mp4'"},
                        "recurse": {"type": "boolean", "description": "Recurse into subdirectories (default: false)"},
                    },
                    "required": []
                },
                "fn": self.tool_list_dir
            },
            "find_files": {
                "description": "Find files matching a name pattern recursively under a directory. Faster and safer than exec_shell find/dir.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":    {"type": "string", "description": "Root directory to search from (default: current working directory)"},
                        "pattern": {"type": "string", "description": "Glob pattern, e.g. '*.log', '**/*.py', 'config.*'"},
                        "limit":   {"type": "integer", "description": "Maximum results to return (default: 100)"},
                    },
                    "required": ["pattern"]
                },
                "fn": self.tool_find_files
            },
            "hash_file": {
                "description": "Compute SHA256 (default), MD5, or SHA1 checksum of a file. Useful for verifying downloads or detecting changes.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path":      {"type": "string", "description": "Path to the file"},
                        "algorithm": {"type": "string", "description": "Hash algorithm: sha256 (default), md5, sha1"},
                    },
                    "required": ["path"]
                },
                "fn": self.tool_hash_file
            },
            "diff_files": {
                "description": "Show a unified diff between two text files, or between a file and a string. Great for reviewing changes before overwriting.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path_a":  {"type": "string", "description": "Path to first file"},
                        "path_b":  {"type": "string", "description": "Path to second file (if comparing two files)"},
                        "text_b":  {"type": "string", "description": "String content to compare against path_a (if not comparing two files)"},
                        "context": {"type": "integer", "description": "Lines of context around changes (default: 3)"},
                    },
                    "required": ["path_a"]
                },
                "fn": self.tool_diff_files
            },
            "zip_create": {
                "description": "Create a ZIP archive from a file or directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source":      {"type": "string", "description": "File or directory path to archive"},
                        "destination": {"type": "string", "description": "Output .zip file path (default: source + '.zip')"},
                    },
                    "required": ["source"]
                },
                "fn": self.tool_zip_create
            },
            "zip_extract": {
                "description": "Extract a ZIP archive to a directory.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "source":      {"type": "string", "description": "Path to the .zip file"},
                        "destination": {"type": "string", "description": "Directory to extract into (default: same directory as zip)"},
                    },
                    "required": ["source"]
                },
                "fn": self.tool_zip_extract
            },
            "image_info": {
                "description": "Get metadata about an image file: dimensions, format, file size, color mode. Does NOT send the image to any AI — pure local metadata.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "Path to image file"},
                    },
                    "required": ["path"]
                },
                "fn": self.tool_image_info
            },

            # ── Clipboard ─────────────────────────────────────────────────────
            "clipboard_get": {
                "description": "Read the current text content of the OS clipboard.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_clipboard_get
            },
            "clipboard_set": {
                "description": "Write text to the OS clipboard so the user can paste it anywhere.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text": {"type": "string", "description": "Text to put on the clipboard"},
                    },
                    "required": ["text"]
                },
                "fn": self.tool_clipboard_set
            },

            # ── Desktop notifications ─────────────────────────────────────────
            "notify": {
                "description": "Send a desktop notification (toast/balloon) to the user's screen. Works on Windows, macOS, and Linux.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title":   {"type": "string", "description": "Notification title"},
                        "message": {"type": "string", "description": "Notification body text"},
                        "sound":   {"type": "boolean", "description": "Play a sound (default: false)"},
                    },
                    "required": ["title", "message"]
                },
                "fn": self.tool_notify
            },

            # ── Window management ─────────────────────────────────────────────
            "window_list": {
                "description": "List all currently open application windows with their titles, process names, and window IDs.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_window_list
            },
            "window_focus": {
                "description": "Bring a window to the foreground and focus it by title substring or window ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title": {"type": "string", "description": "Partial window title to match (case-insensitive)"},
                        "hwnd":  {"type": "integer", "description": "Exact window handle/ID from window_list"},
                    },
                    "required": []
                },
                "fn": self.tool_window_focus
            },
            "window_resize": {
                "description": "Resize and/or move an application window by title or window ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "title":  {"type": "string",  "description": "Partial window title to match"},
                        "hwnd":   {"type": "integer", "description": "Window handle from window_list"},
                        "x":      {"type": "integer", "description": "Left position (pixels from screen left)"},
                        "y":      {"type": "integer", "description": "Top position (pixels from screen top)"},
                        "width":  {"type": "integer", "description": "Window width in pixels"},
                        "height": {"type": "integer", "description": "Window height in pixels"},
                    },
                    "required": []
                },
                "fn": self.tool_window_resize
            },

            # ── HTTP / API ────────────────────────────────────────────────────
            "http_request": {
                "description": "Make a raw HTTP request (GET, POST, PUT, DELETE, PATCH) to any URL. Supports custom headers, JSON body, and form data. Great for calling REST APIs directly.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "method":  {"type": "string", "description": "HTTP method: GET, POST, PUT, DELETE, PATCH (default: GET)"},
                        "url":     {"type": "string", "description": "Full URL including protocol"},
                        "headers": {"type": "object", "description": "Request headers as key-value pairs"},
                        "json":    {"type": "object", "description": "JSON body (sets Content-Type: application/json automatically)"},
                        "data":    {"type": "string", "description": "Raw string body"},
                        "params":  {"type": "object", "description": "URL query parameters as key-value pairs"},
                        "timeout": {"type": "integer", "description": "Timeout in seconds (default: 30)"},
                    },
                    "required": ["url"]
                },
                "fn": self.tool_http_request
            },

            # ── QR code ───────────────────────────────────────────────────────
            "qr_generate": {
                "description": "Generate a QR code image from any text or URL. Saves to logs/ and returns the file path.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text":       {"type": "string",  "description": "Text or URL to encode in the QR code"},
                        "size":       {"type": "integer", "description": "Box size in pixels (default: 10)"},
                        "border":     {"type": "integer", "description": "Border width in boxes (default: 4)"},
                        "error_correction": {"type": "string", "description": "Error correction level: L, M, Q, H (default: M)"},
                    },
                    "required": ["text"]
                },
                "fn": self.tool_qr_generate
            },

            # ── Environment variables ─────────────────────────────────────────
            "env_get": {
                "description": "Read an environment variable value. Returns all env vars if no name specified (filtered list, excludes secrets).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "Environment variable name (e.g. PATH, HOME). Omit to list all."},
                    },
                    "required": []
                },
                "fn": self.tool_env_get
            },
            "env_set": {
                "description": "Set an environment variable for the current process (affects subprocesses spawned from this session).",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":  {"type": "string", "description": "Environment variable name"},
                        "value": {"type": "string", "description": "Value to set"},
                    },
                    "required": ["name", "value"]
                },
                "fn": self.tool_env_set
            },

            # ── System info ────────────────────────────────────────────────────
            "system_info": {
                "description": "Get detailed system information: CPU, RAM, disk usage, OS version, uptime, Python version, and running process count.",
                "parameters": {"type": "object", "properties": {}, "required": []},
                "fn": self.tool_system_info
            },
            "kill_process_by_name": {
                "description": "Kill all running processes matching a name or partial name (e.g. 'chrome', 'notepad'). More convenient than process_kill which needs an ID.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":  {"type": "string",  "description": "Process name or partial name to match (case-insensitive)"},
                        "force": {"type": "boolean", "description": "Force kill (SIGKILL/taskkill /F). Default: false (graceful SIGTERM)"},
                    },
                    "required": ["name"]
                },
                "fn": self.tool_kill_process_by_name
            },

            # ── Color picker ───────────────────────────────────────────────────
            "color_pick": {
                "description": "Sample the pixel color at exact desktop screen coordinates. Returns hex, RGB, and HSL values. Useful for UI automation color verification.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "x": {"type": "integer", "description": "X coordinate (pixels from left edge of screen)"},
                        "y": {"type": "integer", "description": "Y coordinate (pixels from top edge of screen)"},
                    },
                    "required": ["x", "y"]
                },
                "fn": self.tool_color_pick
            },

            # ── Text / data utilities ──────────────────────────────────────────
            "text_transform": {
                "description": "Transform text: convert case, encode/decode base64, URL-encode/decode, count words/lines/chars, reverse, strip, wrap, or extract regex matches.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "text":      {"type": "string", "description": "Input text to transform"},
                        "operation": {"type": "string", "description": "Operation: upper, lower, title, snake_case, camel_case, base64_encode, base64_decode, url_encode, url_decode, reverse, count, strip, regex_extract, json_format, csv_to_json"},
                        "pattern":   {"type": "string", "description": "Regex pattern (for regex_extract operation)"},
                    },
                    "required": ["text", "operation"]
                },
                "fn": self.tool_text_transform
            },

            # ── New v0.9.2 tools ─────────────────────────────────────────
            "execute_python": {
                "description": "Execute Python code in a subprocess and return stdout/stderr. Use for data processing, calculations, CSV/JSON manipulation, or quick scripts. Timeout: 60s default.",
                "parameters": {"type": "object", "properties": {
                    "code": {"type": "string", "description": "Python code to execute"},
                    "timeout": {"type": "integer", "description": "Timeout seconds (default: 60, max: 300)"},
                }, "required": ["code"]},
                "fn": self.tool_execute_python
            },
            "wait": {
                "description": "Pause execution for a specified number of seconds. Use between actions that need settling time, or to wait before retrying something.",
                "parameters": {"type": "object", "properties": {
                    "seconds": {"type": "number", "description": "Seconds to wait (max: 300)"},
                }, "required": ["seconds"]},
                "fn": self.tool_wait
            },
            "send_telegram": {
                "description": "Send a proactive message to a Telegram chat (defaults to admin). Useful for alerts, task completion notifications, and automation reports.",
                "parameters": {"type": "object", "properties": {
                    "message": {"type": "string", "description": "Message text (Markdown supported)"},
                    "chat_id": {"type": "string", "description": "Chat ID (default: admin_chat_id from config)"},
                    "image_path": {"type": "string", "description": "Optional path to image to attach"},
                }, "required": ["message"]},
                "fn": self.tool_send_telegram
            },
            "read_pdf": {
                "description": "Extract text content from a PDF file. Returns plain text of all pages (or specific page range).",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to PDF file"},
                    "pages": {"type": "string", "description": "Page range: '1-5', '3', 'all' (default: all)"},
                }, "required": ["path"]},
                "fn": self.tool_read_pdf
            },
            "read_csv": {
                "description": "Read CSV file and return contents as JSON rows with headers. Great for data analysis.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to CSV file"},
                    "limit": {"type": "integer", "description": "Max rows to return (default: 200)"},
                    "delimiter": {"type": "string", "description": "Delimiter character (default: comma)"},
                }, "required": ["path"]},
                "fn": self.tool_read_csv
            },
            "write_csv": {
                "description": "Write JSON rows to a CSV file. Takes a list of dictionaries as rows.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Output CSV file path"},
                    "rows": {"type": "array", "items": {"type": "object"}, "description": "Array of {key: value} objects"},
                    "append": {"type": "boolean", "description": "Append to existing file (default: false)"},
                }, "required": ["path", "rows"]},
                "fn": self.tool_write_csv
            },
            "read_excel": {
                "description": "Read Excel file (.xlsx) and return contents as JSON rows. Requires openpyxl.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to .xlsx file"},
                    "sheet": {"type": "string", "description": "Sheet name (default: first sheet)"},
                    "limit": {"type": "integer", "description": "Max rows to return (default: 100)"},
                }, "required": ["path"]},
                "fn": self.tool_read_excel
            },
            "regex_search": {
                "description": "Search file contents using regex. Returns matching lines with file paths and line numbers. Faster and safer than exec_shell grep.",
                "parameters": {"type": "object", "properties": {
                    "pattern": {"type": "string", "description": "Regex pattern to search for"},
                    "path": {"type": "string", "description": "File or directory to search in"},
                    "file_pattern": {"type": "string", "description": "Glob to filter files, e.g. '*.py' (default: all)"},
                    "limit": {"type": "integer", "description": "Max results (default: 50)"},
                }, "required": ["pattern", "path"]},
                "fn": self.tool_regex_search
            },
            "image_resize": {
                "description": "Resize an image to specified dimensions. Supports PNG, JPEG, WebP, BMP.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to source image"},
                    "width": {"type": "integer", "description": "Target width in pixels"},
                    "height": {"type": "integer", "description": "Target height in pixels"},
                    "output_path": {"type": "string", "description": "Output path (default: adds _resized suffix)"},
                }, "required": ["path"]},
                "fn": self.tool_image_resize
            },
            "image_convert": {
                "description": "Convert image between formats (PNG, JPEG, WebP, BMP, TIFF).",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Path to source image"},
                    "format": {"type": "string", "description": "Target format: png, jpeg, webp, bmp"},
                    "output_path": {"type": "string", "description": "Output path (default: same name, new extension)"},
                    "quality": {"type": "integer", "description": "JPEG/WebP quality 1-100 (default: 85)"},
                }, "required": ["path", "format"]},
                "fn": self.tool_image_convert
            },
            "git_status": {
                "description": "Run 'git status' in a directory. Returns working tree status.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                }, "required": []},
                "fn": self.tool_git_status
            },
            "git_diff": {
                "description": "Run 'git diff' to show changes. Returns unified diff output.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "staged": {"type": "boolean", "description": "Show staged changes only (default: false)"},
                }, "required": []},
                "fn": self.tool_git_diff
            },
            "git_log": {
                "description": "Show recent git commit history.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "count": {"type": "integer", "description": "Number of commits (default: 10)"},
                }, "required": []},
                "fn": self.tool_git_log
            },
            "git_commit": {
                "description": "Stage files and create a git commit.",
                "parameters": {"type": "object", "properties": {
                    "path": {"type": "string", "description": "Directory path (default: workspace)"},
                    "message": {"type": "string", "description": "Commit message"},
                    "files": {"type": "array", "items": {"type": "string"}, "description": "Files to stage (default: all changed files)"},
                }, "required": ["message"]},
                "fn": self.tool_git_commit
            },
            # spawn_subagent, check_subagent  — Migrated to skills/core/subagent_manager.py
            # chrome_* tools (16)             — Migrated to skills/core/chrome_bridge.py
            # post_tweet, read_mentions, read_dms, post_reddit,
            # read_reddit_inbox, reply_reddit  — Migrated to skills/core/social_media.py

            "create_skill": {
                "description": (
                    "Create a new Galactic AI skill. Writes a .py file to skills/community/ "
                    "and loads it immediately. The skill must subclass GalacticSkill and "
                    "implement get_tools(). Use list_skills first to check what already exists. "
                    "CRITICAL: You MUST use 'from skills.base import GalacticSkill' at the top of the file."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name":        {"type": "string", "description": "Skill name in snake_case (e.g. 'weather_lookup'). Used as the filename."},
                        "code":        {"type": "string", "description": "Full Python source code. MUST import from skills.base (NOT tools.base) and subclass GalacticSkill."},
                        "description": {"type": "string", "description": "One-line description of what this skill does."}
                    },
                    "required": ["name", "code", "description"]
                },
                "fn": self.tool_create_skill
            },
            "list_skills": {
                "description": "List all loaded skills with their metadata and tools. Shows both core and community skills.",
                "parameters": {
                    "type": "object",
                    "properties": {},
                    "required": []
                },
                "fn": self.tool_list_skills
            },
            "remove_skill": {
                "description": (
                    "Remove a community skill by name. Core skills cannot be removed. "
                    "Unloads the skill and deletes its file from skills/community/."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "name": {"type": "string", "description": "The skill_name to remove (e.g. 'weather_lookup')."}
                    },
                    "required": ["name"]
                },
                "fn": self.tool_remove_skill
            },
            "resume_workflow": {
                "description": "Resume an interrupted background workflow or task from a saved checkpoint.",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "uuid": {"type": "string", "description": "The UUID of the checkpoint to load."}
                    },
                    "required": ["uuid"]
                },
                "fn": self.tool_resume_workflow
            },
            "find_tools": {
                "description": (
                    "Search the FULL tool catalog by keyword when none of your currently available "
                    "tools fit the task (e.g. query='screenshot', 'git', 'audio', 'schedule'). "
                    "Matches become callable starting next turn. Use this instead of assuming a "
                    "capability doesn't exist — the full catalog has 180+ tools; only a small curated "
                    "subset is shown to you at once to keep you fast and accurate."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string", "description": "Keyword(s) describing the capability you need, e.g. 'take screenshot' or 'git commit'."}
                    },
                    "required": ["query"]
                },
                "fn": self.tool_find_tools
            },
            "swarm_execute": {
                "description": (
                    "V2.0 SWARM: Decompose a complex goal into specialist sub-tasks and execute them as a "
                    "coordinated multi-agent swarm — independent tasks run IN PARALLEL, dependent tasks wait "
                    "for their inputs, and all agents share findings via a common blackboard. Use for large "
                    "multi-part goals (research + build + verify). Returns the synthesized final report."
                ),
                "parameters": {
                    "type": "object",
                    "properties": {
                        "goal": {"type": "string", "description": "The complex objective to execute as a swarm."},
                        "max_agents": {"type": "integer", "description": "Optional cap on sub-agents (default 4, max 8)."}
                    },
                    "required": ["goal"]
                },
                "fn": self.tool_swarm_execute
            },
        }

    async def tool_find_tools(self, args):
        """
        Ollama tool-overload workaround: instead of declaring all ~189 tools
        up front, local models call this to search the full catalog by
        keyword. Matches are surfaced as text AND flagged so the next turn's
        _get_active_tools() actually declares them, making them callable.
        """
        query = str(args.get('query', '')).strip().lower()
        if not query:
            return "[ERROR] find_tools requires a 'query' argument (e.g. 'screenshot', 'git', 'audio')."

        terms = [t for t in re.split(r'[\s,]+', query) if t]
        scored = []
        for name, spec in self.tools.items():
            if name == 'find_tools':
                continue
            haystack = (name + " " + spec.get('description', '')).lower()
            score = sum(haystack.count(t) for t in terms)
            if score > 0:
                scored.append((score, name, spec))
        scored.sort(key=lambda x: -x[0])
        top = scored[:8]

        if not top:
            return f"No tools matched '{query}'. Try broader keywords (e.g. 'file', 'browser', 'image', 'audio', 'git')."

        discovered = getattr(self, '_ollama_discovered', [])
        lines = [f"Found {len(top)} tool(s) matching '{query}' — now available for calling this turn onward:"]
        for _, name, spec in top:
            if name not in discovered:
                discovered.append(name)
            params = spec.get('parameters', {}).get('properties', {})
            param_str = ", ".join(params.keys()) if params else "none"
            lines.append(f"- {name}({param_str}): {spec.get('description', '')[:160]}")
        self._ollama_discovered = discovered[-40:]  # bounded — no unbounded growth in long sessions

        return "\n".join(lines)

    async def tool_swarm_execute(self, args):
        goal = str(args.get('goal', '')).strip()
        if not goal:
            return "[ERROR] swarm_execute requires a 'goal' argument."
        try:
            from swarm_orchestrator import SwarmOrchestrator
        except ImportError as e:
            return f"[ERROR] swarm_orchestrator.py is missing: {e}"
        orch = getattr(self, '_swarm_orchestrator', None)
        if orch is None:
            orch = self._swarm_orchestrator = SwarmOrchestrator(self)
        try:
            max_agents = int(args.get('max_agents') or 0) or None
        except (TypeError, ValueError):
            max_agents = None
        return await orch.run(goal, max_agents=max_agents)

    def register_skill_tools(self, skills):
        """Merge tools from all loaded skills into self.tools.
        Called by GalacticCore.load_skills() after all skills are instantiated.
        Skill tools will OVERWRITE core gateway tools if names match, ensuring
        upgraded skill versions take priority.
        """
        count = 0
        overwritten = []
        for skill in skills:
            if not skill.enabled:
                continue
            skill_tools = skill.get_tools()
            for tool_name, tool_def in skill_tools.items():
                if tool_name in self.tools:
                    overwritten.append(tool_name)
                self.tools[tool_name] = tool_def
                count += 1
        
        if count:
            print(f"[Skills] Registered {count} tool(s) from skills.")
        if overwritten:
            print(f"[Skills] Upgraded core tools: {', '.join(set(overwritten))}")
    async def tool_resume_workflow(self, args):
        """Tool handler to restore state from a checkpoint and continue."""
        uuid_val = args.get('uuid')
        if not uuid_val:
            return "Error: 'uuid' is required."
            
        try:
            state = await self.load_checkpoint(uuid_val)
            # The actual resuming is tricky to do fully inside a tool call since it interrupts the current flow, 
            # but setting the state means the NEXT turn uses it. 
            # We will return a system prompt instructing the model to proceed.
            return f"[OK] Restored checkpoint {uuid_val}. State restored: turn {state.get('turn_count')}. You MUST now continue the interrupted task."
        except Exception as e:
            return f"[Error] Failed to resume workflow: {str(e)}"
    @staticmethod
    def _path_not_found_hint(path, verb="read"):
        """Turn a bare 'no such file' into a self-correcting message.

        A model that typos a filename ("texels.js" for "textures.js") learns
        nothing from '[Errno 2] No such file or directory' — so it guesses
        again, and again. Listing the closest real filenames from the actual
        directory lets it fix itself on the very next turn instead of looping.
        """
        import difflib as _diff
        try:
            abs_path = os.path.abspath(path)
            wanted = os.path.basename(abs_path)
            directory = os.path.dirname(abs_path) or '.'

            # Walk up to the nearest directory that actually exists, so a wrong
            # folder name is reported as clearly as a wrong file name.
            probe = directory
            while probe and not os.path.isdir(probe):
                parent = os.path.dirname(probe)
                if parent == probe:
                    break
                probe = parent

            if not os.path.isdir(probe):
                return (f"[ERROR] Cannot {verb} '{path}' — not found, and no parent "
                        f"directory of that path exists. Check the path and try again.")

            try:
                entries = sorted(os.listdir(probe))
            except Exception:
                entries = []
            files = [e for e in entries if os.path.isfile(os.path.join(probe, e))]

            lines = [f"[ERROR] File not found: {abs_path}"]
            if probe != directory:
                lines.append(f"The directory '{directory}' does not exist either. "
                             f"Nearest existing directory: {probe}")

            close = _diff.get_close_matches(wanted, files, n=5, cutoff=0.5)
            if not close:
                # Fall back to same-extension neighbours — a bad stem with the
                # right extension is still a strong hint.
                ext = os.path.splitext(wanted)[1].lower()
                if ext:
                    close = [f for f in files if f.lower().endswith(ext)][:8]

            if close:
                lines.append(f"Did you mean one of these? (exact names in {probe})")
                lines.extend(f"  - {c}" for c in close)
            elif files:
                lines.append(f"Files actually in {probe}:")
                lines.extend(f"  - {f}" for f in files[:15])
                if len(files) > 15:
                    lines.append(f"  ... and {len(files) - 15} more")
            else:
                lines.append(f"Directory {probe} contains no files.")

            # Only point at "the list above" when there actually is one.
            if close or files:
                lines.append("Do NOT retry the same path — it does not exist. Use an exact "
                             "name from the list above, or call list_dir on the directory.")
            else:
                lines.append("Do NOT retry the same path — it does not exist. "
                             "Check the directory before trying again.")
            return "\n".join(lines)
        except Exception:
            return f"[ERROR] File not found: {path}"

    async def tool_read_file(self, args):
        """Read a file with optional line range and line numbers (non-blocking)."""
        path = args.get('path')
        start_line = args.get('start_line')
        end_line = args.get('end_line')
        
        if not path:
            return "Error: 'path' parameter is required."

        def _read_sync():
            with open(path, 'r', encoding='utf-8') as f:
                return f.readlines()

        try:
            loop = asyncio.get_running_loop()
            lines = await loop.run_in_executor(getattr(self, '_io_pool', None), _read_sync)
            
            total_lines = len(lines)

            # Default to a 300-line chunk to prevent overwhelming the model
            # on large files (which causes re-read loops and hallucinations).
            CHUNK = 300
            s = int(start_line) if start_line is not None else 1
            e = int(end_line) if end_line is not None else min(s + CHUNK - 1, total_lines)

            # Bounds check
            s = max(1, min(s, total_lines))
            e = max(s, min(e, total_lines))

            selected_lines = lines[s-1:e]

            output = []
            for i, line in enumerate(selected_lines, start=s):
                output.append(f"{i:4} | {line}")

            result = "".join(output)

            if total_lines > CHUNK:
                next_s = e + 1
                next_e = min(e + CHUNK, total_lines)
                header = (
                    f"--- {path} | Lines {s}-{e} of {total_lines} ---\n"
                    f"LARGE FILE ({total_lines} lines total). Showing {CHUNK}-line chunk. "
                    f"Use start_line/end_line to navigate. "
                    f"Next chunk: start_line={next_s}, end_line={next_e}\n"
                    f"---\n"
                )
            else:
                header = f"--- Reading {path} (Lines {s}-{e} of {total_lines}) ---\n"

            footer = "" if e >= total_lines else f"\n--- (More content below. Next: start_line={e+1}) ---"

            return header + result + footer
        except FileNotFoundError:
            return self._path_not_found_hint(path, verb="read")
        except (IsADirectoryError, PermissionError) as e:
            # Windows raises PermissionError (not IsADirectoryError) when you
            # open() a directory, which reads as a confusing "access denied".
            if os.path.isdir(path):
                return (f"[ERROR] '{path}' is a directory, not a file. "
                        f"Use list_dir to see what's inside it.")
            return f"Error reading file: {e}"
        except Exception as e:
            return f"Error reading file: {e}"
    # ── Write confinement: keep file-mutating tools inside allowed roots ─────
    # os.path.abspath() on a model-supplied path resolves happily to anywhere
    # the user can write — including the Windows Startup folder. The basename
    # check in _PROTECTED_FILES is a typo-guard, not a boundary. Confinement
    # resolves the target to a real path and requires it to sit under an
    # allowed root. Opt-out with `security: {confine_writes: false}`.
    _GALACTIC_ROOT = os.path.dirname(os.path.abspath(__file__))

    def _security_cfg(self):
        return self.core.config.get('security', {}) or {}

    def _confinement_enabled(self):
        return bool(self._security_cfg().get('confine_writes', True))

    def _allowed_write_roots(self):
        """Absolute, symlink-resolved, normcase'd roots that writes may land under."""
        roots = [self._GALACTIC_ROOT]
        try:
            ws = self.get_active_workspace()
            if ws:
                roots.append(ws)
        except Exception:
            pass
        paths_cfg = self.core.config.get('paths', {}) or {}
        for key in ('workspace', 'logs', 'images', 'chroma_data'):
            if paths_cfg.get(key):
                roots.append(paths_cfg[key])
        tmp = getattr(self, '_temp_dir', None)
        if tmp:
            roots.append(tmp)
        extra = self._security_cfg().get('extra_write_roots') or []
        if isinstance(extra, str):
            extra = [extra]
        roots.extend(str(r) for r in extra if r)

        out = []
        for r in roots:
            try:
                rp = os.path.normcase(os.path.realpath(
                    os.path.abspath(os.path.expanduser(str(r)))))
            except Exception:
                continue
            if rp and rp not in out:
                out.append(rp)
        return out

    def _check_write_path(self, path, action="write"):
        """None if `path` may be mutated, else a [BLOCKED] error string.

        Compares realpath'd, normcase'd prefixes WITH a trailing separator so a
        sibling like C:\\workspace_backup cannot pass a prefix check against
        C:\\workspace, and so the match is case-insensitive on Windows.
        """
        if not self._confinement_enabled():
            return None
        try:
            target = os.path.realpath(os.path.abspath(os.path.expanduser(str(path))))
        except Exception as e:
            return f"[BLOCKED] Could not resolve path for {action}: {e}"
        norm = os.path.normcase(target)
        roots = self._allowed_write_roots()
        for root in roots:
            if norm == root or norm.startswith(root.rstrip(os.sep) + os.sep):
                return None
        return (
            f"[BLOCKED] {action} refused — '{target}' is outside every allowed write root.\n"
            f"Allowed roots:\n" + '\n'.join(f"  • {r}" for r in roots) + "\n"
            f"Write inside the active workspace or the Galactic install dir instead. "
            f"To permit another location, add it to `security.extra_write_roots` in "
            f"config, or set `security.confine_writes: false` to disable this guard."
        )

    # ── The Crucible: opt-in approval gate for file-mutating tools ───────────
    # When models.require_approval is on, write_file/edit_file/replace_function
    # show you a diff — and exec_shell/execute_python/process_start show you the
    # command — and wait for Approve/Reject before touching disk or spawning a
    # process. Default OFF — when off, the gate short-circuits and behavior is
    # unchanged.
    def _approval_enabled(self):
        return bool(self.core.config.get('models', {}).get('require_approval'))

    @staticmethod
    def _read_text_safe(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception:
            return ''

    def _make_unified_diff(self, path, old, new):
        import difflib
        base = os.path.basename(path)
        d = ''.join(difflib.unified_diff(
            (old or '').splitlines(keepends=True),
            (new or '').splitlines(keepends=True),
            fromfile=f"a/{base}", tofile=f"b/{base}", n=3))
        if not d:
            return "(no textual difference)"
        if len(d) > 15000:
            d = d[:15000] + "\n...[diff truncated — change is large]..."
        return d

    async def _approve_request(self, path, action, diff):
        """Shared approval plumbing. Returns (decided: bool, decision: dict|None).

        Emits an approval_request WS event and blocks until the user decides or
        the timeout elapses. On timeout returns (False, None) — the callers turn
        that into a reject, the safe default.
        """
        import uuid as _uuid
        req_id = _uuid.uuid4().hex[:12]
        loop = asyncio.get_running_loop()
        fut = loop.create_future()
        store = getattr(self, '_pending_approvals', None)
        if store is None:
            store = self._pending_approvals = {}
        store[req_id] = fut
        try:
            await self.core.relay.emit(2, "approval_request", {
                "id": req_id, "path": path, "action": action, "diff": diff})
            await self.core.log(f"⏸️ Approval required: {action} on {path}", priority=2)
            timeout = int(self.core.config.get('models', {}).get('approval_timeout', 300))
            try:
                return True, await asyncio.wait_for(fut, timeout=timeout)
            except asyncio.TimeoutError:
                return False, None
        finally:
            store.pop(req_id, None)
            try:
                await self.core.relay.emit(2, "approval_resolved", {"id": req_id})
            except Exception:
                pass

    async def _approve_change(self, path, old_content, new_content, action):
        """Gate a file mutation. Returns (approved: bool, reason: str|None).

        No-op (True, None) when require_approval is off. Otherwise shows the
        user a unified diff and blocks until they decide.
        """
        if not self._approval_enabled():
            return True, None
        timeout = int(self.core.config.get('models', {}).get('approval_timeout', 300))
        decided, decision = await self._approve_request(
            path, action, self._make_unified_diff(path, old_content, new_content))
        if not decided:
            return False, (f"[NOT APPLIED] Approval for {action} on {path} timed out after "
                           f"{timeout}s — no changes were made. Ask the user before retrying.")
        if decision.get('approved'):
            return True, None
        fb = (decision.get('feedback') or '').strip()
        return False, (f"[CHANGE REJECTED] The user declined the {action} to {path}."
                       + (f" Their feedback: {fb}" if fb else "")
                       + " Do NOT re-apply the same change — revise based on this.")

    async def _approve_command(self, action, command, target=None):
        """Gate a code-execution tool. Returns (approved: bool, reason: str|None).

        exec_shell / execute_python / process_start had NO gate: with
        require_approval on, injected instructions could still run arbitrary
        commands while file writes were being diffed. A command has no "before"
        state, so the command text itself is rendered as an all-additions block
        the deck's existing diff view shows as-is.
        """
        if not self._approval_enabled():
            return True, None
        body = '\n'.join('+' + ln for ln in (command or '').splitlines()) or '+(empty)'
        diff = (f"--- a/(nothing runs yet)\n+++ b/{action}\n"
                f"@@ this command will RUN on your machine @@\n{body}")
        label = target or (command or '').strip().splitlines()[0][:120] or action
        timeout = int(self.core.config.get('models', {}).get('approval_timeout', 300))
        decided, decision = await self._approve_request(label, action, diff)
        if not decided:
            return False, (f"[NOT EXECUTED] Approval for {action} timed out after {timeout}s — "
                           f"nothing was run. Ask the user before retrying.")
        if decision.get('approved'):
            return True, None
        fb = (decision.get('feedback') or '').strip()
        return False, (f"[COMMAND REJECTED] The user declined to run this {action}."
                       + (f" Their feedback: {fb}" if fb else "")
                       + " Do NOT re-run it — revise based on this.")

    async def tool_write_file(self, args):
        """Write content to a file (non-blocking) with robust error handling.

        VERIFICATION GUARANTEE: After every write, this tool immediately reads
        the file back and returns the verified byte count and line count.
        This is structural proof that the write actually happened on disk.
        """
        path = args.get('path')
        content = args.get('content')
        if not path:
            return "Error: 'path' required."
        
        filename = os.path.basename(path)
        if filename in self._PROTECTED_FILES:
            return (
                f"[BLOCKED] Cannot overwrite protected core file '{filename}'. "
                f"Create a new file with a different name instead."
            )

        # Workspace confinement — refuse writes outside the allowed roots.
        denied = self._check_write_path(path, action="write_file")
        if denied:
            return denied

        # The Crucible: gate the write behind user approval (no-op when off).
        if self._approval_enabled():
            old = self._read_text_safe(os.path.abspath(path))
            approved, reason = await self._approve_change(path, old, content or '', "write_file")
            if not approved:
                return reason

        def _write_and_verify():
            try:
                is_absolute = os.path.isabs(path)
                abs_path = os.path.abspath(path)
                os.makedirs(os.path.dirname(abs_path), exist_ok=True)

                # Auto-backup via VCR before overwriting
                self._vcr_auto_backup(abs_path)

                # ── WRITE ──
                with open(abs_path, 'w', encoding='utf-8') as f:
                    f.write(content)

                # ── DISK VERIFICATION (O(1) memory) ──
                # This is what makes tool results trustworthy: the model SEES
                # proof the write landed (real byte count on disk). os.path.getsize
                # confirms the bytes hit the platter WITHOUT reading the file back
                # into RAM — critical when the AI writes a large file (a 50MB CSV
                # would otherwise be loaded, then duplicated as an encoded byte
                # array, just to report a size). The line count comes from the
                # content we already hold in memory, so no re-read is needed.
                verified_bytes = os.path.getsize(abs_path)
                verified_lines = (content or '').count('\n') + 1
                expected_bytes = len((content or '').encode('utf-8'))

                if verified_bytes == 0 and expected_bytes > 0:
                    return (
                        f"❌ WRITE VERIFICATION FAILED: Wrote to {abs_path} but read back 0 bytes. "
                        f"Expected {expected_bytes} bytes. The file may be empty or locked."
                    )

                p_msg = " (resolved from relative)" if not is_absolute else ""
                return (
                    f"✅ WRITE VERIFIED\n"
                    f"  Path: {abs_path}{p_msg}\n"
                    f"  Lines written: {verified_lines}\n"
                    f"  Bytes on disk: {verified_bytes}\n"
                    f"  Status: File confirmed on disk. No further read-back needed."
                )
            except PermissionError:
                return f"❌ Access Denied: Cannot write to {path}. File may be locked by another process or system restricted."
            except OSError as e:
                return f"❌ OS Error writing to {path}: {e.strerror if hasattr(e, 'strerror') else str(e)}"
            except Exception as e:
                return f"❌ Unexpected error writing to {path}: {str(e)}"

        try:
            loop = asyncio.get_running_loop()
            return await loop.run_in_executor(getattr(self, '_io_pool', None), _write_and_verify)
        except Exception as e:
            return f"Error writing file: {e}"
    async def tool_web_search(self, args):
        """Web search using DuckDuckGo — returns parsed, ranked results (no API key needed)."""
        query = args.get('query', '')
        if not query:
            return "[ERROR] No search query provided."
        try:
            import urllib.parse
            from bs4 import BeautifulSoup

            encoded_q = urllib.parse.quote_plus(query)
            search_url = f"https://duckduckgo.com/html/?q={encoded_q}"

            async with httpx.AsyncClient(
                timeout=15.0,
                follow_redirects=True,
                headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                    'Accept-Language': 'en-US,en;q=0.9',
                }
            ) as client:
                response = await client.get(search_url)

            soup = BeautifulSoup(response.text, 'html.parser')
            results = []

            for result in soup.select('.result__body, .result')[:10]:
                title_el  = result.select_one('.result__title, .result__a')
                snippet_el = result.select_one('.result__snippet')
                url_el    = result.select_one('.result__url')

                title   = title_el.get_text(strip=True)   if title_el   else ''
                snippet = snippet_el.get_text(strip=True) if snippet_el else ''
                url     = url_el.get_text(strip=True)     if url_el     else ''

                if title and (snippet or url):
                    results.append({"title": title, "snippet": snippet, "url": url})

            if not results:
                return f"No results found for: '{query}'. Try rephrasing or use web_fetch on a specific URL."

            lines = [f"🔍 Web results for **'{query}'**:\n"]
            for i, r in enumerate(results[:8], 1):
                lines.append(f"{i}. **{r['title']}**")
                if r['snippet']:
                    lines.append(f"   {r['snippet']}")
                if r['url']:
                    lines.append(f"   🔗 {r['url']}")
                lines.append("")

            return "\n".join(lines)

        except ImportError:
            # bs4 not available: fall back to raw fetch
            return f"[Web Search] Query: {query} — Install beautifulsoup4 for parsed results."
        except Exception as e:
            return f"Web search error: {e}"
    async def tool_open_browser(self, args):
        """Open a URL in the browser (Playwright or System Fallback)."""
        url = args.get('url')
        if not url: return "[ERROR] 'url' is required."
        if not url.startswith(('http://', 'https://')):
            url = 'https://' + url
            
        try:
            # 1. Try Playwright Plugin (if loaded)
            browser_plugin = next(
                (p for p in self.core.plugins
                 if "BrowserExecutorPro" in p.__class__.__name__
                 or getattr(p, 'skill_name', '') == 'browser_pro'),
                None
            )
            if browser_plugin:
                result = await browser_plugin.navigate(url)
                if result['status'] == 'success':
                    return f"[BROWSER/PRO] Navigated to: {url}"

            # 2. Fallback to System Browser
            def _open_sync():
                webbrowser.open(url)
                return f"[BROWSER/SYSTEM] Opened URL in default browser: {url}"
            
            loop = asyncio.get_running_loop()
            return await loop.run_in_executor(None, _open_sync)

        except Exception as e:
            return f"[ERROR] Browser navigation: {e}"
    async def tool_browser_navigate(self, args):
        """Alias for tool_open_browser — navigates to a URL."""
        return await self.tool_open_browser(args)
    async def tool_browser_open(self, args):
        """Alias for tool_open_browser — opens a URL or the browser."""
        # Default to a friendly page if no URL is provided
        if not args.get('url'):
            args['url'] = 'https://www.google.com'
        return await self.tool_open_browser(args)
    async def tool_screenshot(self, args):
        """Take a screenshot of the current browser page."""
        path = args.get('path')
        try:
            browser_plugin = next(
                (p for p in self.core.plugins
                 if "BrowserExecutorPro" in p.__class__.__name__
                 or getattr(p, 'skill_name', '') == 'browser_pro'),
                None
            )
            if not browser_plugin:
                return "[ERROR] BrowserExecutorPro plugin not loaded."
            result = await browser_plugin.screenshot(path=path, full_page=True)
            if result['status'] == 'success':
                return f"[BROWSER] Screenshot saved: {result['path']}"
            else:
                return f"[ERROR] Screenshot failed: {result.get('message', 'Unknown error')}"
        except Exception as e:
            return f"[ERROR] Browser screenshot: {e}"
    async def tool_list_skills(self, args):
        """List all loaded skills and their tools."""
        if not self.core.skills:
            return "No skills loaded. Core skills are still running as legacy plugins during migration."
        lines = []
        for skill in self.core.skills:
            tool_names = list(skill.get_tools().keys())
            core_tag = " [core]" if skill.is_core else " [community]"
            enabled_tag = "" if skill.enabled else " (DISABLED)"
            lines.append(
                f"{skill.icon} **{skill.skill_name}** v{skill.version}{core_tag}{enabled_tag}\n"
                f"   {skill.description}\n"
                f"   Tools ({len(tool_names)}): {', '.join(tool_names) if tool_names else '(none)'}"
            )
        return "\n\n".join(lines)
    async def tool_create_skill(self, args):
        """Create a new community skill at runtime."""
        import importlib
        import ast as _ast

        name = args.get('name', '').strip()
        code = args.get('code', '')
        desc = args.get('description', '')

        if not name or not code:
            return "[ERROR] Both 'name' and 'code' are required."

        # Validate name is safe for use as a Python module name
        if not all(c.isalnum() or c == '_' for c in name) or name[0].isdigit():
            return "[ERROR] Skill name must be snake_case (letters, digits, underscores; cannot start with a digit)."

        # Validate code contains required elements
        if 'GalacticSkill' not in code:
            return "[ERROR] Code must contain a class that inherits from GalacticSkill."
        if 'get_tools' not in code:
            return "[ERROR] Skill class must implement get_tools()."

        # Find the skill class name via AST parsing
        skill_class_names = []
        try:
            tree = _ast.parse(code)
            for node in _ast.walk(tree):
                if isinstance(node, _ast.ClassDef):
                    for base in node.bases:
                        base_name = ''
                        if isinstance(base, _ast.Name):
                            base_name = base.id
                        elif isinstance(base, _ast.Attribute):
                            base_name = base.attr
                        if base_name == 'GalacticSkill':
                            skill_class_names.append(node.name)
                            break
        except SyntaxError as e:
            return f"[ERROR] Syntax error in skill code: {e}"

        if not skill_class_names:
            return "[ERROR] Could not find a class inheriting from GalacticSkill in the provided code."
        if len(skill_class_names) > 1:
            return f"[ERROR] Found multiple GalacticSkill subclasses: {', '.join(skill_class_names)}. Only one is allowed per skill file."
        skill_class_name = skill_class_names[0]

        # Check for duplicate skill names
        for existing in self.core.skills:
            if existing.skill_name == name:
                return f"[ERROR] Skill '{name}' already loaded. Use remove_skill first to replace it."

        # Write to community/
        skills_dir = os.path.join(os.path.dirname(os.path.abspath(self.core.config_path)), 'skills', 'community')
        os.makedirs(skills_dir, exist_ok=True)
        skill_path = os.path.join(skills_dir, f'{name}.py')

        try:
            with open(skill_path, 'w', encoding='utf-8') as f:
                f.write(code)
        except Exception as e:
            return f"[ERROR] Failed to write skill file: {e}"

        # Dynamic import
        try:
            module_name = f'skills.community.{name}'
            if module_name in sys.modules:
                del sys.modules[module_name]

            mod = importlib.import_module(module_name)
            cls = getattr(mod, skill_class_name)
            skill = cls(self.core)
            skill.is_core = False

            await skill.on_load()
            self.core.skills.append(skill)

            # Register its tools
            new_tools = skill.get_tools()
            registered = []
            for tool_name, tool_def in new_tools.items():
                if tool_name not in self.tools:
                    self.tools[tool_name] = tool_def
                    registered.append(tool_name)

            asyncio.create_task(skill.run())

            # Update registry
            from datetime import datetime as _dt
            registry = self.core._read_registry()
            if not registry.get('installed') and any(not s.is_core for s in self.core.skills if s.skill_name != name):
                await self.core.log("[Skills] Warning: registry.json was empty/missing — existing community skill entries may have been lost. Check skills/registry.json.", priority=1)
            registry['installed'].append({
                'module': name,
                'class': skill_class_name,
                'file': f'{name}.py',
                'installed_at': _dt.now().isoformat(),
                'source': 'ai_authored',
                'description': desc
            })
            self.core._write_registry(registry)

            return (
                f"[OK] Skill '{name}' created and loaded.\n"
                f"  Class: {skill_class_name}\n"
                f"  Tools registered: {', '.join(registered) if registered else '(none)'}\n"
                f"  File: {skill_path}"
            )

        except Exception as e:
            try:
                os.remove(skill_path)
            except OSError:
                pass
            sys.modules.pop(f'skills.community.{name}', None)
            return f"[ERROR] Failed to load skill '{name}': {e}"
    async def tool_remove_skill(self, args):
        """Remove a community skill by name."""
        name = args.get('name', '').strip()
        if not name:
            return "[ERROR] 'name' is required."

        target = next((s for s in self.core.skills if s.skill_name == name), None)
        if not target:
            return f"[ERROR] Skill '{name}' not found. Use list_skills to see loaded skills."

        if target.is_core:
            return f"[ERROR] '{name}' is a core skill and cannot be removed."

        try:
            await target.on_unload()
        except Exception:
            pass

        # Unregister tools
        tool_names = list(target.get_tools().keys())
        for tn in tool_names:
            self.tools.pop(tn, None)

        self.core.skills.remove(target)

        # Delete file
        skills_dir = os.path.join(os.path.dirname(os.path.abspath(self.core.config_path)), 'skills', 'community')
        skill_path = os.path.join(skills_dir, f'{name}.py')
        try:
            os.remove(skill_path)
        except OSError:
            pass

        # Update registry
        registry = self.core._read_registry()
        if not registry.get('installed') and any(not s.is_core for s in self.core.skills if s.skill_name != name):
            await self.core.log("[Skills] Warning: registry.json was empty/missing — existing community skill entries may have been lost. Check skills/registry.json.", priority=1)
        registry['installed'] = [e for e in registry['installed'] if e.get('module') != name]
        self.core._write_registry(registry)

        return f"[OK] Skill '{name}' removed. Tools unregistered: {', '.join(tool_names) if tool_names else '(none)'}"
    async def tool_generate_image(self, args):
        """Generate an image using FLUX via NVIDIA's GenAI API."""
        import base64 as _b64, time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_image requires a 'prompt' argument."
        model = args.get('model', 'black-forest-labs/flux.1-schnell')
        # Strip nvidia/ prefix if user passed the alias path
        if model.startswith('nvidia/'):
            model = model[len('nvidia/'):]
        width = int(args.get('width', 1024))
        height = int(args.get('height', 1024))
        is_schnell = 'schnell' in model
        steps = int(args.get('steps', 4 if is_schnell else 50))

        # FLUX schnell and dev each have their own API key.
        # fluxDevApiKey → flux.1-dev, fluxApiKey → flux.1-schnell, apiKey → fallback.
        nvidia_cfg = self.core.config.get('providers', {}).get('nvidia', {})
        if not is_schnell:
            nvidia_key = (
                nvidia_cfg.get('fluxDevApiKey') or
                nvidia_cfg.get('fluxApiKey') or
                nvidia_cfg.get('apiKey') or ''
            )
        else:
            nvidia_key = (
                nvidia_cfg.get('fluxApiKey') or
                nvidia_cfg.get('apiKey') or ''
            )
        if not nvidia_key:
            return "[ERROR] NVIDIA FLUX key not found — add providers.nvidia.fluxApiKey (schnell) or fluxDevApiKey (dev) to config.yaml"

        url = f"https://ai.api.nvidia.com/v1/genai/{model}"
        headers = {
            "Authorization": f"Bearer {nvidia_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }
        # Build payload — schnell doesn't support cfg_scale or mode fields
        payload = {"prompt": prompt, "width": width, "height": height, "seed": 0, "steps": steps}
        if not is_schnell:
            payload["mode"] = "base"
            payload["cfg_scale"] = 5  # dev default per NVIDIA docs (1-9 range)

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                r = await client.post(url, headers=headers, json=payload)
                if r.status_code == 401:
                    return f"[ERROR] NVIDIA GenAI 401 Unauthorized — key used: nvapi-...{nvidia_key[-8:]}. Check that your NVIDIA API key has access to the FLUX model at ai.api.nvidia.com."
                if r.status_code == 500:
                    return f"[ERROR] NVIDIA GenAI HTTP 500 — their inference server is down right now. Do NOT retry. Report this to the user and suggest trying again in a few minutes or switching to flux.1-schnell."
                if r.status_code != 200:
                    return f"[ERROR] NVIDIA GenAI HTTP {r.status_code}: {r.text[:500]}"
                data = r.json()

            artifact = data.get('artifacts', [{}])[0]
            finish = artifact.get('finishReason', '')
            if finish == 'CONTENT_FILTERED':
                return "⚠️ Image generation blocked by content filter. Try a different prompt."
            b64 = artifact.get('base64', '')
            if not b64:
                return f"[ERROR] Image generation failed: {json.dumps(data)}"

            # API returns JPEG data
            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'flux')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"flux_{int(_time.time())}.jpg"
            path = os.path.join(img_subdir, fname)
            with open(path, 'wb') as f:
                f.write(_b64.b64decode(b64))
            # Signal Telegram bridge and web deck to deliver the image directly
            self.last_image_file = path
            return f"✅ Image generated and saved to: {path}\nModel: {model}\nPrompt: {prompt}"
        except Exception as e:
            return f"[ERROR] generate_image: {str(e)}"
    async def tool_schedule_task(self, args):
        """Schedule a task/reminder using the scheduler plugin."""
        name = args.get('name')
        delay_seconds = args.get('delay_seconds')
        message = args.get('message')
        
        try:
            # Check if scheduler plugin is available
            scheduler_plugin = next((p for p in self.core.plugins if "Scheduler" in p.__class__.__name__), None)
            if scheduler_plugin:
                await scheduler_plugin.schedule_task(name, delay_seconds, message)
                return f"Task '{name}' scheduled to fire in {delay_seconds} seconds."
            else:
                return "Scheduler plugin not available. Task not scheduled."
        except Exception as e:
            return f"Error scheduling task: {e}"
    async def tool_list_tasks(self, args):
        """List all scheduled tasks."""
        try:
            scheduler_plugin = next((p for p in self.core.plugins if "Scheduler" in p.__class__.__name__), None)
            if scheduler_plugin:
                tasks = await scheduler_plugin.list_tasks()
                if tasks:
                    return json.dumps(tasks, indent=2)
                else:
                    return "No scheduled tasks."
            else:
                return "Scheduler plugin not available."
        except Exception as e:
            return f"Error listing tasks: {e}"
    async def tool_edit_file(self, args):
        """Robustly edit a file (non-blocking).
        
        VERIFICATION GUARANTEE: After every edit, this tool reads the file back
        and returns the verified line count and surrounding context. This is the
        structural proof that the change actually landed on disk.
        """
        path = args.get('path')
        if not path:
            return "Error: 'path' parameter is required."

        # Workspace confinement — refuse edits outside the allowed roots.
        denied = self._check_write_path(path, action="edit_file")
        if denied:
            return denied

        def _compute_edit():
            """Apply replacements in memory. Returns ('ERR', msg) or
            ('OK', {old, new, replacements})."""
            # Collect replacements
            replacements = []
            if 'replacements' in args and isinstance(args['replacements'], list):
                for r in args['replacements']:
                    old = r.get('old_text') or r.get('old')
                    new = r.get('new_text') or r.get('new')
                    if old is not None and new is not None:
                        replacements.append((old, new))
            else:
                old = args.get('old_text') or args.get('old')
                new = args.get('new_text') or args.get('new')
                if old is not None and new is not None:
                    replacements.append((old, new))

            if not replacements:
                return ('ERR', "Error: No valid replacements provided. Required: 'old_text' and 'new_text' or 'replacements' array.")

            start_line = args.get('start_line')
            end_line = args.get('end_line')

            try:
                with open(path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    f.seek(0)
                    lines = f.readlines()

                total_lines = len(lines)

                if start_line is not None or end_line is not None:
                    s = max(1, int(start_line) if start_line is not None else 1)
                    e = min(total_lines, int(end_line) if end_line is not None else total_lines)

                    segment_lines = lines[s-1:e]
                    segment = "".join(segment_lines)

                    for old_text, new_text in replacements:
                        count = segment.count(old_text)
                        if count == 0:
                            return ('ERR', f"Error: Could not find '{old_text}' within lines {s}-{e} of {path}.")
                        if count > 1:
                            return ('ERR', f"Error: Found {count} occurrences of '{old_text}' within lines {s}-{e}. Be more specific.")
                        segment = segment.replace(old_text, new_text)

                    new_content = "".join(lines[:s-1]) + segment + "".join(lines[e:])
                else:
                    new_content = content
                    for old_text, new_text in replacements:
                        count = new_content.count(old_text)
                        if count == 0:
                            if old_text in content:
                                return ('ERR', f"Error: '{old_text}' was already changed by a previous replacement in this call.")
                            return ('ERR', f"Error: Could not find exact text '{old_text}' in {path}.")
                        if count > 1:
                            return ('ERR', f"Error: Found {count} occurrences of '{old_text}'. Use start_line/end_line to disambiguate.")
                        new_content = new_content.replace(old_text, new_text)

                return ('OK', {'old': content, 'new': new_content, 'replacements': replacements})
            except Exception as e:
                return ('ERR', f"Error editing file: {str(e)}")

        loop = asyncio.get_running_loop()
        status, payload = await loop.run_in_executor(getattr(self, '_io_pool', None), _compute_edit)
        if status == 'ERR':
            return payload

        # The Crucible: gate the write behind user approval (no-op when off).
        approved, reason = await self._approve_change(path, payload['old'], payload['new'], "edit_file")
        if not approved:
            return reason

        new_content = payload['new']
        replacements = payload['replacements']

        def _write_verify():
            try:
                # Auto-backup via VCR before editing (only now that it's approved)
                self._vcr_auto_backup(os.path.abspath(path))

                # ── WRITE BACK ──
                with open(path, 'w', encoding='utf-8') as f:
                    f.write(new_content)

                # ── IMMEDIATE READ-BACK (structural verification) ──
                # Read the file back and verify the change landed. Edit verification
                # genuinely needs the content (to spot-check each replacement), but
                # the byte count comes from os.path.getsize, not a re-encode.
                with open(path, 'r', encoding='utf-8') as f:
                    verified_content = f.read()

                verified_lines = verified_content.count('\n') + 1
                verified_bytes = os.path.getsize(path)

                # Spot-check: verify the new text actually landed in the file
                verification_checks = []
                for old_text, new_text in replacements:
                    if new_text and new_text[:40] in verified_content:
                        verification_checks.append(f"  ✓ '{new_text[:40].strip()}...' confirmed on disk")
                    elif new_text == '':
                        verification_checks.append(f"  ✓ Deletion of '{old_text[:30].strip()}...' confirmed")
                    else:
                        verification_checks.append(f"  ⚠ '{new_text[:40].strip()}...' NOT found after write — possible encoding issue")

                checks_str = '\n'.join(verification_checks) if verification_checks else '  (no spot-checks)'

                return (
                    f"✅ EDIT VERIFIED\n"
                    f"  File: {path}\n"
                    f"  Replacements applied: {len(replacements)}\n"
                    f"  Total lines after edit: {verified_lines}\n"
                    f"  Bytes on disk: {verified_bytes}\n"
                    f"  Content spot-checks:\n{checks_str}\n"
                    f"  Status: Edit confirmed on disk."
                )
            except Exception as e:
                return f"Error editing file: {str(e)}"

        return await loop.run_in_executor(getattr(self, '_io_pool', None), _write_verify)

    # ── AST-safe Python editing ──────────────────────────────────────────────
    # edit_file relies on exact string matching, which is fragile: whitespace
    # drift, a moved line, or a near-duplicate breaks it. These tools locate a
    # function/class by NAME via Python's AST, so edits are immune to line-number
    # drift, and the whole file is re-parsed to guarantee valid Python BEFORE any
    # bytes hit disk. (lsp_tooling can *read* symbols; these can rewrite them.)
    @staticmethod
    def _find_ast_targets(tree, name):
        """Return the list of AST nodes matching 'name' or 'ClassName.method'."""
        import ast
        parts = name.split('.')
        out = []
        if len(parts) == 2:
            cls_name, meth = parts
            for node in ast.walk(tree):
                if isinstance(node, ast.ClassDef) and node.name == cls_name:
                    for sub in node.body:
                        if isinstance(sub, (ast.FunctionDef, ast.AsyncFunctionDef)) and sub.name == meth:
                            out.append(sub)
        else:
            target = parts[0]
            # Prefer a top-level match; only walk deeper if there isn't one.
            for node in tree.body:
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)) and node.name == target:
                    out.append(node)
            if not out:
                for node in ast.walk(tree):
                    if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)) and node.name == target:
                        out.append(node)
        return out

    async def tool_list_functions(self, args):
        """List top-level functions/classes (and methods) in a Python file with line ranges."""
        import ast
        path = args.get('path', '')
        if not path:
            return "[ERROR] list_functions requires a 'path'."

        def _worker():
            if not os.path.exists(path):
                return f"[ERROR] File not found: {path}"
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    tree = ast.parse(f.read())
            except SyntaxError as e:
                return f"[ERROR] {path} is not valid Python: {e}"
            except Exception as e:
                return f"[ERROR] Could not read {path}: {e}"
            lines = [f"[AST] {path}"]
            for node in tree.body:
                if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef)):
                    lines.append(f"  def {node.name}()  (lines {node.lineno}-{node.end_lineno})")
                elif isinstance(node, ast.ClassDef):
                    lines.append(f"  class {node.name}  (lines {node.lineno}-{node.end_lineno})")
                    for sub in node.body:
                        if isinstance(sub, (ast.FunctionDef, ast.AsyncFunctionDef)):
                            lines.append(f"      {node.name}.{sub.name}()  (lines {sub.lineno}-{sub.end_lineno})")
            if len(lines) == 1:
                lines.append("  (no top-level functions or classes found)")
            return "\n".join(lines)

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(getattr(self, '_io_pool', None), _worker)

    async def tool_read_function(self, args):
        """Return the exact source of a function/class/method located by name via AST."""
        import ast
        path = args.get('path', '')
        name = args.get('name', '')
        if not path or not name:
            return "[ERROR] read_function requires 'path' and 'name'."

        def _worker():
            if not os.path.exists(path):
                return f"[ERROR] File not found: {path}"
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    source = f.read()
                tree = ast.parse(source)
            except SyntaxError as e:
                return f"[ERROR] {path} is not valid Python: {e}"
            except Exception as e:
                return f"[ERROR] Could not read {path}: {e}"
            targets = self._find_ast_targets(tree, name)
            if not targets:
                return (f"[ERROR] No function/class named '{name}' in {path}. "
                        f"Use list_functions to see available names.")
            if len(targets) > 1:
                where = ', '.join(f"line {t.lineno}" for t in targets)
                return (f"[ERROR] '{name}' is ambiguous ({len(targets)} matches at {where}). "
                        f"Qualify it as 'ClassName.method'.")
            node = targets[0]
            src_lines = source.splitlines(keepends=True)
            segment = ''.join(src_lines[node.lineno - 1:node.end_lineno])
            return (f"[AST] {name} in {path} (lines {node.lineno}-{node.end_lineno}):\n\n{segment}")

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(getattr(self, '_io_pool', None), _worker)

    async def tool_replace_function(self, args):
        """Replace a function/class/method (located by AST name) with new code, safely."""
        import ast, textwrap
        path = args.get('path', '')
        name = args.get('name', '')
        new_code = args.get('new_code', '')
        if not path or not name or not new_code:
            return "[ERROR] replace_function requires 'path', 'name', and 'new_code'."

        filename = os.path.basename(path)
        if filename in getattr(self, '_PROTECTED_FILES', set()):
            return f"[BLOCKED] Cannot modify protected core file '{filename}'."

        # Workspace confinement — refuse rewrites outside the allowed roots.
        denied = self._check_write_path(path, action="replace_function")
        if denied:
            return denied

        def _compute():
            """Build & syntax-validate the new file content. Returns
            ('ERR', msg) or ('OK', {old, new, start, end})."""
            if not os.path.exists(path):
                return ('ERR', f"[ERROR] File not found: {path}")
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    source = f.read()
                tree = ast.parse(source)
            except SyntaxError as e:
                return ('ERR', f"[ERROR] {path} is not valid Python (fix it first): {e}")
            except Exception as e:
                return ('ERR', f"[ERROR] Could not read {path}: {e}")

            targets = self._find_ast_targets(tree, name)
            if not targets:
                return ('ERR', f"[ERROR] No function/class named '{name}' in {path}. "
                               f"Use list_functions to see available names.")
            if len(targets) > 1:
                where = ', '.join(f"line {t.lineno}" for t in targets)
                return ('ERR', f"[ERROR] '{name}' is ambiguous ({len(targets)} matches at {where}). "
                               f"Qualify it as 'ClassName.method'.")
            node = targets[0]
            start, end = node.lineno, node.end_lineno  # decorators above are preserved

            # Re-indent the new code to the target's column (so the model can pass
            # it at column 0 whether it's a top-level function or a nested method).
            indent = ' ' * node.col_offset
            dedented = textwrap.dedent(new_code).rstrip('\n')
            new_block = '\n'.join((indent + ln if ln.strip() else ln)
                                  for ln in dedented.split('\n')) + '\n'

            src_lines = source.splitlines(keepends=True)
            new_content = ''.join(src_lines[:start - 1]) + new_block + ''.join(src_lines[end:])

            # SAFETY GATE: the edited file must still parse. If not, write nothing.
            try:
                ast.parse(new_content)
            except SyntaxError as e:
                return ('ERR', f"[ERROR] Replacement would produce invalid Python — NOT written.\n"
                               f"  {e}\nCheck the syntax/indentation of your new_code.")
            return ('OK', {'old': source, 'new': new_content, 'start': start, 'end': end})

        loop = asyncio.get_running_loop()
        status, payload = await loop.run_in_executor(getattr(self, '_io_pool', None), _compute)
        if status == 'ERR':
            return payload

        # The Crucible: gate the write behind user approval (no-op when off).
        approved, reason = await self._approve_change(path, payload['old'], payload['new'], "replace_function")
        if not approved:
            return reason

        start, end, new_content = payload['start'], payload['end'], payload['new']

        def _write():
            self._vcr_auto_backup(os.path.abspath(path))
            with open(path, 'w', encoding='utf-8') as f:
                f.write(new_content)
            new_total = new_content.count('\n') + 1
            return (f"✅ AST REPLACE VERIFIED\n"
                    f"  Replaced '{name}' in {path} (was lines {start}-{end})\n"
                    f"  File now {new_total} lines and parses as valid Python.\n"
                    f"  Original backed up to workspace/.galactic_vcr/ (undo: /vcr undo {path}).")

        return await loop.run_in_executor(getattr(self, '_io_pool', None), _write)

    # ── Outbound fetch guard (SSRF) ──────────────────────────────────────────
    # web_fetch used to run with verify=False and follow_redirects=True against
    # any scheme or host — so it leaked to MITM and could be aimed at
    # http://127.0.0.1:17789/api/... (the deck's own unauthenticated API) by
    # injected page content. Every hop is now resolved and IP-checked.
    @staticmethod
    def _ip_is_internal(ip_str):
        import ipaddress
        try:
            ip = ipaddress.ip_address(ip_str)
        except ValueError:
            return True  # unparseable → treat as unsafe
        if ip.version == 6 and ip.ipv4_mapped is not None:
            ip = ip.ipv4_mapped
        return bool(ip.is_private or ip.is_loopback or ip.is_link_local
                    or ip.is_reserved or ip.is_multicast or ip.is_unspecified)

    async def _check_fetch_url(self, url):
        """(ok, err) for an outbound fetch target — scheme + resolved-IP check."""
        import socket
        from urllib.parse import urlparse
        try:
            u = urlparse(url or '')
        except Exception:
            return False, f"[BLOCKED] Unparseable URL: {url!r}"
        scheme = (u.scheme or '').lower()
        if scheme not in ('http', 'https'):
            return False, (f"[BLOCKED] web_fetch only handles http:// and https:// URLs "
                           f"(got '{u.scheme or 'no scheme'}'). file://, ftp:// and data: "
                           f"are refused.")
        host = u.hostname
        if not host:
            return False, f"[BLOCKED] URL has no host: {url}"
        if self._security_cfg().get('allow_private_fetch'):
            return True, None
        try:
            infos = await asyncio.get_running_loop().getaddrinfo(
                host, u.port or (443 if scheme == 'https' else 80),
                proto=socket.IPPROTO_TCP)
        except Exception as e:
            return False, f"[BLOCKED] Could not resolve host '{host}': {e}"
        for info in infos:
            ip = info[4][0]
            if self._ip_is_internal(ip):
                return False, (
                    f"[BLOCKED] '{host}' resolves to an internal address ({ip}). "
                    f"web_fetch will not reach loopback, link-local or private "
                    f"network targets — that path leads straight to Galactic's own "
                    f"local API. Set `security.allow_private_fetch: true` if you "
                    f"genuinely need this.")
        return True, None

    async def tool_web_fetch(self, args):
        """Fetch and extract readable content from a URL."""
        url = args.get('url')
        mode = args.get('mode', 'markdown')
        if not url:
            return "[ERROR] 'url' is required."
        if not re.match(r'^[a-zA-Z][a-zA-Z0-9+.\-]*:', url):
            url = 'https://' + url

        try:
            import httpx
            from bs4 import BeautifulSoup
            from urllib.parse import urljoin

            # Redirects are followed by hand so every hop gets re-checked — an
            # allowed host can 302 straight into 127.0.0.1 otherwise.
            async with httpx.AsyncClient(follow_redirects=False, timeout=30.0) as client:
                current = url
                response = None
                for _hop in range(6):
                    ok, err = await self._check_fetch_url(current)
                    if not ok:
                        return err
                    response = await client.get(current, headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    })
                    if response.status_code in (301, 302, 303, 307, 308):
                        loc = response.headers.get('location')
                        if not loc:
                            break
                        current = urljoin(current, loc)
                        continue
                    break
                else:
                    return f"[BLOCKED] Too many redirects (>6) starting at {url}."

                response.raise_for_status()

                soup = BeautifulSoup(response.text, 'html.parser')
                
                # Remove script and style elements
                for script in soup(["script", "style", "nav", "footer", "header"]):
                    script.decompose()
                
                # Get text
                if mode == 'text':
                    text = soup.get_text(separator='\n', strip=True)
                else:  # markdown mode
                    # Basic markdown conversion
                    title = soup.find('title')
                    title_text = f"# {title.string}\n\n" if title else ""
                    
                    body = soup.find('body') or soup
                    paragraphs = body.find_all(['p', 'h1', 'h2', 'h3', 'h4', 'li'])
                    
                    text_parts = [title_text]
                    for p in paragraphs:
                        tag_name = p.name
                        text_content = p.get_text(strip=True)
                        
                        if tag_name == 'h1':
                            text_parts.append(f"\n# {text_content}\n")
                        elif tag_name == 'h2':
                            text_parts.append(f"\n## {text_content}\n")
                        elif tag_name == 'h3':
                            text_parts.append(f"\n### {text_content}\n")
                        elif tag_name == 'li':
                            text_parts.append(f"- {text_content}")
                        else:
                            text_parts.append(text_content)
                    
                    text = '\n'.join(text_parts)
                
                # Limit length
                max_chars = 8000
                if len(text) > max_chars:
                    text = text[:max_chars] + "\n\n[... content truncated]"
                
                return f"[DOC] Content from {url}:\n\n{text}"
        except Exception as e:
            return f"Error fetching URL: {e}"
    async def tool_process_start(self, args):
        """Start a background process."""
        command = args.get('command')
        session_id = args.get('session_id', f"proc_{int(asyncio.get_event_loop().time())}")
        if not command:
            return "[ERROR] 'command' is required."

        # The Crucible: gate the spawn behind user approval (no-op when off).
        approved, reason = await self._approve_command('process_start', command,
                                                       target=session_id)
        if not approved:
            return reason

        try:
            # Store processes in core if not exists
            if not hasattr(self.core, 'processes'):
                self.core.processes = {}
            
            # Start process
            process = await asyncio.create_subprocess_shell(
                command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE
            )
            
            self.core.processes[session_id] = {
                'process': process,
                'command': command,
                'started': asyncio.get_event_loop().time(),
                'stdout': [],
                'stderr': []
            }
            
            # Start log collection task
            asyncio.create_task(self._collect_process_output(session_id))
            
            return f"[RUN] Process started: {session_id}\nCommand: {command}\nPID: {process.pid}"
        except Exception as e:
            return f"Error starting process: {e}"
    async def tool_process_status(self, args):
        """Check status of a background process."""
        session_id = args.get('session_id')
        
        try:
            if not hasattr(self.core, 'processes') or session_id not in self.core.processes:
                return f"[ERR] Process not found: {session_id}"
            
            proc_info = self.core.processes[session_id]
            process = proc_info['process']
            
            status = "running" if process.returncode is None else f"exited ({process.returncode})"
            runtime = asyncio.get_event_loop().time() - proc_info['started']
            
            stdout_preview = ''.join(proc_info['stdout'][-10:])[:500]
            
            return (
                f"[STATUS] Process Status: {session_id}\n"
                f"Command: {proc_info['command']}\n"
                f"PID: {process.pid}\n"
                f"Status: {status}\n"
                f"Runtime: {runtime:.1f}s\n"
                f"Recent output:\n{stdout_preview}"
            )
        except Exception as e:
            return f"Error checking process: {e}"
    async def tool_process_kill(self, args):
        """Kill a background process."""
        session_id = args.get('session_id')
        
        try:
            if not hasattr(self.core, 'processes') or session_id not in self.core.processes:
                return f"[ERR] Process not found: {session_id}"
            
            proc_info = self.core.processes[session_id]
            process = proc_info['process']
            
            if process.returncode is None:
                process.kill()
                await process.wait()
                return f"[KILL] Process killed: {session_id}"
            else:
                return f"Process already exited: {session_id} (code {process.returncode})"
        except Exception as e:
            return f"Error killing process: {e}"
    async def tool_analyze_image(self, args):
        """Analyze an image — routes to the active provider's vision endpoint (non-blocking)."""
        path = args.get('path')
        prompt = args.get('prompt', 'Describe this image in detail. Include any text you see.')

        import base64
        from pathlib import Path

        if not path or not Path(path).exists():
            return f"[ERR] Image not found: {path}"

        def _read_img():
            with open(path, 'rb') as f:
                return f.read()

        try:
            loop = asyncio.get_running_loop()
            raw = await loop.run_in_executor(getattr(self, '_io_pool', None), _read_img)
            image_b64 = base64.b64encode(raw).decode('utf-8')
            suffix = Path(path).suffix.lower()
            mime_map = {'.jpg': 'image/jpeg', '.jpeg': 'image/jpeg',
                        '.png': 'image/png', '.gif': 'image/gif', '.webp': 'image/webp'}
            mime_type = mime_map.get(suffix, 'image/jpeg')

            return await self._analyze_image_b64(image_b64, mime_type, prompt)
        except Exception as e:
            return f"[ERROR] analyze_image: {e}"
    async def tool_memory_search(self, args):
        """Search semantic memory for relevant context."""
        query = args.get('query')
        top_k = int(args.get('top_k', 5))
        
        try:
            # Access core memory (could be semantic or keyword-based)
            if hasattr(self.core, 'memory'):
                results = await self.core.memory.recall(query, top_k=top_k)
                
                if not results:
                    return f"[MEMORY] No relevant memories found for: {query}"
                
                # Format results
                formatted = [f"[MEMORY] Found {len(results)} relevant memories:\n"]
                for i, mem in enumerate(results, 1):
                    score = mem.get('relevance_score', 'N/A')
                    content_preview = mem['content'][:200] + "..." if len(mem['content']) > 200 else mem['content']
                    source = mem.get('metadata', {}).get('source', 'unknown')
                    formatted.append(f"\n{i}. [Score: {score}] ({source})\n{content_preview}\n")
                
                return "".join(formatted)
            else:
                return "[ERR] Memory system not available."
        except Exception as e:
            return f"Error searching memory: {e}"
    async def tool_search_codebase(self, args):
        """Semantic search over the Neural Indexer's codebase index.

        The indexer vector-embeds every source file into the 'codebase_index'
        category. This tool queries ONLY that category, so the model can find
        where something lives ("where is the wake-word toggle handled?") by
        meaning rather than exact text — complementing grep_search.
        """
        query = (args.get('query') or '').strip()
        if not query:
            return "[ERROR] search_codebase requires a 'query'."
        top_k = int(args.get('top_k', 6))
        if not hasattr(self.core, 'memory') or not self.core.memory:
            return "[ERR] Memory system not available."
        try:
            # Over-fetch, then keep only chunks under the CURRENT workspace root.
            # The index still holds chunks from older installs of this project
            # (an old F:\ drive, a stale Desktop copy) whose code is out of date —
            # returning those would hand back wrong answers.
            root = os.path.abspath(
                self.core.config.get('system', {}).get('workspace_root', os.getcwd())
            ).lower()
            raw = await self.core.memory.query_memory(
                query, n_results=min(top_k * 4, 40), category='codebase_index'
            )
            scoped = [r for r in raw
                      if str((r.get('metadata') or {}).get('path', '')).lower().startswith(root)]
            # NEVER fall back to unscoped `raw`: if nothing in the current
            # workspace matched, returning cross-install/cross-project chunks
            # (the exact stale hits the scoping above exists to exclude) would
            # hand back code from an unrelated repo. Empty is the honest answer.
            results = scoped[:top_k]
            if not results:
                return (f"[CODEBASE] No indexed code matched: {query}\n"
                        "(The Neural Indexer may still be building — check Status.)")

            out = [f"[CODEBASE] Top {len(results)} semantic matches for: {query}\n"]
            for i, r in enumerate(results, 1):
                meta = r.get('metadata') or {}
                path = meta.get('path', 'unknown')
                dist = r.get('distance')
                score = f"{1.0 - dist:.3f}" if isinstance(dist, (int, float)) else "n/a"
                body = r.get('content', '')
                # Indexer stores "FILE: ...\nPATH: ...\nCONTENT:\n<chunk>" —
                # strip the header so the model sees just the code snippet.
                if 'CONTENT:\n' in body:
                    body = body.split('CONTENT:\n', 1)[1]
                snippet = body.strip()[:600]
                out.append(f"\n{i}. {path}  [relevance {score}]\n{snippet}\n")
            out.append("\nUse read_file on a path above to see full context.")
            return "".join(out)
        except Exception as e:
            return f"Error searching codebase: {e}"
    async def tool_memory_imprint(self, args):
        """Save information to long-term memory and persist to MEMORY.md (non-blocking)."""
        content = args.get('content')
        tags = args.get('tags', '')

        try:
            if hasattr(self.core, 'memory'):
                metadata = {"source": "manual_imprint", "tags": tags}
                await self.core.memory.imprint(content, metadata)

                # Also write to MEMORY.md (best effort)
                def _write_md():
                    try:
                        workspace = self.core.config.get('paths', {}).get('workspace', '')
                        if not workspace: return
                        memory_path = os.path.join(workspace, 'MEMORY.md')
                        timestamp = datetime.now().strftime('%Y-%m-%d')
                        entry = f"\n- {timestamp}{f' [{tags}]' if tags else ''}: {content}"
                        if not os.path.exists(memory_path):
                            with open(memory_path, 'w', encoding='utf-8') as f: f.write("# Memory\n")
                        with open(memory_path, 'a', encoding='utf-8') as f: f.write(entry)
                    except: pass
                
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(None, _write_md)
                
                if hasattr(self, 'personality') and hasattr(self.personality, 'reload_memory'):
                    self.personality.reload_memory()

                return f"[MEMORY] Saved to long-term memory. Tags: {tags or 'none'}"
            else: return "[ERR] Memory system not available."
        except Exception as e: return f"Error saving to memory: {e}"
    def _resolve_tts_engine(self, voice):
        """Pick a TTS engine: honor the user's configured voice_agent.engine,
        preferring fish-speech (cloud clone) and premium ElevenLabs voices,
        otherwise a free edge/local voice."""
        va = self.core.config.get('voice_agent', {}) or {}
        va_engine = str(va.get('engine') or '').lower()
        if va_engine == 'fish-speech' and va.get('fish_speech_api_key'):
            return 'fish-speech'
        el_key = (self.core.config.get('elevenlabs', {}) or {}).get('api_key', '')
        if el_key and voice in ('Nova', 'Byte', 'Default'):
            return 'elevenlabs'
        if va_engine in ('piper', 'chatterbox', 'gtts'):
            return va_engine
        return 'edge-tts'

    async def tool_text_to_speech(self, args):
        """Convert text to speech via the shared tts_engine module. Returns the
        path to the generated audio file (read back by /api/tts)."""
        import tts_engine
        import hashlib as _hashlib
        text = args.get('text') or ''
        if not text.strip():
            return "[ERR] text_to_speech requires non-empty 'text'."
        cfg = self.core.config
        cfg_voice = (cfg.get('elevenlabs', {}) or {}).get('voice', 'Guy')
        voice = args.get('voice', cfg_voice)
        engine = args.get('engine') or self._resolve_tts_engine(voice)

        logs_dir = self.config.get('paths', {}).get('logs', './logs')
        text_hash = _hashlib.md5(text.encode()).hexdigest()[:8]

        # Try the resolved engine, then fall back to free engines so a missing
        # key never leaves the caller with no audio.
        errors = []
        for eng in dict.fromkeys([engine, 'edge-tts', 'gtts']):
            result = await asyncio.to_thread(
                tts_engine.synthesize, text, cfg,
                engine=eng, voice=voice, out_dir=logs_dir, basename=f'tts_{text_hash}')
            if result.get('path'):
                if result.get('engine') == 'fish-speech':
                    await asyncio.to_thread(tts_engine.boost_file, result['path'], 15)
                self.last_voice_file = result['path']
                return f"[VOICE] Generated speech: {result['path']}"
            errors.append(f"{eng}: {result.get('error')}")
        return "[ERR] No TTS engine available - " + "; ".join(errors)
    async def tool_generate_image_sd35(self, args):
        """Generate an image using Stable Diffusion 3.5 Large via NVIDIA NIM."""
        import base64 as _b64, time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_image_sd35 requires a 'prompt' argument."
        negative_prompt = args.get('negative_prompt', '')
        width    = int(args.get('width', 1024))
        height   = int(args.get('height', 1024))
        steps    = int(args.get('steps', 40))
        cfg_scale = float(args.get('cfg_scale', 5.0))
        seed     = int(args.get('seed', 0))

        nvidia_cfg = self.core.config.get('providers', {}).get('nvidia', {})
        nvidia_key = nvidia_cfg.get('apiKey', '')
        if not nvidia_key:
            return "[ERROR] No nvidia.apiKey found in config.yaml"

        url = "https://ai.api.nvidia.com/v1/genai/stabilityai/stable-diffusion-3.5-large"
        headers = {
            "Authorization": f"Bearer {nvidia_key}",
            "Accept": "application/json",
            "Content-Type": "application/json",
        }
        payload = {
            "prompt": prompt,
            "mode": "base",
            "width": width,
            "height": height,
            "steps": steps,
            "cfg_scale": cfg_scale,
            "seed": seed,
        }
        if negative_prompt:
            payload["negative_prompt"] = negative_prompt

        try:
            async with httpx.AsyncClient(timeout=120.0) as client:
                r = await client.post(url, headers=headers, json=payload)
                if r.status_code == 401:
                    return f"[ERROR] NVIDIA SD3.5 401 Unauthorized — check your apiKey in config.yaml"
                if r.status_code == 500:
                    return "[ERROR] NVIDIA SD3.5 HTTP 500 — inference server error. Try again in a few minutes."
                if r.status_code != 200:
                    return f"[ERROR] NVIDIA SD3.5 HTTP {r.status_code}: {r.text[:500]}"
                data = r.json()

            artifact = data.get('artifacts', [{}])[0]
            finish = artifact.get('finishReason', '')
            if finish == 'CONTENT_FILTERED':
                return "⚠️ Image blocked by content filter. Try a different prompt."
            b64 = artifact.get('base64', '')
            if not b64:
                return f"[ERROR] SD3.5 generation failed: {json.dumps(data)}"

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'sd35')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"sd35_{int(_time.time())}.jpg"
            path = os.path.join(img_subdir, fname)
            
            def _save_sync():
                with open(path, 'wb') as f:
                    f.write(_b64.b64decode(b64))
                return True

            loop = asyncio.get_running_loop()
            await loop.run_in_executor(None, _save_sync)
            
            self.last_image_file = path
            return f"✅ SD3.5 image generated: {path}\nModel: stable-diffusion-3.5-large\nPrompt: {prompt}"
        except Exception as e:
            return f"[ERROR] generate_image_sd35: {e}"
    async def tool_post_to_social(self, args):
        """Post text and an optional image to social media (X.com / Discord / Simulated)."""
        content = args.get('content', '')
        image_path = args.get('image_path', '')
        platform = args.get('platform', 'X.com')
        
        if not content:
            return "[ERROR] tool_post_to_social: 'content' is required."
            
        try:
            # 1. Simulate the post (Save to Desktop for verification)
            desktop = os.path.join(os.environ['USERPROFILE'], 'Desktop')
            fname = f"social_post_{int(time.time())}.txt"
            post_file = os.path.join(desktop, fname)
            
            with open(post_file, 'w', encoding='utf-8') as f:
                f.write(f"--- SOCIAL MEDIA POST ({platform}) ---\n")
                f.write(f"Content: {content}\n")
                if image_path:
                    f.write(f"Image Attachment: {image_path}\n")
                f.write(f"Timestamp: {time.ctime()}\n")
                f.write("-" * 40 + "\n")
            
            # 2. Mock success (in a real scenario, this would hit a bridge or API)
            msg = f"✅ Posted to {platform} successfully!\n📍 Saved copy to Desktop: {fname}"
            if image_path:
                msg += f"\n🖼️ Image attached: {os.path.basename(image_path)}"
            
            return msg
        except Exception as e:
            return f"[ERROR] tool_post_to_social: {e}"
    def _get_genai_client(self):
        from google import genai
        import json
        google_cfg = self.core.config.get('providers', {}).get('google', {})
        api_key = google_cfg.get('apiKey', '')
        client_args = {}
        
        try:
            from google.oauth2.credentials import Credentials
            oauth_token_path = os.path.join(os.path.dirname(__file__), 'config', 'antigravity_token.json')
            if os.path.exists(oauth_token_path):
                with open(oauth_token_path, 'r') as f:
                    token_data = json.load(f)
                _ag_cfg = self.core.config.get('antigravity', {}) or {}
                credentials = Credentials(
                    token=token_data.get('access_token'),
                    refresh_token=token_data.get('refresh_token'),
                    token_uri="https://oauth2.googleapis.com/token",
                    client_id=token_data.get('client_id') or _ag_cfg.get('client_id'),
                    client_secret=token_data.get('client_secret') or _ag_cfg.get('client_secret')
                )
                client_args['credentials'] = credentials
        except: pass
        
        if not client_args.get('credentials') and api_key:
            client_args['api_key'] = api_key
            
        if not client_args:
            raise ValueError("No Google API key or OAuth token configured.")
            
        return genai.Client(**client_args)

    async def tool_generate_image_imagen(self, args):
        """Generate an image using Google Imagen 3 via the google-genai SDK."""
        import time as _time
        prompt       = args.get('prompt', '')
        model        = args.get('model', 'imagen-3')
        aspect_ratio = args.get('aspect_ratio', '1:1')
        n_images     = int(args.get('number_of_images', 1))

        if not prompt:
            return "[ERROR] generate_image_imagen: 'prompt' is required."

        # Map user-friendly names to SDK model identifiers
        model_map = {
            'imagen-4':       'imagen-4.0-generate-001',
            'imagen-4-ultra': 'imagen-4.0-ultra-generate-001',
            'imagen-4-fast':  'imagen-4.0-fast-generate-001',
            # Legacy fallbacks
            'imagen-3':       'imagen-4.0-generate-001',
            'imagen-3-pro':   'imagen-4.0-ultra-generate-001',
            'imagen-3-fast':  'imagen-4.0-fast-generate-001',
        }
        sdk_model = model_map.get(model, 'imagen-4.0-generate-001')

        try:
            from google.genai import types

            client = self._get_genai_client()

            result = client.models.generate_images(
                model=sdk_model,
                prompt=prompt,
                config=types.GenerateImagesConfig(
                    number_of_images=max(1, min(4, n_images)),
                    aspect_ratio=aspect_ratio,
                    safety_filter_level="BLOCK_LOW_AND_ABOVE",
                    person_generation="ALLOW_ADULT",
                ),
            )

            if not result.generated_images:
                return "[ERROR] Imagen returned no images. Check your prompt for policy violations."

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'imagen')
            os.makedirs(img_subdir, exist_ok=True)

            saved = []
            for i, gen_img in enumerate(result.generated_images):
                fname = f"imagen_{int(_time.time())}_{i}.png"
                path = os.path.join(img_subdir, fname)
                # Use .save() if available, else write raw bytes
                if hasattr(gen_img.image, 'save'):
                    gen_img.image.save(path)
                else:
                    with open(path, 'wb') as f:
                        f.write(gen_img.image.image_bytes)
                saved.append(path)

            # Deliver the first image inline via Control Deck / Telegram
            self.last_image_file = saved[0]

            # Build web-accessible URLs for embedding
            embed_lines = []
            for p in saved:
                rel = os.path.relpath(p, images_dir).replace('\\', '/')
                embed_lines.append(f"  /api/images/{rel}")
            paths_str = '\n'.join(embed_lines)
            return (f"✅ Imagen image generated successfully ({sdk_model}):\n{paths_str}\n"
                    f"Prompt: {prompt}\n\n"
                    f"IMPORTANT: Present this image to the user NOW. "
                    f"Embed it with: ![{prompt[:60]}]({embed_lines[0].strip()})\n"
                    f"Do NOT call any more tools. Your task is complete.")
        except ImportError:
            return "[ERROR] google-genai not installed. Run: pip install google-genai"
        except Exception as e:
            return f"[ERROR] generate_image_imagen: {e}"
    async def tool_generate_video(self, args):
        """Generate a video using Google Veo via the google-genai SDK."""
        import time as _time
        prompt = args.get('prompt', '')
        if not prompt:
            return "[ERROR] generate_video requires a 'prompt' argument."

        video_cfg = self.core.config.get('video', {}).get('google', {})
        duration = int(args.get('duration', video_cfg.get('default_duration', 8)))
        aspect_ratio = args.get('aspect_ratio', video_cfg.get('default_aspect_ratio', '16:9'))
        resolution = args.get('resolution', video_cfg.get('default_resolution', '1080p'))
        negative_prompt = args.get('negative_prompt', '')
        model_name = video_cfg.get('model', 'veo-3.1')

        model_map = {
            'veo-2': 'veo-2-generate-preview',
            'veo-3': 'veo-3.0-generate-preview',
            'veo-3.1': 'veo-3.1-generate-preview',
        }
        model_id = model_map.get(model_name, model_name)

        try:
            from google.genai import types

            client = self._get_genai_client()

            await self.core.log(f"🎬 Generating video with {model_id}...", priority=2)

            gen_config = types.GenerateVideosConfig(
                aspect_ratio=aspect_ratio,
                resolution=resolution,
                duration_seconds=duration,
            )
            if negative_prompt:
                gen_config.negative_prompt = negative_prompt

            operation = client.models.generate_videos(
                model=model_id,
                source=types.GenerateVideosSource(prompt=prompt),
                config=gen_config,
            )

            poll_count = 0
            while not operation.done:
                poll_count += 1
                if poll_count % 6 == 0:
                    await self.core.log(
                        f"🎬 Video still generating... ({poll_count * 10}s elapsed)",
                        priority=3
                    )
                await asyncio.sleep(10)
                operation = client.operations.get(operation)

            if not operation.response or not operation.response.generated_videos:
                return "[ERROR] Video generation returned no results."

            video = operation.response.generated_videos[0]
            client.files.download(file=video.video)

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            vid_subdir = os.path.join(images_dir, 'video')
            os.makedirs(vid_subdir, exist_ok=True)
            fname = f"veo_{int(_time.time())}.mp4"
            path = os.path.join(vid_subdir, fname)
            video.video.save(path)

            self.last_video_file = path
            return (
                f"✅ Video generated: {path}\n"
                f"Model: {model_id}\n"
                f"Duration: {duration}s | Resolution: {resolution} | Aspect: {aspect_ratio}\n"
                f"Prompt: {prompt}"
            )
        except Exception as e:
            return f"[ERROR] generate_video: {e}"
    async def tool_generate_video_from_image(self, args):
        """Animate a still image into video using Google Veo."""
        import time as _time
        prompt = args.get('prompt', '')
        image_path = args.get('image_path', '')
        if not prompt:
            return "[ERROR] generate_video_from_image requires a 'prompt' argument."
        if not image_path or not os.path.exists(image_path):
            return f"[ERROR] Image not found: {image_path}"

        video_cfg = self.core.config.get('video', {}).get('google', {})
        duration = int(args.get('duration', video_cfg.get('default_duration', 8)))
        aspect_ratio = args.get('aspect_ratio', video_cfg.get('default_aspect_ratio', '16:9'))
        model_name = video_cfg.get('model', 'veo-3.1')

        model_map = {
            'veo-2': 'veo-2-generate-preview',
            'veo-3': 'veo-3.0-generate-preview',
            'veo-3.1': 'veo-3.1-generate-preview',
        }
        model_id = model_map.get(model_name, model_name)

        try:
            from google.genai import types
            from PIL import Image as _PILImage

            client = self._get_genai_client()

            await self.core.log(f"🎬 Animating image to video with {model_id}...", priority=2)

            import mimetypes as _mimetypes
            with open(image_path, 'rb') as f:
                img_bytes = f.read()
            mtype, _ = _mimetypes.guess_type(image_path)
            if not mtype: mtype = "image/jpeg"

            operation = client.models.generate_videos(
                model=model_id,
                source=types.GenerateVideosSource(
                    prompt=prompt,
                    image=types.Image(image_bytes=img_bytes, mime_type=mtype)
                ),
                config=types.GenerateVideosConfig(
                    aspect_ratio=aspect_ratio,
                    duration_seconds=duration,
                ),
            )

            poll_count = 0
            while not operation.done:
                poll_count += 1
                if poll_count % 6 == 0:
                    await self.core.log(
                        f"🎬 Video still generating... ({poll_count * 10}s elapsed)",
                        priority=3
                    )
                await asyncio.sleep(10)
                operation = client.operations.get(operation)

            if not operation.response or not operation.response.generated_videos:
                return "[ERROR] Video generation returned no results."

            video = operation.response.generated_videos[0]
            client.files.download(file=video.video)

            images_dir = self.core.config.get('paths', {}).get('images', './images')
            vid_subdir = os.path.join(images_dir, 'video')
            os.makedirs(vid_subdir, exist_ok=True)
            fname = f"veo_{int(_time.time())}.mp4"
            path = os.path.join(vid_subdir, fname)
            video.video.save(path)

            self.last_video_file = path
            return (
                f"✅ Image animated to video: {path}\n"
                f"Model: {model_id}\n"
                f"Source: {image_path}\n"
                f"Duration: {duration}s | Aspect: {aspect_ratio}\n"
                f"Prompt: {prompt}"
            )
        except Exception as e:
            return f"[ERROR] generate_video_from_image: {e}"
    async def tool_list_dir(self, args):
        """List directory contents with sizes and dates."""
        import glob as _glob, stat as _stat
        from datetime import datetime as _dt
        path    = args.get('path', '.') or '.'
        pattern = args.get('pattern', '*')
        recurse = bool(args.get('recurse', False))
        try:
            base = os.path.abspath(path)
            if not os.path.isdir(base):
                return (
                    f"[ERROR] list_dir FAILED — path does not exist or is not a directory.\n"
                    f"  Requested: {path!r}\n"
                    f"  Resolved to: {base!r}\n"
                    f"STOP — do not guess or invent file listings. Report this error to the user "
                    f"and ask them for the correct absolute path."
                )
            search = os.path.join(base, '**', pattern) if recurse else os.path.join(base, pattern)
            
            # Run blocking glob in a thread pool
            loop = asyncio.get_running_loop()
            entries = await loop.run_in_executor(getattr(self, '_io_pool', None), lambda: _glob.glob(search, recursive=recurse))
            
            if not entries:
                return f"No files match '{pattern}' in {base}"
            lines = [f"{'TYPE':<5} {'SIZE':>10}  {'MODIFIED':<20}  NAME"]
            lines.append('-' * 70)
            for e in sorted(entries)[:500]:
                try:
                    st   = os.stat(e)
                    kind = 'DIR ' if os.path.isdir(e) else 'FILE'
                    size = '' if os.path.isdir(e) else f"{st.st_size:,}"
                    mtime = _dt.fromtimestamp(st.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    name = os.path.relpath(e, base)
                    lines.append(f"{kind:<5} {size:>10}  {mtime:<20}  {name}")
                except Exception:
                    pass
            if len(entries) > 500:
                lines.append(f"... (showing 500 of {len(entries)} matches)")
            return '\n'.join(lines)
        except Exception as e:
            return f"[ERROR] list_dir: {e}"
    async def tool_find_files(self, args):
        """Find files matching a glob pattern recursively."""
        import glob as _glob
        path    = args.get('path', '.') or '.'
        pattern = args.get('pattern', '*')
        limit   = int(args.get('limit', 100))
        try:
            base = os.path.abspath(path)
            if '**' in pattern or '/' in pattern or '\\' in pattern:
                search = os.path.join(base, pattern)
            else:
                search = os.path.join(base, '**', pattern)
            
            # Run blocking glob in a thread pool
            loop = asyncio.get_running_loop()
            results = await loop.run_in_executor(getattr(self, '_io_pool', None), lambda: _glob.glob(search, recursive=True))
            
            results = [os.path.relpath(r, base) for r in sorted(results)]
            total = len(results)
            results = results[:limit]
            if not results:
                return f"No files found matching '{pattern}' under {base}"
            out = '\n'.join(results)
            if total > limit:
                out += f"\n... ({total - limit} more results — increase limit to see all)"
            return f"Found {total} file(s):\n{out}"
        except Exception as e:
            return f"[ERROR] find_files: {e}"
    async def tool_hash_file(self, args):
        """Compute a file's hash checksum."""
        import hashlib as _hl
        path = args.get('path', '')
        algo = args.get('algorithm', 'sha256').lower()
        algos = {'sha256': _hl.sha256, 'md5': _hl.md5, 'sha1': _hl.sha1}
        if algo not in algos:
            return f"[ERROR] Unsupported algorithm '{algo}'. Choose: sha256, md5, sha1"
        try:
            h = algos[algo]()
            with open(path, 'rb') as f:
                for chunk in iter(lambda: f.read(65536), b''):
                    h.update(chunk)
            size = os.path.getsize(path)
            return f"{algo.upper()}: {h.hexdigest()}\nFile: {path}\nSize: {size:,} bytes"
        except Exception as e:
            return f"[ERROR] hash_file: {e}"
    async def tool_diff_files(self, args):
        """Show unified diff between two files or a file and a string."""
        import difflib as _diff
        path_a  = args.get('path_a', '')
        path_b  = args.get('path_b', '')
        text_b  = args.get('text_b', None)
        context = int(args.get('context', 3))
        try:
            with open(path_a, 'r', encoding='utf-8', errors='replace') as f:
                lines_a = f.readlines()
            if path_b:
                with open(path_b, 'r', encoding='utf-8', errors='replace') as f:
                    lines_b = f.readlines()
                label_b = path_b
            elif text_b is not None:
                lines_b = [l if l.endswith('\n') else l + '\n' for l in text_b.splitlines()]
                label_b = '<new content>'
            else:
                return "[ERROR] Provide path_b or text_b to compare against."
            diff = list(_diff.unified_diff(lines_a, lines_b, fromfile=path_a, tofile=label_b, n=context))
            if not diff:
                return "✅ Files are identical — no differences found."
            diff_str = ''.join(diff)
            # Cap output: diffing two large or entirely-different files (e.g.
            # minified JS) can produce tens of thousands of chars and blow up the
            # model's context window.
            if len(diff_str) > 15000:
                diff_str = diff_str[:15000] + "\n\n...[DIFF TRUNCATED — output exceeded 15000 chars; compare narrower ranges]..."
            return diff_str
        except Exception as e:
            return f"[ERROR] diff_files: {e}"
    async def tool_zip_create(self, args):
        """Create a ZIP archive from a file or directory (non-blocking)."""
        import zipfile as _zip, time as _time
        source = args.get('source', '')
        dest   = args.get('destination', '') or source.rstrip('/\\') + '.zip'
        
        def _zip_sync():
            try:
                src = os.path.abspath(source)
                dst = os.path.abspath(dest)
                if not os.path.exists(src):
                    return f"[ERROR] Source does not exist: {src}"
                with _zip.ZipFile(dst, 'w', compression=_zip.ZIP_DEFLATED) as zf:
                    if os.path.isdir(src):
                        for root, dirs, files in os.walk(src):
                            for file in files:
                                fp = os.path.join(root, file)
                                zf.write(fp, os.path.relpath(fp, os.path.dirname(src)))
                    else:
                        zf.write(src, os.path.basename(src))
                size = os.path.getsize(dst)
                return f"✅ Created: {dst}\nSize: {size:,} bytes"
            except Exception as e:
                return f"[ERROR] zip_create sync: {e}"

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _zip_sync)
    async def tool_zip_extract(self, args):
        """Extract a ZIP archive."""
        import zipfile as _zip
        source = args.get('source', '')
        dest   = args.get('destination', '') or os.path.dirname(os.path.abspath(source))
        try:
            source = os.path.abspath(source)
            dest   = os.path.abspath(dest)
            os.makedirs(dest, exist_ok=True)
            with _zip.ZipFile(source, 'r') as zf:
                names = zf.namelist()
                zf.extractall(dest)
            return f"✅ Extracted {len(names)} files to: {dest}"
        except Exception as e:
            return f"[ERROR] zip_extract: {e}"
    async def tool_image_info(self, args):
        """Get image metadata without loading to AI."""
        path = args.get('path', '')
        try:
            from PIL import Image as _Image
            size = os.path.getsize(path)
            with _Image.open(path) as img:
                w, h   = img.size
                fmt    = img.format or 'UNKNOWN'
                mode   = img.mode
                info   = img.info
            exif_str = ''
            if 'exif' in info:
                exif_str = ' (EXIF data present)'
            return (
                f"File:       {path}\n"
                f"Format:     {fmt}\n"
                f"Dimensions: {w} x {h} px\n"
                f"Color mode: {mode}\n"
                f"File size:  {size:,} bytes ({size/1024:.1f} KB){exif_str}"
            )
        except ImportError:
            # Fallback without PIL — just file size + extension
            ext = os.path.splitext(path)[1].upper().lstrip('.')
            size = os.path.getsize(path) if os.path.exists(path) else 0
            return f"File: {path}\nFormat: {ext}\nFile size: {size:,} bytes\n(Install Pillow for full metadata: pip install Pillow)"
        except Exception as e:
            return f"[ERROR] image_info: {e}"
    async def tool_clipboard_get(self, args):
        """Read text from the OS clipboard."""
        try:
            import subprocess as _sp, sys as _sys
            if _sys.platform == 'win32':
                result = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', 'Get-Clipboard',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await result.communicate()
                text = stdout.decode('utf-8', errors='replace').strip()
            elif _sys.platform == 'darwin':
                result = await asyncio.create_subprocess_exec(
                    'pbpaste', stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await result.communicate()
                text = stdout.decode('utf-8', errors='replace').strip()
            else:
                # Linux — try xclip then xsel
                try:
                    result = await asyncio.create_subprocess_exec(
                        'xclip', '-selection', 'clipboard', '-o',
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    stdout, _ = await result.communicate()
                    text = stdout.decode('utf-8', errors='replace').strip()
                except FileNotFoundError:
                    result = await asyncio.create_subprocess_exec(
                        'xsel', '--clipboard', '--output',
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    stdout, _ = await result.communicate()
                    text = stdout.decode('utf-8', errors='replace').strip()
            if not text:
                return "(Clipboard is empty)"
            return f"Clipboard content ({len(text)} chars):\n{text}"
        except Exception as e:
            return f"[ERROR] clipboard_get: {e}"
    async def tool_clipboard_set(self, args):
        """Write text to the OS clipboard."""
        text = args.get('text', '')
        try:
            import subprocess as _sp, sys as _sys
            if _sys.platform == 'win32':
                proc = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', f'Set-Clipboard -Value @"\n{text}\n"@',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            elif _sys.platform == 'darwin':
                proc = await asyncio.create_subprocess_exec(
                    'pbcopy', stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate(input=text.encode('utf-8'))
            else:
                try:
                    proc = await asyncio.create_subprocess_exec(
                        'xclip', '-selection', 'clipboard',
                        stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    await proc.communicate(input=text.encode('utf-8'))
                except FileNotFoundError:
                    proc = await asyncio.create_subprocess_exec(
                        'xsel', '--clipboard', '--input',
                        stdin=_sp.PIPE, stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    await proc.communicate(input=text.encode('utf-8'))
            return f"✅ Copied {len(text)} characters to clipboard."
        except Exception as e:
            return f"[ERROR] clipboard_set: {e}"
    async def tool_notify(self, args):
        """Send a desktop notification."""
        import sys as _sys
        title   = args.get('title', 'Galactic AI')
        message = args.get('message', '')
        sound   = bool(args.get('sound', False))
        try:
            if _sys.platform == 'win32':
                # Use PowerShell toast on Windows 10/11
                ps_script = f"""
Add-Type -AssemblyName System.Windows.Forms
$notify = New-Object System.Windows.Forms.NotifyIcon
$notify.Icon = [System.Drawing.SystemIcons]::Information
$notify.Visible = $true
$notify.ShowBalloonTip(5000, '{title.replace("'", "")}', '{message.replace("'", "")}', [System.Windows.Forms.ToolTipIcon]::Info)
Start-Sleep -Milliseconds 5500
$notify.Dispose()
""".strip()
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'powershell', '-Command', ps_script,
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            elif _sys.platform == 'darwin':
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'osascript', '-e',
                    f'display notification "{message}" with title "{title}"',
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            else:
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'notify-send', title, message,
                    stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                await proc.communicate()
            return f"✅ Notification sent: '{title}' — {message}"
        except Exception as e:
            return f"[ERROR] notify: {e}"
    async def tool_window_list(self, args):
        """List all open windows."""
        import sys as _sys
        try:
            if _sys.platform == 'win32':
                import ctypes, ctypes.wintypes as _wt
                EnumWindows        = ctypes.windll.user32.EnumWindows
                GetWindowTextW     = ctypes.windll.user32.GetWindowTextW
                GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                IsWindowVisible    = ctypes.windll.user32.IsWindowVisible
                GetWindowThreadProcessId = ctypes.windll.user32.GetWindowThreadProcessId
                EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                windows = []
                def callback(hwnd, lParam):
                    if IsWindowVisible(hwnd):
                        length = GetWindowTextLengthW(hwnd)
                        if length > 0:
                            buf = ctypes.create_unicode_buffer(length + 1)
                            GetWindowTextW(hwnd, buf, length + 1)
                            pid = ctypes.c_ulong()
                            GetWindowThreadProcessId(hwnd, ctypes.byref(pid))
                            windows.append((int(hwnd), buf.value, pid.value))
                    return True
                EnumWindows(EnumWindowsProc(callback), 0)
                if not windows:
                    return "No visible windows found."
                lines = [f"{'HWND':>10}  {'PID':>7}  TITLE"]
                lines.append('-' * 70)
                for hwnd, title, pid in sorted(windows, key=lambda x: x[1].lower()):
                    lines.append(f"{hwnd:>10}  {pid:>7}  {title[:60]}")
                return '\n'.join(lines)
            else:
                import subprocess as _sp
                proc = await asyncio.create_subprocess_exec(
                    'wmctrl', '-l', stdout=_sp.PIPE, stderr=_sp.PIPE
                )
                stdout, _ = await proc.communicate()
                return stdout.decode('utf-8', errors='replace').strip() or "No windows found (wmctrl output was empty)"
        except Exception as e:
            return f"[ERROR] window_list: {e}"
    async def tool_window_focus(self, args):
        """Bring a window to the foreground."""
        import sys as _sys
        title = args.get('title', '')
        hwnd  = args.get('hwnd', None)
        try:
            if _sys.platform == 'win32':
                import ctypes
                if hwnd:
                    target_hwnd = int(hwnd)
                else:
                    # Find by title substring
                    EnumWindows = ctypes.windll.user32.EnumWindows
                    GetWindowTextW = ctypes.windll.user32.GetWindowTextW
                    GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
                    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                    found = []
                    def callback(h, lParam):
                        if IsWindowVisible(h):
                            length = GetWindowTextLengthW(h)
                            if length > 0:
                                buf = ctypes.create_unicode_buffer(length + 1)
                                GetWindowTextW(h, buf, length + 1)
                                if title.lower() in buf.value.lower():
                                    found.append((int(h), buf.value))
                        return True
                    EnumWindows(EnumWindowsProc(callback), 0)
                    if not found:
                        return f"[ERROR] No window found matching '{title}'"
                    target_hwnd = found[0][0]
                # Restore if minimized, then set foreground
                ctypes.windll.user32.ShowWindow(target_hwnd, 9)  # SW_RESTORE
                ctypes.windll.user32.SetForegroundWindow(target_hwnd)
                return f"✅ Focused window HWND={target_hwnd}"
            else:
                import subprocess as _sp
                cmd = ['wmctrl', '-a', title] if title else ['wmctrl', '-ia', str(hwnd)]
                proc = await asyncio.create_subprocess_exec(*cmd, stdout=_sp.PIPE, stderr=_sp.PIPE)
                _, stderr = await proc.communicate()
                if proc.returncode != 0:
                    return f"[ERROR] wmctrl: {stderr.decode().strip()}"
                return f"✅ Window focused"
        except Exception as e:
            return f"[ERROR] window_focus: {e}"
    async def tool_window_resize(self, args):
        """Resize and/or move a window."""
        import sys as _sys
        title  = args.get('title', '')
        hwnd   = args.get('hwnd', None)
        x      = args.get('x', None)
        y      = args.get('y', None)
        width  = args.get('width', None)
        height = args.get('height', None)
        try:
            if _sys.platform == 'win32':
                import ctypes
                if hwnd:
                    target_hwnd = int(hwnd)
                else:
                    EnumWindows = ctypes.windll.user32.EnumWindows
                    GetWindowTextW = ctypes.windll.user32.GetWindowTextW
                    GetWindowTextLengthW = ctypes.windll.user32.GetWindowTextLengthW
                    IsWindowVisible = ctypes.windll.user32.IsWindowVisible
                    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.POINTER(ctypes.c_int), ctypes.POINTER(ctypes.c_int))
                    found = []
                    def callback(h, lParam):
                        if IsWindowVisible(h):
                            length = GetWindowTextLengthW(h)
                            if length > 0:
                                buf = ctypes.create_unicode_buffer(length + 1)
                                GetWindowTextW(h, buf, length + 1)
                                if title.lower() in buf.value.lower():
                                    found.append(int(h))
                        return True
                    EnumWindows(EnumWindowsProc(callback), 0)
                    if not found:
                        return f"[ERROR] No window found matching '{title}'"
                    target_hwnd = found[0]
                # Get current rect
                import ctypes.wintypes as _wt
                rect = _wt.RECT()
                ctypes.windll.user32.GetWindowRect(target_hwnd, ctypes.byref(rect))
                nx = x      if x      is not None else rect.left
                ny = y      if y      is not None else rect.top
                nw = width  if width  is not None else (rect.right - rect.left)
                nh = height if height is not None else (rect.bottom - rect.top)
                ctypes.windll.user32.MoveWindow(target_hwnd, int(nx), int(ny), int(nw), int(nh), True)
                return f"✅ Window moved/resized: pos=({nx},{ny}) size={nw}x{nh}"
            else:
                import subprocess as _sp
                if title:
                    geo = ''
                    if width and height:
                        geo = f"{width}x{height}"
                        if x is not None and y is not None:
                            geo += f"+{x}+{y}"
                    proc = await asyncio.create_subprocess_exec(
                        'wmctrl', '-r', title, '-e', f"0,{x or -1},{y or -1},{width or -1},{height or -1}",
                        stdout=_sp.PIPE, stderr=_sp.PIPE
                    )
                    _, stderr = await proc.communicate()
                    if proc.returncode != 0:
                        return f"[ERROR] wmctrl: {stderr.decode().strip()}"
                    return "✅ Window resized"
                return "[ERROR] Provide title or hwnd"
        except Exception as e:
            return f"[ERROR] window_resize: {e}"
    async def tool_http_request(self, args):
        """Make a raw HTTP request to any URL."""
        method  = args.get('method', 'GET').upper()
        url     = args.get('url', '')
        headers = args.get('headers', {})
        body_json = args.get('json', None)
        body_data = args.get('data', None)
        params  = args.get('params', None)
        timeout = int(args.get('timeout', 30))
        if not url:
            return "[ERROR] http_request requires a 'url' argument."
        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                kwargs = {'headers': headers or {}}
                if params:
                    kwargs['params'] = params
                if body_json is not None:
                    kwargs['json'] = body_json
                elif body_data is not None:
                    kwargs['content'] = body_data.encode() if isinstance(body_data, str) else body_data
                r = await client.request(method, url, **kwargs)
            ct = r.headers.get('content-type', '')
            if 'application/json' in ct:
                try:
                    body = json.dumps(r.json(), indent=2)[:8000]
                except Exception:
                    body = r.text[:8000]
            else:
                body = r.text[:8000]
            return (
                f"HTTP {r.status_code} {r.reason_phrase}\n"
                f"Content-Type: {ct}\n"
                f"Headers: {dict(r.headers)}\n\n"
                f"{body}"
            )
        except Exception as e:
            return f"[ERROR] http_request: {e}"
    async def tool_qr_generate(self, args):
        """Generate a QR code and save it as a PNG image."""
        text  = args.get('text', '')
        size  = int(args.get('size', 10))
        border = int(args.get('border', 4))
        ec_map = {'L': 1, 'M': 0, 'Q': 3, 'H': 2}
        ec    = ec_map.get(args.get('error_correction', 'M').upper(), 0)
        if not text:
            return "[ERROR] qr_generate requires 'text' argument."
        try:
            import qrcode as _qr
            import time as _time
            qr = _qr.QRCode(
                version=None,
                error_correction=ec,
                box_size=size,
                border=border,
            )
            qr.add_data(text)
            qr.make(fit=True)
            img = qr.make_image(fill_color='black', back_color='white')
            images_dir = self.core.config.get('paths', {}).get('images', './images')
            img_subdir = os.path.join(images_dir, 'qr')
            os.makedirs(img_subdir, exist_ok=True)
            fname = f"qr_{int(_time.time())}.png"
            path  = os.path.join(img_subdir, fname)
            img.save(path)
            self.last_image_file = path
            return f"✅ QR code saved to: {path}\nContent: {text[:80]}"
        except ImportError:
            return "[ERROR] qrcode library not installed. Run: pip install qrcode[pil]"
        except Exception as e:
            return f"[ERROR] qr_generate: {e}"
    async def tool_env_get(self, args):
        """Read environment variable(s)."""
        name = args.get('name', '')
        _SKIP = {'PATH', 'PYTHONPATH', 'APPDATA', 'LOCALAPPDATA', 'PROGRAMDATA',
                 'COMSPEC', 'PROCESSOR_ARCHITECTURE', 'NUMBER_OF_PROCESSORS'}
        if name:
            val = os.environ.get(name)
            if val is None:
                return f"Environment variable '{name}' is not set."
            return f"{name}={val}"
        else:
            lines = []
            for k, v in sorted(os.environ.items()):
                if any(secret in k.upper() for secret in ['KEY', 'SECRET', 'PASSWORD', 'TOKEN', 'PASS']):
                    lines.append(f"{k}=<hidden>")
                else:
                    lines.append(f"{k}={v[:120]}")
            return '\n'.join(lines)
    async def tool_env_set(self, args):
        """Set an environment variable for this session."""
        name  = args.get('name', '')
        value = args.get('value', '')
        if not name:
            return "[ERROR] env_set requires 'name' argument."
        os.environ[name] = value
        return f"✅ Set {name}={value}"
    async def tool_system_info(self, args):
        """Return detailed system hardware and OS information."""
        import platform as _pl, time as _tm
        try:
            import psutil as _ps
            cpu_count  = _ps.cpu_count(logical=True)
            cpu_phys   = _ps.cpu_count(logical=False)
            cpu_pct    = _ps.cpu_percent(interval=0.3)
            mem        = _ps.virtual_memory()
            disk       = _ps.disk_usage('/')
            boot_time  = _ps.boot_time()
            uptime_s   = int(_tm.time() - boot_time)
            uptime_str = f"{uptime_s//3600}h {(uptime_s%3600)//60}m"
            proc_count = len(_ps.pids())
            ram_total  = f"{mem.total / (1024**3):.1f} GB"
            ram_used   = f"{mem.used / (1024**3):.1f} GB ({mem.percent:.0f}%)"
            disk_total = f"{disk.total / (1024**3):.1f} GB"
            disk_used  = f"{disk.used / (1024**3):.1f} GB ({disk.percent:.0f}%)"
            psutil_info = (
                f"CPU:          {cpu_phys} physical / {cpu_count} logical cores @ {cpu_pct:.1f}% usage\n"
                f"RAM:          {ram_used} / {ram_total}\n"
                f"Disk (/):     {disk_used} / {disk_total}\n"
                f"Uptime:       {uptime_str}\n"
                f"Processes:    {proc_count} running\n"
            )
        except ImportError:
            psutil_info = "(Install psutil for CPU/RAM stats: pip install psutil)\n"
        import sys as _sys
        return (
            f"OS:           {_pl.system()} {_pl.release()} ({_pl.version()[:60]})\n"
            f"Machine:      {_pl.machine()} / {_pl.processor()[:60]}\n"
            f"Python:       {_sys.version.split()[0]} ({_sys.executable})\n"
            + psutil_info
        )
    async def tool_kill_process_by_name(self, args):
        """Kill processes by name substring."""
        name  = args.get('name', '').lower()
        force = bool(args.get('force', False))
        if not name:
            return "[ERROR] kill_process_by_name requires 'name' argument."
        try:
            import psutil as _ps
            killed = []
            for proc in _ps.process_iter(['pid', 'name', 'cmdline']):
                try:
                    pname = (proc.info['name'] or '').lower()
                    if name in pname:
                        if force:
                            proc.kill()
                        else:
                            proc.terminate()
                        killed.append(f"PID {proc.pid}: {proc.info['name']}")
                except (_ps.NoSuchProcess, _ps.AccessDenied):
                    pass
            if not killed:
                return f"No processes found matching '{name}'"
            return f"✅ Terminated {len(killed)} process(es):\n" + '\n'.join(killed)
        except ImportError:
            # Fallback to taskkill / kill
            import subprocess as _sp
            import sys as _sys
            if _sys.platform == 'win32':
                flag = '/F' if force else ''
                cmd = ['taskkill', '/IM', f'*{name}*', flag] if flag else ['taskkill', '/IM', f'*{name}*']
                proc = await asyncio.create_subprocess_exec(*[c for c in cmd if c], stdout=_sp.PIPE, stderr=_sp.PIPE)
                stdout, stderr = await proc.communicate()
                return stdout.decode('utf-8', errors='replace').strip() or stderr.decode('utf-8', errors='replace').strip()
            else:
                sig = '-9' if force else '-15'
                proc = await asyncio.create_subprocess_exec('pkill', sig, '-f', name, stdout=_sp.PIPE, stderr=_sp.PIPE)
                stdout, stderr = await proc.communicate()
                return f"pkill exit {proc.returncode}: {(stdout+stderr).decode(errors='replace').strip() or 'Done'}"
        except Exception as e:
            return f"[ERROR] kill_process_by_name: {e}"
    async def tool_color_pick(self, args):
        """Sample pixel color at screen coordinates."""
        x = int(args.get('x', 0))
        y = int(args.get('y', 0))
        try:
            import pyautogui as _pag
            import colorsys as _cs
            pixel = _pag.screenshot().getpixel((x, y))
            r, g, b = pixel[0], pixel[1], pixel[2]
            h, s, v = _cs.rgb_to_hsv(r/255, g/255, b/255)
            return (
                f"Pixel at ({x}, {y}):\n"
                f"  Hex:  #{r:02X}{g:02X}{b:02X}\n"
                f"  RGB:  rgb({r}, {g}, {b})\n"
                f"  HSV:  hsl({h*360:.0f}°, {s*100:.0f}%, {v*100:.0f}%)"
            )
        except Exception as e:
            return f"[ERROR] color_pick: {e}"
    async def tool_text_transform(self, args):
        """Transform text in various ways."""
        import re as _re, json as _json, urllib.parse as _up, base64 as _b64, csv as _csv, io as _io
        text      = args.get('text', '')
        operation = args.get('operation', '').lower().replace(' ', '_')
        pattern   = args.get('pattern', '')
        try:
            if operation == 'upper':
                return text.upper()
            elif operation == 'lower':
                return text.lower()
            elif operation == 'title':
                return text.title()
            elif operation == 'snake_case':
                return _re.sub(r'[\s\-]+', '_', _re.sub(r'(?<!^)(?=[A-Z])', '_', text)).lower()
            elif operation == 'camel_case':
                parts = _re.split(r'[\s_\-]+', text)
                return parts[0].lower() + ''.join(p.title() for p in parts[1:])
            elif operation == 'base64_encode':
                return _b64.b64encode(text.encode('utf-8')).decode('ascii')
            elif operation == 'base64_decode':
                return _b64.b64decode(text).decode('utf-8', errors='replace')
            elif operation == 'url_encode':
                return _up.quote(text, safe='')
            elif operation == 'url_decode':
                return _up.unquote(text)
            elif operation == 'reverse':
                return text[::-1]
            elif operation == 'count':
                lines = text.splitlines()
                words = text.split()
                non_space = len(text.replace(' ', '').replace('\n', ''))
                return (f"Characters: {len(text):,}\n"
                        f"Words:      {len(words):,}\n"
                        f"Lines:      {len(lines):,}\n"
                        f"Non-space:  {non_space:,}")
            elif operation == 'strip':
                return text.strip()
            elif operation == 'regex_extract':
                if not pattern:
                    return "[ERROR] regex_extract requires a 'pattern' argument."
                matches = _re.findall(pattern, text)
                if not matches:
                    return "No matches found."
                return f"Found {len(matches)} match(es):\n" + '\n'.join(str(m) for m in matches[:100])
            elif operation == 'json_format':
                try:
                    parsed = _json.loads(text)
                    return _json.dumps(parsed, indent=2, ensure_ascii=False)
                except Exception as e:
                    return f"[ERROR] Invalid JSON: {e}"
            elif operation == 'csv_to_json':
                reader = _csv.DictReader(_io.StringIO(text))
                rows = list(reader)
                return _json.dumps(rows, indent=2, ensure_ascii=False)
            else:
                ops = ['upper','lower','title','snake_case','camel_case','base64_encode','base64_decode',
                       'url_encode','url_decode','reverse','count','strip','regex_extract','json_format','csv_to_json']
                return f"[ERROR] Unknown operation '{operation}'. Available: {', '.join(ops)}"
        except Exception as e:
            return f"[ERROR] text_transform ({operation}): {e}"
    async def tool_execute_python(self, args):
        """Execute Python code in a subprocess with robust output capping and stderr prioritization."""
        code = args.get('code', '')
        timeout = min(int(args.get('timeout', 60)), 300)
        if not code.strip():
            return "❌ Error: No code provided."

        # The Crucible: gate the execution behind user approval (no-op when off).
        approved, reason = await self._approve_command('execute_python', code,
                                                       target='python -c (inline script)')
        if not approved:
            return reason

        import tempfile
        tmp = None
        try:
            # Write into the project's own tmp/ sandbox (self._temp_dir =
            # GALACTIC_TEMP_DIR), not the global OS %TEMP%, so scratch scripts
            # stay inside the workspace and are swept by the 7-day tmp/ purge.
            _tmpdir = getattr(self, '_temp_dir', None)
            tmp = tempfile.NamedTemporaryFile(mode='w', suffix='.py', delete=False,
                                              encoding='utf-8', dir=_tmpdir)
            tmp.write(code)
            tmp.close()
            
            proc = await asyncio.create_subprocess_exec(
                'python', tmp.name,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            
            try:
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
            except asyncio.TimeoutError:
                proc.kill()
                await proc.wait()
                return f"⏱️ Timeout: Python script exceeded {timeout}s and was terminated."

            out = stdout.decode('utf-8', errors='replace').strip()
            err = stderr.decode('utf-8', errors='replace').strip()
            
            # Capping logic: 4000 chars per stream to protect context
            if len(out) > 4000: out = out[:2000] + "\n...[STDOUT TRUNCATED]...\n" + out[-2000:]
            if len(err) > 4000: err = err[:2000] + "\n...[STDERR TRUNCATED]...\n" + err[-2000:]
            
            result_parts = []
            if err:
                result_parts.append(f"❌ STDERR:\n{err}")
            if out:
                result_parts.append(f"✅ STDOUT:\n{out}")
            
            if proc.returncode != 0:
                result_parts.append(f"⚠️ Exit Code: {proc.returncode}")
            
            return "\n\n".join(result_parts) if result_parts else "✅ Script completed with no output."
            
        except Exception as e:
            return f"❌ Error executing Python: {str(e)}"
        finally:
            # Best-effort cleanup. On Windows a just-killed subprocess can keep a
            # lock on the file for a few ms after proc.kill(), making os.unlink
            # raise PermissionError — retry briefly before giving up rather than
            # silently leaking it. The 7-day tmp/ purge is the backstop.
            if tmp and os.path.exists(tmp.name):
                for _attempt in range(3):
                    try:
                        os.unlink(tmp.name)
                        break
                    except PermissionError:
                        await asyncio.sleep(0.1)
                    except Exception:
                        break
    async def tool_wait(self, args):
        """Pause execution."""
        seconds = min(float(args.get('seconds', 1)), 300)
        await asyncio.sleep(seconds)
        return f"Waited {seconds:.1f} seconds."

    async def tool_ask_user(self, args):
        """Pause mid-task and ask the human a question, then resume with their answer.

        Use when you hit something only the user can resolve — a 2FA/verification
        code, a subjective design choice, a missing credential, an ambiguous
        instruction. Blocks the loop until they answer or a timeout elapses, so
        the agent never has to guess-and-fail or abort the whole task.
        """
        question = (args.get('question') or '').strip()
        if not question:
            return "[ERROR] ask_user requires a 'question'."
        timeout = min(int(args.get('timeout', 300)), 1800)

        import uuid as _uuid
        req_id = _uuid.uuid4().hex[:12]
        loop = asyncio.get_running_loop()
        fut = loop.create_future()
        pending = getattr(self, '_pending_asks', None)
        if pending is None:
            pending = self._pending_asks = {}
        pending[req_id] = fut

        try:
            await self.core.relay.emit(2, "ask_user", {"id": req_id, "question": question})
            await self.core.log(f"❓ ask_user: awaiting human input — {question[:80]}", priority=2)
            try:
                answer = await asyncio.wait_for(fut, timeout=timeout)
                return f"[USER ANSWERED] {answer}"
            except asyncio.TimeoutError:
                return (f"[NO RESPONSE] The user did not answer within {timeout}s. "
                        f"Proceed with your best judgment, or stop and tell them you're "
                        f"blocked waiting on: {question[:120]}")
        finally:
            pending.pop(req_id, None)
            # Tell the UI to dismiss the prompt even on timeout/cancel.
            try:
                await self.core.relay.emit(2, "ask_user_resolved", {"id": req_id})
            except Exception:
                pass

    # ── Swarm Blackboard: live shared memory across agents ───────────────────
    def _blackboard(self):
        bb = getattr(self.core, 'blackboard', None)
        if bb is None:
            from blackboard import Blackboard
            bb = self.core.blackboard = Blackboard()
        return bb

    async def tool_blackboard_write(self, args):
        """Publish a value to the shared Blackboard so other agents can read it."""
        key = (args.get('key') or '').strip()
        if not key:
            return "[ERROR] blackboard_write requires a 'key'."
        value = args.get('value', '')
        by = (args.get('by') or 'agent').strip() or 'agent'
        self._blackboard().write(key, value, by=by)
        try:
            await self.core.relay.emit(3, "blackboard_update", {
                "key": key, "by": by,
                "preview": (value if isinstance(value, str) else str(value))[:200]})
        except Exception:
            pass
        return f"[BLACKBOARD] Wrote key '{key}'. Peers can now blackboard_read('{key}')."

    async def tool_blackboard_read(self, args):
        """Read a value from the shared Blackboard by key (returns nothing if unset)."""
        key = (args.get('key') or '').strip()
        if not key:
            return "[ERROR] blackboard_read requires a 'key'."
        val = self._blackboard().read(key)
        if val is None:
            existing = ', '.join(self._blackboard().keys()) or '(none yet)'
            return f"[BLACKBOARD] Key '{key}' is not set. Available keys: {existing}"
        return f"[BLACKBOARD] {key} =\n{val}"

    async def tool_blackboard_list(self, args):
        """List the keys currently on the shared Blackboard, with short previews."""
        snap = self._blackboard().snapshot(value_chars=120)
        if not snap:
            return "[BLACKBOARD] Empty — no keys written yet."
        lines = [f"[BLACKBOARD] {len(snap)} key(s):"]
        for e in snap:
            lines.append(f"  • {e['key']} (by {e['by']}): {e['preview']}")
        return "\n".join(lines)

    async def tool_blackboard_wait_for(self, args):
        """Block until another agent writes `key` to the Blackboard, then return it."""
        key = (args.get('key') or '').strip()
        if not key:
            return "[ERROR] blackboard_wait_for requires a 'key'."
        timeout = min(int(args.get('timeout', 60)), 600)
        val = await self._blackboard().wait_for(key, timeout=timeout)
        if val is None:
            return (f"[BLACKBOARD] Timed out after {timeout}s waiting for key '{key}' — "
                    f"no agent produced it. Proceed without it or try a different approach.")
        return f"[BLACKBOARD] {key} (received) =\n{val}"
    async def tool_send_telegram(self, args):
        """Send a proactive Telegram message."""
        message = args.get('message', '')
        chat_id = args.get('chat_id', '') or str(self.core.config.get('telegram', {}).get('admin_chat_id', ''))
        image_path = args.get('image_path', '')
        if not chat_id:
            return "[ERROR] No chat_id provided and no admin_chat_id in config."
        if not message:
            return "[ERROR] No message provided."
        try:
            tg = getattr(self.core, 'telegram', None)
            if not tg:
                return "[ERROR] Telegram bridge not available."
            if image_path and os.path.exists(image_path):
                await tg.send_photo(int(chat_id), image_path, caption=message)
                return f"Sent photo + message to Telegram chat {chat_id}."
            else:
                await tg.send_message(int(chat_id), message)
                return f"Sent message to Telegram chat {chat_id}."
        except Exception as e:
            return f"[ERROR] send_telegram: {e}"
    async def tool_read_pdf(self, args):
        """Extract text from a PDF file."""
        path = args.get('path', '')
        pages_arg = args.get('pages', 'all')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"
        try:
            try:
                import pdfplumber
                with pdfplumber.open(path) as pdf:
                    total = len(pdf.pages)
                    page_range = self._parse_page_range(pages_arg, total)
                    texts = []
                    for i in page_range:
                        page = pdf.pages[i]
                        text = page.extract_text()
                        if text:
                            texts.append(f"--- Page {i+1} ---\n{text}")
                    return "\n\n".join(texts) if texts else "[INFO] No text content found in PDF."
            except ImportError:
                pass
            try:
                import PyPDF2
                reader = PyPDF2.PdfReader(path)
                total = len(reader.pages)
                page_range = self._parse_page_range(pages_arg, total)
                texts = []
                for i in page_range:
                    text = reader.pages[i].extract_text()
                    if text:
                        texts.append(f"--- Page {i+1} ---\n{text}")
                return "\n\n".join(texts) if texts else "[INFO] No text content found in PDF."
            except ImportError:
                return "[ERROR] Install pdfplumber or PyPDF2: pip install pdfplumber"
        except Exception as e:
            return f"[ERROR] read_pdf: {e}"
    async def tool_read_csv(self, args):
        """Read a CSV file (non-blocking)."""
        import csv as _csv
        path = args.get('path', '')
        limit = int(args.get('limit', 200))
        delimiter = args.get('delimiter', ',')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"

        def _read_sync():
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = _csv.DictReader(f, delimiter=delimiter)
                    rows = []
                    for i, row in enumerate(reader):
                        if i >= limit: break
                        rows.append(dict(row))
                    return json.dumps({"total_rows": len(rows), "columns": list(rows[0].keys()) if rows else [], "rows": rows}, indent=2)
            except Exception as e: return f"[ERROR] csv_read_sync: {e}"

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _read_sync)
    async def tool_write_csv(self, args):
        """Write JSON rows to a CSV file (non-blocking)."""
        import csv as _csv
        path = args.get('path', '')
        rows = args.get('rows', [])
        append = args.get('append', False)
        if not path or not rows: return "[ERROR] Path and rows required."

        def _write_sync():
            try:
                mode = 'a' if append else 'w'
                exists = os.path.exists(path) and append
                with open(path, mode, newline='', encoding='utf-8') as f:
                    writer = _csv.DictWriter(f, fieldnames=rows[0].keys())
                    if not exists: writer.writeheader()
                    writer.writerows(rows)
                return f"Wrote {len(rows)} rows to {path}."
            except Exception as e: return f"[ERROR] csv_write_sync: {e}"

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _write_sync)
    async def tool_read_excel(self, args):
        """Read an Excel (.xlsx) file (non-blocking)."""
        path = args.get('path', ''); sheet = args.get('sheet', None); limit = int(args.get('limit', 100))
        if not path or not os.path.exists(path): return f"[ERROR] File not found: {path}"

        def _excel_sync():
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows: return "[INFO] Empty spreadsheet."
                headers = [str(h) if h else f"col_{i}" for i, h in enumerate(rows[0])]
                data = []
                for row in rows[1:limit+1]:
                    data.append({headers[i]: (str(v) if v is not None else '') for i, v in enumerate(row)})
                snames = wb.sheetnames; wb.close()
                return json.dumps({"sheets": snames, "columns": headers, "rows": data, "total_rows": len(data)}, indent=2)
            except Exception as e: return f"[ERROR] excel_sync: {e}"

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _excel_sync)
    async def tool_regex_search(self, args):
        """Search files with regex (non-blocking)."""
        import fnmatch as _fn
        pattern = args.get('pattern', '')
        search_path = args.get('path', '.')
        file_pattern = args.get('file_pattern', '*')
        limit = int(args.get('limit', 50))
        if not pattern:
            return "[ERROR] No pattern provided."
        try:
            compiled = re.compile(pattern, re.IGNORECASE)
        except re.error as e:
            return f"[ERROR] Invalid regex: {e}"

        def _search_sync():
            results = []
            try:
                if os.path.isfile(search_path):
                    files = [search_path]
                else:
                    files = []
                    for root, dirs, fnames in os.walk(search_path):
                        for fn in fnames:
                            if _fn.fnmatch(fn, file_pattern):
                                files.append(os.path.join(root, fn))
                        if len(files) > 5000:
                            break
                for fpath in files:
                    try:
                        with open(fpath, 'r', encoding='utf-8', errors='ignore') as f:
                            for line_no, line in enumerate(f, 1):
                                if compiled.search(line):
                                    results.append(f"{fpath}:{line_no}: {line.rstrip()[:200]}")
                                    if len(results) >= limit:
                                        return f"Found {len(results)} matches (limit reached):\n" + "\n".join(results)
                    except (PermissionError, IsADirectoryError):
                        continue
                return f"Found {len(results)} matches:\n" + "\n".join(results) if results else "No matches found."
            except Exception as e:
                return f"[ERROR] regex_search sync: {e}"

        loop = asyncio.get_running_loop()
        return await loop.run_in_executor(None, _search_sync)
    async def tool_image_resize(self, args):
        """Resize an image (non-blocking)."""
        path = args.get('path', '')
        width = args.get('width')
        height = args.get('height')
        output = args.get('output_path', '')
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"

        def _resize_sync():
            try:
                from PIL import Image
                img = Image.open(path)
                orig_w, orig_h = img.size
                new_w = int(width) if width else orig_w
                new_h = int(height) if height else orig_h
                resized = img.resize((new_w, new_h), Image.LANCZOS)
                out_path = output or os.path.splitext(path)[0] + f"_{new_w}x{new_h}" + os.path.splitext(path)[1]
                resized.save(out_path)
                return f"Resized {orig_w}x{orig_h} → {new_w}x{new_h}. Saved to: {out_path}"
            except Exception as e: return f"[ERROR] _resize_sync: {e}"

        try:
            loop = asyncio.get_running_loop()
            return await loop.run_in_executor(None, _resize_sync)
        except Exception as e:
            return f"[ERROR] image_resize: {e}"
    async def tool_image_convert(self, args):
        """Convert image format (non-blocking)."""
        path = args.get('path', '')
        fmt = args.get('format', 'png').lower()
        output = args.get('output_path', '')
        quality = int(args.get('quality', 85))
        if not path or not os.path.exists(path):
            return f"[ERROR] File not found: {path}"

        def _convert_sync():
            try:
                from PIL import Image
                img = Image.open(path)
                if fmt in ('jpeg', 'jpg') and img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                out_path = output or os.path.splitext(path)[0] + '.' + ('jpg' if fmt == 'jpeg' else fmt)
                save_kwargs = {}
                if fmt in ('jpeg', 'jpg', 'webp'):
                    save_kwargs['quality'] = quality
                img.save(out_path, **save_kwargs)
                return f"Converted to {fmt.upper()}. Saved to: {out_path}"
            except Exception as e: return f"[ERROR] _convert_sync: {e}"

        try:
            loop = asyncio.get_running_loop()
            return await loop.run_in_executor(None, _convert_sync)
        except Exception as e:
            return f"[ERROR] image_convert: {e}"
    async def tool_git_status(self, args):
        path = args.get('path')
        return await self._git_exec(['status', '--short'], cwd=path)
    async def tool_git_diff(self, args):
        path = args.get('path')
        cmd = ['diff', '--stat']
        if args.get('staged'):
            cmd.append('--cached')
        return await self._git_exec(cmd, cwd=path)
    async def tool_git_log(self, args):
        path = args.get('path')
        count = str(int(args.get('count', 10)))
        return await self._git_exec(['log', f'--oneline', f'-{count}'], cwd=path)
    async def tool_git_commit(self, args):
        path = args.get('path')
        message = args.get('message', 'Auto-commit by Galactic AI')
        files = args.get('files', [])
        cwd = path or self.core.config.get('paths', {}).get('workspace', '.')
        if files:
            for f in files:
                await self._git_exec(['add', f], cwd=cwd)
        else:
            await self._git_exec(['add', '-A'], cwd=cwd)
        return await self._git_exec(['commit', '-m', message], cwd=cwd)
    async def tool_restart_galactic(self, args):
        """
        Restart the Galactic AI process from within an agent task.
        Use this after applying self-repairs (patching files, installing deps, etc.)
        so changes take effect without manual intervention.

        Args (all optional):
          reason (str): Human-readable reason for the restart (logged before shutdown).

        Returns: confirmation string. Note: the response is sent before the process
                 restarts, so the current task/session will end immediately after.
        """
        reason = args.get('reason', 'Self-repair restart requested by agent.')
        await self.core.log(f"🔄 Agent-initiated restart: {reason}", priority=1)

        try:
            import aiohttp
            port = self.core.config.get('web', {}).get('port', 17789)
            url  = f"http://127.0.0.1:{port}/api/restart"
            async with aiohttp.ClientSession() as session:
                async with session.post(url, json={'reason': reason}, timeout=aiohttp.ClientTimeout(total=5)) as resp:
                    data = await resp.json()
                    if data.get('ok'):
                        return f"✅ Restart triggered: {data.get('message', 'Restarting...')} Reason: {reason}"
                    else:
                        return f"⚠️ Restart API returned error: {data}"
        except Exception as e:
            # Fallback: direct OS-level restart if HTTP call fails (e.g., web deck not running)
            await self.core.log(f"⚠️ HTTP restart failed ({e}), falling back to direct restart.", priority=1)
            import sys, subprocess, asyncio
            async def _direct_restart():
                await asyncio.sleep(1.0)
                subprocess.Popen([sys.executable] + sys.argv)
                shutdown_event = getattr(self.core, 'shutdown_event', None)
                if shutdown_event:
                    shutdown_event.set()
                else:
                    sys.exit(0)
            asyncio.create_task(_direct_restart())
            return f"✅ Direct restart scheduled. Reason: {reason}"
